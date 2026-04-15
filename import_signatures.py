import sqlite3
import openpyxl
import glob
import sys
import os

# DB 연결
db_path = "./backend/data/signatures.db"
conn = sqlite3.connect(db_path)
cur = conn.cursor()

# 엑셀 파일 찾기
target = None
if len(sys.argv) > 1:
    target = sys.argv[1]
else:
    files = glob.glob("./spams/*20260413_A.xlsx")
    for f in files:
        if "SD Output" not in f:
            target = f
            break

if not target or not os.path.exists(target):
    print("엑셀 파일을 찾을 수 없습니다.")
    sys.exit(1)

print(f"Reading from: {os.path.basename(target)}")
wb = openpyxl.load_workbook(target, data_only=True)

sheets_to_read = ["문자열중복제거", "문장중복제거"]
totals_inserted = 0
totals_ignored = 0

for s in sheets_to_read:
    if s not in wb.sheetnames:
        continue
    ws = wb[s]
    inserted_in_sheet = 0
    ignored_in_sheet = 0
    
    for row in range(2, ws.max_row + 1):
        signature = ws.cell(row=row, column=1).value
        byte_length = ws.cell(row=row, column=2).value
        category_code = ws.cell(row=row, column=3).value
        
        if signature:
            # 카테고리 로직 (기본 spam, 필요에 따라 맵핑 가능하지만 현재 구조에서는 카테고리 코드로 저장하거나 기본값 사용)
            cat_str = str(category_code) if category_code else 'spam'
            # 현재 코드 상 source는 파일명에서 동적으로 추출
            src_str = f"excel_import_{os.path.basename(target).replace('.xlsx', '')}"
            
            try:
                # signature 컬럼이 UNIQUE 제약 조건이 있으므로 IGNORE 사용
                cur.execute('''
                    INSERT OR IGNORE INTO signatures 
                    (signature, byte_length, category, source) 
                    VALUES (?, ?, ?, ?)
                ''', (str(signature), byte_length, cat_str, src_str))
                
                if cur.rowcount > 0:
                    inserted_in_sheet += 1
                    totals_inserted += 1
                else:
                    ignored_in_sheet += 1
                    totals_ignored += 1
            except Exception as e:
                print(f"Error inserting {signature}: {e}")
                
    print(f"[{s}] 처리 완료: {inserted_in_sheet}건 삽입, {ignored_in_sheet}건 중복 무시")

conn.commit()
conn.close()

print(f"\n최종 완료! 총 {totals_inserted}건 새로 삽입, {totals_ignored}건 중복 무시되었습니다.")
