import sys
import collections
from openpyxl import load_workbook

file_path = sys.argv[1]
try:
    wb = load_workbook(file_path, data_only=True)
    # The user mentioned '육안분석(시뮬결과35_150)'. We will try to find it.
    sheet_name = '육안분석(시뮬결과35_150)'
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Fallback to the second sheet if name differs slightly
        ws = wb.worksheets[1]
    
    headers = [str(cell.value).strip() if cell.value else f'Col{i}' for i, cell in enumerate(ws[1])]
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not row[0].value: break  
        
        # Determine color of first cell
        fill = row[0].fill
        color = 'None'
        if fill and fill.start_color:
            if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                color = str(fill.start_color.rgb)
            elif hasattr(fill.start_color, 'index'):
                color = str(fill.start_color.index)
            elif hasattr(fill.start_color, 'theme'):
                color = f"theme-{fill.start_color.theme}"
                
        row_data = {'_row': row_idx, '_color': color}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value
        data.append(row_data)

    groups = collections.defaultdict(list)
    for d in data:
        groups[d['_color']].append(d)
        
    for color, items in groups.items():
        print(f'\n=== Color Group: {color} (Count: {len(items)}) ===')
        for i in range(min(50, len(items))): 
            item = items[i]
            code = item.get('분류', '')
            gubun = item.get('구분', '')
            prob = item.get('Probability', '')
            msg_len = item.get('메시지 길이', '')
            url_len = item.get('URL 길이', '')
            msg_head = str(item.get('메시지', ''))[:15].replace('\n', ' ')
            print(f"Row {item['_row']:04d} | 구분:{gubun} | 분류:{code} | 확률:{prob} | ms_len:{msg_len} | ur_len:{url_len} | {msg_head}")
            
except Exception as e:
    import traceback
    traceback.print_exc()
