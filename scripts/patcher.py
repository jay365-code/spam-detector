# Apply patching logic safely
import sys

TARGET = r"c:\Users\leejo\Project\AI Agent\Spam Detector\backend\app\utils\excel_handler.py"
with open(TARGET, "r", encoding="utf-8") as f:
    text = f.read()

# 1. Update imports
text = text.replace(
    "from openpyxl.styles import Font, Alignment, PatternFill",
    "from openpyxl.styles import Font, Alignment, PatternFill, Border, Side"
)

# 2. update create_template_workbook (remove 금융.SPAM headers)
text = text.replace(
    '        for name in ["육안분석(시뮬결과35_150)", "TRAP.육안분석(시뮬결과35_150)", "금융.SPAM"]:\n            ws_target = wb[name]',
    '        for name in ["육안분석(시뮬결과35_150)", "TRAP.육안분석(시뮬결과35_150)"]:\n            ws_target = wb[name]'
)

# 3. update process_file
old_process_file = '''                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        if finance_ws.max_row <= 1 and not finance_ws.cell(row=1, column=1).value:
                            finance_headers = ["메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이", "Probability", "Semantic Class", "Reason", "Red Group"]
                            finance_ws.append(finance_headers)
                            header_font = Font(bold=True)
                            header_align = Alignment(horizontal='center', vertical='center')
                            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            for cell in finance_ws[1]:
                                cell.font = header_font
                                cell.alignment = header_align
                                cell.fill = header_fill
                        
                        _msg = ws.cell(row=row_idx, column=msg_col_idx).value
                        _url = ""
                        try:
                            _url_col_idx = headers.index("URL") + 1
                            _url = ws.cell(row=row_idx, column=_url_col_idx).value
                        except ValueError:
                            pass
                        
                        _msg_len = self._lenb(_msg) if _msg else 0
                        _url_len = self._lenb(_url) if _url else 0
                        
                        finance_ws.append([
                            self._sanitize_cell_value(_msg), 
                            self._sanitize_cell_value(_url), 
                            self._sanitize_cell_value(gubun_val), 
                            self._sanitize_cell_value(code_val), 
                            _msg_len, 
                            _url_len, 
                            self._sanitize_cell_value(prob_val),
                            self._sanitize_cell_value(semantic_val),
                            self._sanitize_cell_value(reason_val),
                            self._sanitize_cell_value("O" if result.get("red_group") else "")
                        ])'''
new_process_file = '''                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        _msg = ws.cell(row=row_idx, column=msg_col_idx).value
                        if _msg:
                            finance_ws.append([self._sanitize_cell_value(_msg)])'''
text = text.replace(old_process_file, new_process_file)


# 4. update process_kisa_txt
old_process_txt = '''                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        if finance_ws.max_row <= 1 and not finance_ws.cell(row=1, column=1).value:
                            headers = ["메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이", "Probability", "Semantic Class", "Reason", "Red Group"]
                            finance_ws.append(headers)
                            header_font = Font(bold=True)
                            header_align = Alignment(horizontal='center', vertical='center')
                            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            for cell in finance_ws[1]:
                                cell.font = header_font
                                cell.alignment = header_align
                                cell.fill = header_fill
                        finance_ws.append([
                            self._sanitize_cell_value(msg_val), 
                            self._sanitize_cell_value(url_val), 
                            self._sanitize_cell_value(gubun_val), 
                            self._sanitize_cell_value(code_val), 
                            msg_len, 
                            url_len, 
                            self._sanitize_cell_value(prob_val),
                            self._sanitize_cell_value(semantic_val),
                            self._sanitize_cell_value(reason_val),
                            self._sanitize_cell_value("O" if result.get("red_group") else "")
                        ])'''
new_process_txt = '''                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        if msg_val:
                            finance_ws.append([self._sanitize_cell_value(msg_val)])'''
text = text.replace(old_process_txt, new_process_txt)


# 5. update _update_summary_table
import re
new_summary = '''    def _update_summary_table(self, wb: Workbook, is_trap: bool, filename: str, spam_cnt: int, url_cnt: int, str_cnt: int, sen_cnt: int):
        """
        Updates the summary statistics table on 'TRAP.문장 중복제거' sheet.
        """
        target_sheet = "TRAP.문장 중복제거" if is_trap else "문장중복제거"
        if target_sheet not in wb.sheetnames:
            ws = wb.create_sheet(target_sheet)
        else:
            ws = wb[target_sheet]

        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # 헤더가 비어있으면 초기 뼈대 세팅
        if not ws["E2"].value:
            # 1) 헤더 및 날짜명칭
            match = re.search(r'(\\d{4})(\\d{2})(\\d{2})_([A-Za-z0-9]+)', filename)
            if match:
                m, d, group = match.group(2), match.group(3), match.group(4)
                try:
                    dt = datetime(int(match.group(1)), int(m), int(d))
                    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
                    wd = weekdays[dt.weekday()]
                    date_str = f"● {m}월 {d}일 ({wd})_{group}"
                except ValueError:
                    date_str = f"● {m}월 {d}일_{group}"
            else:
                date_str = "● 요약"

            ws["E1"] = date_str
            ws["E1"].font = Font(bold=True)

            headers = ["구분", "스팸태깅", "URL", "문자열", "문장"]
            for col_idx, h in enumerate(headers, start=5): # E=5
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            ws["E3"] = "SPAM"
            ws["E4"] = "TRAP"
            ws["E3"].font = Font(bold=True)
            ws["E4"].font = Font(bold=True)
            ws["E3"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws["E4"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws["E3"].alignment = Alignment(horizontal='center', vertical='center')
            ws["E4"].alignment = Alignment(horizontal='center', vertical='center')

            # 셀 초기화 (빈값은 '-'로 표시되게 하거나 그냥 '-' 문자 입력)
            for r in [3, 4]:
                for c in range(6, 10):
                    ws.cell(row=r, column=c, value="-").alignment = Alignment(horizontal='right')

        # 데이터 업데이트 (SPAM 이면 3행, TRAP 이면 4행)
        row_idx = 4 if is_trap else 3
        # 기존 값이 "숫자"인 경우 덮어쓰지 않고 최신값 유지 (빈 값은 덮어씀)
        if str(ws.cell(row=row_idx, column=6).value) == "-":
            ws.cell(row=row_idx, column=6, value=spam_cnt if spam_cnt else "-").alignment = Alignment(horizontal='right')     
        if str(ws.cell(row=row_idx, column=7).value) == "-":
            ws.cell(row=row_idx, column=7, value=url_cnt if url_cnt else "-").alignment = Alignment(horizontal='right')       
        if str(ws.cell(row=row_idx, column=8).value) == "-":
            ws.cell(row=row_idx, column=8, value=str_cnt if str_cnt else "-").alignment = Alignment(horizontal='right')       
        if str(ws.cell(row=row_idx, column=9).value) == "-":
            ws.cell(row=row_idx, column=9, value=sen_cnt if sen_cnt else "-").alignment = Alignment(horizontal='right')       

        # 테두리 적용 (E2:I4)
        for r in range(2, 5):
            for c in range(5, 10):
                ws.cell(row=r, column=c).border = thin_border

        # 너비 조정
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
'''

text = re.sub(r'    def _update_summary_table\(self, wb: Workbook, is_trap: bool, filename: str, spam_cnt: int, url_cnt: int, str_cnt: int, sen_cnt: int\):.*?(?=(?:\n    def |\Z))', new_summary, text, flags=re.DOTALL)

with open(TARGET, "w", encoding="utf-8") as f:
    f.write(text)

print("Patch applied successfully, updated summary table code.")
