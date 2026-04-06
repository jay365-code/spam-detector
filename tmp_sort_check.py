# -*- coding: utf-8 -*-
import sys
import collections
from openpyxl import load_workbook

wb = load_workbook(r'c:\Users\leejo\Project\AI Agent\Spam Detector\spams\MMSC스팸추출_20260327_A.xlsx', data_only=True)
ws = wb['육안분석(시뮬결과35_150)']

headers = [str(c.value).strip() if c.value else f'Col{i}' for i, c in enumerate(ws[1])]

data = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    if not row[0].value: break
    color = 'None'
    if row[0].fill and row[0].fill.start_color and hasattr(row[0].fill.start_color, 'rgb'):
        color = str(row[0].fill.start_color.rgb)
        
    row_dict = {'row': row_idx, 'color': color}
    for i, c in enumerate(row):
        if i < len(headers):
            row_dict[headers[i]] = c.value
    data.append(row_dict)

groups = collections.defaultdict(list)
for d in data:
    groups[d['color']].append(d)

for color in ['00000000', 'Values must be of type <class \'str\'>', 'FFFFD1D1', 'None']:
    if color not in groups: continue
    items = groups[color]
    print(f'\n--- Group {color} (Count: {len(items)}) ---')
    print("Row | 구분 | 분류 | prob | sem | msg_len | url_len | msg")
    for item in items[:20]:
        msg_snip = str(item.get('메시지', ''))[:20].replace('\n', ' ')
        print(f"{item['row']:04} | {item.get('구분')} | {item.get('분류')} | {item.get('Probability')} | {item.get('Semantic Class')} | {item.get('메시지 길이')} | {item.get('URL 길이')} | {msg_snip}")
