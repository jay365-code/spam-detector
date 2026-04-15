import re
from openpyxl.styles import PatternFill, Alignment

def generate_excel_from_json(self, logs: list, output_path: str, is_trap: bool, original_filename: str = None) -> dict:
    """
    Re-generate the Excel file entirely from the UI's JSON state (logs).
    """
    wb = self.Workbook()
    # 1. 시트 초기화 및 헤더 작성
    main_sheet_name = "TRAP.시뮬결과전체" if is_trap else "시뮬결과전체"
    ws = wb.active
    ws.title = main_sheet_name
    
    headers = ["메시지", "추출된 URL(직접접속금지)", "구분(A,B)", "유형", "문자", "URL", "Probability", "Semantic Class", "Reason", "Red Group"]
    ws.append(headers)
    
    # 헤더 스타일
    header_fill = self.openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = self.openpyxl.styles.Font(bold=True)
        cell.alignment = self.openpyxl.styles.Alignment(horizontal="center")
        
    unique_urls = {}
    unique_short_urls = {}
    blocklist_data = []
    stats = {"spam_count": 0}
    
    for log_item in logs:
        # User requested exclusion
        if log_item.get("result", {}).get("exclude_from_excel"):
            continue
            
        result = log_item.get("result", {})
        msg_val = log_item.get("message", "") # Fallback to original message
        if not msg_val:
            msg_val = result.get("message", "") # Just in case
            
        # Ensure we capture modified urls if any. Wait, the user might have modified extracted_urls? 
        # The edit modal doesn't let them edit URLs directly right now. It uses what's in 'message_extracted_url' or the original.
        url_val = result.get("message_extracted_url", "")
            
        is_spam = result.get("is_spam")
        semantic_class = result.get("semantic_class", "")
        reason_val = result.get("reason", "")
        
        is_type_b = str(semantic_class).startswith("Type_B")
        is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(reason_val)
        
        if is_separated:
            gubun_val = ""
            raw_code = str(result.get("classification_code", ""))
            match = re.search(r'\d+', raw_code)
            code_val = match.group(0) if match else raw_code
        elif is_type_b or is_spam is True:
            gubun_val = "o"
            stats["spam_count"] += 1
            raw_code = str(result.get("classification_code", ""))
            match = re.search(r'\d+', raw_code)
            code_val = match.group(0) if match else raw_code
        else:
            gubun_val = ""
            code_val = ""

        prob_float = result.get("spam_probability", 0.0)
        prob_val = f"{int(float(prob_float) * 100)}%" if str(prob_float).replace('.','').isdigit() else "0%"
        
        # Lengths
        msg_len = self._lenb(msg_val)
        url_len = self._lenb(url_val)
        
        # Write Row
        ws.append([
            self._sanitize_cell_value(msg_val), 
            self._sanitize_cell_value(url_val), 
            self._sanitize_cell_value(gubun_val), 
            self._sanitize_cell_value(code_val), 
            msg_len, 
            url_len, 
            self._sanitize_cell_value(prob_val),
            self._sanitize_cell_value(semantic_class),
            self._sanitize_cell_value(reason_val),
            self._sanitize_cell_value("O" if result.get("red_group") else "")
        ])
        
        if str(code_val) == "3":
            finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
            if msg_val:
                finance_ws.append([self._sanitize_cell_value(msg_val)])
                
        # Handle Dedup tracking
        if url_val:
            for u in url_val.split(","):
                u = u.strip()
                if u:
                    if u not in unique_urls:
                        unique_urls[u] = code_val
                        
        # Blocklist tracking - IBSE Signature
        ibse_sig = result.get("ibse_signature")
        if ibse_sig and is_spam:
            ibse_len = result.get("ibse_len", self._lenb(ibse_sig))
            code_int = int(code_val) if str(code_val).isdigit() else 2
            blocklist_data.append({
                "sig": self._sanitize_cell_value(ibse_sig),
                "len": ibse_len,
                "code": code_int
            })

    # Sort & Format
    self._sort_sheet_by_type(ws, headers)
    self._apply_formatting(ws, headers)
    
    # Create sheets
    dedup_sheet_name = "TRAP.URL중복 제거" if is_trap else "URL중복 제거"
    self._create_dedup_sheet(wb, unique_urls, unique_short_urls, sheet_name=dedup_sheet_name)
    
    blocklist_sheet_name = "TRAP.문자문장차단등록" if is_trap else "문자문장차단등록"
    self._create_blocklist_sheet(wb, blocklist_data, sheet_name=blocklist_sheet_name)
    
    sheet_name_str = "TRAP.문자열 중복제거" if is_trap else "문자열중복제거"
    sheet_name_sen = "TRAP.문장 중복제거" if is_trap else "문장중복제거"
    str_cnt, sen_cnt = self._create_split_dedup_sheets(wb, blocklist_data, sheet_name_str, sheet_name_sen)
    
    url_cnt = len(unique_urls) + len(unique_short_urls)
    self._update_summary_table(wb, is_trap, original_filename or "generated.xlsx", stats["spam_count"], url_cnt, str_cnt, sen_cnt)
    
    wb.save(output_path)
    return {"success": True, "output_path": output_path, "filename": original_filename, "total_rows": len(logs)}
