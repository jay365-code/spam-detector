import logging
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

class ExcelHandler:
    def __init__(self):
        self.shortener_domains = self._load_shortener_domains()
        
    def create_template_workbook(self, output_path: str):
        """
        KISA 및 TRAP 데이터를 저장할 엑셀 기본 구조(13개 시트)를 생성하여 저장합니다.
        """
        wb = Workbook()
        
        sheet_names = [
            "시뮬결과전체", "육안분석(시뮬결과35_150)", "URL중복 제거", "문자문장차단등록",
            "문자열중복제거", "문장중복제거", 
            "TRAP.시뮬결과전체", "TRAP.육안분석(시뮬결과35_150)", "TRAP.중간작업", 
            "TRAP.URL중복 제거", "TRAP.문자문장차단등록",
            "TRAP.문자열 중복제거", "TRAP.문장 중복제거", "금융.SPAM"
        ]
        
        # 0번 시트 이름 변경
        ws = wb.active
        ws.title = sheet_names[0]
        
        # 나머지 생성
        for name in sheet_names[1:]:
            wb.create_sheet(name)
            
        # ═══ 공통 스타일 정의 ═══
        hdr_font = Font(name='맑은 고딕', bold=True, size=11)
        hdr_font_10 = Font(name='맑은 고딕', bold=True, size=10)
        hdr_align = Alignment(horizontal='center', vertical='center')
        hdr_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        def _apply_hdr(ws_t, headers_t, col_widths=None, font=None):
            _f = font or hdr_font
            for ci, hv in enumerate(headers_t, 1):
                cell = ws_t.cell(row=1, column=ci, value=hv)
                if hv:
                    cell.font = _f
                    cell.alignment = hdr_align
                    cell.fill = hdr_fill
            if col_widths:
                for cl, w in col_widths.items():
                    ws_t.column_dimensions[cl].width = w
        
        _apply_hdr(wb['시뮬결과전체'],
            ["메시지", "URL", "구분", "메시지길이", "URL길이"],
            {'A': 81.0, 'B': 25.6, 'C': 10.6, 'D': 11.9})
        
        _apply_hdr(wb['육안분석(시뮬결과35_150)'],
            ["메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이",
             "Probability", "Semantic Class", "Reason", "Red Group"],
            {'A': 90.0, 'B': 22.6, 'C': 10.6, 'D': 6.6, 'E': 12.6, 'F': 10.6},
            font=hdr_font_10)
        
        _url_hdr = ["URL(중복제거)", "길이", "분류"]
        for _ in range(16):
            _url_hdr.append("")
        _url_hdr.extend(["URL(단축URL)", "길이", "분류"])
        _apply_hdr(wb['URL중복 제거'], _url_hdr,
            {'A': 30.3, 'C': 8.6, 'T': 29.6}, font=hdr_font_10)
        
        _apply_hdr(wb['문자문장차단등록'],
            ["메시지", "문자열", "길이", "문장열", "길이", "분류"],
            {'A': 80.6, 'B': 20.6, 'C': 7.6, 'D': 37.6, 'E': 7.6})
        
        _apply_hdr(wb['문자열중복제거'],
            ["문자열(중복제거)", "길이", "분류"], {'A': 25.9})
        
        _apply_hdr(wb['문장중복제거'],
            ["문장(중복제거)", "길이", "분류"], {'A': 46.1})
        
        _apply_hdr(wb['TRAP.시뮬결과전체'],
            ["메시지", "URL", "구분", "메시지길이", "URL길이"],
            {'A': 89.6, 'B': 25.6, 'C': 10.6})
        
        _apply_hdr(wb['TRAP.육안분석(시뮬결과35_150)'],
            ["메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이",
             "Probability", "Semantic Class", "Reason", "Red Group"],
            {'A': 89.6, 'B': 21.6, 'C': 13.6},
            font=hdr_font_10)
        
        _apply_hdr(wb['TRAP.중간작업'],
            ["메시지", "URL", "문자열(9byte이상)", "문자열(길이)", "문장(40byte)", "길이", "분류"],
            {'A': 65.6, 'B': 25.6, 'C': 19.6, 'D': 10.6, 'E': 40.6, 'F': 10.6})
        
        _apply_hdr(wb['TRAP.URL중복 제거'],
            ["TRAP URL(중복제거)", "길이", "분류"],
            {'A': 25.6, 'B': 10.6})
        
        _apply_hdr(wb['TRAP.문자문장차단등록'],
            ["메시지", "문자열", "길이", "문장열", "길이", "분류"],
            {'A': 80.6, 'B': 20.6, 'C': 7.6, 'D': 37.6, 'E': 7.6})
        
        _apply_hdr(wb['TRAP.문자열 중복제거'],
            ["TRAP 문자열(중복제거)", "길이", "분류"],
            {'A': 25.6, 'B': 10.6})
        
        _apply_hdr(wb['TRAP.문장 중복제거'],
            ["TRAP 문장(중복제거)", "길이", "분류"],
            {'A': 40.6, 'B': 10.6, 'E': 12.6, 'F': 10.6})
        
        wb['금융.SPAM'].column_dimensions['A'].width = 89.6
                
        # 자동 저장
        dirname = os.path.dirname(output_path)
        if dirname:
            os.makedirs(dirname, exist_ok=True)
        wb.save(output_path)
        return True
        
    def _load_shortener_domains(self) -> set:
        # 기본 하드코딩된 단축 도메인 세트
        domains = {
            "a.to", "abit.ly", "adf.ly", "adfoc.us", "agshort.link", "aka.ms", "amzn.to", "apple.co", "asq.kr", 
            "bit.do", "bit.ly", "bitly.com", "bitly.cx", "bitly.kr", "bl.ink", "blow.pw", "buff.ly", "buly.kr", 
            "c11.kr", "clic.ke", "cogi.cc", "coupa.ng", "cutt.it", "cutt.ly", 
            "di.do", "dokdo.in", "dub.co", 
            "fb.me", 
            "gmarket.it", "goo.gl", "goo.su", "gooal.kr", 
            "han.gl", "horturl.at", 
            "ii.ad", "iii.ad", "instagr.am", "is.gd", 
            "j.mp", 
            "kakaolink.com", "ko.fm", "ko.gl", "koe.kr", 
            "link24.kr", "linktr.ee", "lrl.kr", 
            "mcaf.ee", "me2.do", "muz.so", "myip.kr", 
            "naver.me", 
            "ouo.io", "ow.ly", 
            "qrco.de", 
            "rb.gy", "rebrand.ly", "reurl.kr", "rul.kr",
            "sbz.kr", "short.io", "shorter.me", "shorturl.at", "shrl.me", "shrtco.de", 
            "t.co", "t.ly", "t.me", "t2m.kr", "tiny.cc", "tinyurl.com", "tne.kr", "tny.im", "tr.ee", "tuney.kr",
            "url.kr", "uto.kr", 
            "v.gd", "vo.la", "vvd.bz", "vvd.im", 
            "wp.me", 
            "youtu.be", "yun.kr", 
            "zrr.kr",
            # 새롭게 식별된 단축 URL 추가
            "tnia.cc", "2.lnkme.net", "lnkme.net", "alie.kr"
        }
        
        # 외부 클론잡 동작용 txt 파일이 존재하면 정적으로 병합 구성
        try:
            list_path = os.path.join(os.path.dirname(__file__), "shorteners_list.txt")
            if os.path.exists(list_path):
                with open(list_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip().lower()
                        if line and not line.startswith("#"):
                            domains.add(line)
        except Exception as e:
            logger.warning(f"Failed to load shorteners_list.txt: {e}")
            
        return domains

    def _lenb(self, text: str) -> int:
        """Calculate byte length for CP949 encoding (Korean standard)."""
        if not isinstance(text, str):
            text = str(text) if text is not None else ""
        try:
            return len(text.encode('cp949'))
        except UnicodeEncodeError:
            # Fallback for chars not in cp949
            return len(text.encode('utf-8'))

    def _sanitize_cell_value(self, value):
        """
        Prevent Excel Formula Injection (CSV Injection).
        If value starts with =, +, -, @, prepend ' to treat as text.
        Also remove illegal characters.
        """
        if not isinstance(value, str):
            return value
        
        # 1. OpenPyXL handles illegal XML chars (mostly), but let's be safe against nulls
        value = value.replace('\0', '')
        
        # 2. Formula Injection prevention
        if value.startswith(('=', '+', '-', '@')):
            return "'" + value
        
        return value

    def _to_code_int(self, code) -> int | str:
        """
        분류 코드를 정수(int)로 변환합니다.
        엑셀에서 숫자는 자동으로 우측 정렬되므로, 샘플과 동일한 표시 형식을 위해
        '2', '3' 등 문자열로 저장된 코드를 int로 변환합니다.
        변환 불가능한 경우 원본 값을 그대로 반환합니다.
        """
        if code is None or code == '':
            return code
        try:
            return int(str(code).strip())
        except (ValueError, TypeError):
            return code


    def _sort_sheet_by_type(self, ws, headers: list):
        """
        분류(Semantic Class) 기준으로 자동 그룹핑 및 정렬:
        1. Type A
        2. Type B (서브타입 - URL, SIGNATURE 등 - 알파벳 정렬)
        3. 텍스트 HAM + 악성 URL 발췌 (분리 감지)
        4. 기타 미분류 SPAM
        5. 일반 HAM
        """
        if ws.max_row <= 1 or not headers:
            return  # 헤더만 있거나 데이터 없음

        # 헤더 이름을 기반으로 동적 컬럼 인덱스 찾기(1-based)
        def get_col_idx(name: str):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None
                
        msg_col_idx      = get_col_idx("메시지")
        url_col_idx      = get_col_idx("URL")
        gubun_col_idx    = get_col_idx("구분")
        reason_col_idx   = get_col_idx("Reason")
        semantic_col_idx = get_col_idx("Semantic Class")
        red_col_idx      = get_col_idx("Red Group")   # [Fix] Red Group 컬럼 인덱스 추가
        
        # 모든 데이터 행 읽기 (헤더 제외)
        data_rows = []
        for row_idx in range(2, ws.max_row + 1):
            row_data = []
            for col_idx in range(1, ws.max_column + 1):
                row_data.append(ws.cell(row=row_idx, column=col_idx).value)
            data_rows.append(row_data)
        
        # 정렬 기준 안전 변환 함수
        def safe_get(row, idx):
            return str(row[idx - 1] if idx and len(row) >= idx else "").strip()
            
        # ── 정렬 키 함수 ──────────────────────────────────────────────────────
        def rank_key(row):
            semantic_val = safe_get(row, semantic_col_idx)
            reason_val   = safe_get(row, reason_col_idx)
            gubun_val    = safe_get(row, gubun_col_idx)

            # [Fix] Red Group 판단: Reason 텍스트 대신 Red Group 컬럼 값('O')으로 직접 확인
            # 기존: "[수동 Red Group 지정]" in reason_val  → AI 자동 Red Group 항목 누락 버그
            # 수정: Red Group 컬럼이 'O'이면 수동/자동 구분 없이 모두 Red Group으로 인식
            is_red_group = safe_get(row, red_col_idx).upper() == "O" if red_col_idx else False
            is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in reason_val
            is_type_b    = semantic_val.startswith("Type_B") or "[FP Sentinel Override]" in reason_val
            is_type_a    = semantic_val.startswith("Type_A")
            is_spam      = (gubun_val.lower() == "o")

            # 색상/그룹 우대 순위:
            # 0: 일반 SPAM (황금색 - 최상단)
            # 1: Red Group 및 텍스트 HAM + 악성 URL 분리 (핑크색 - 스팸 하단)
            # 2: HAM (투명)
            # 3: 기타
            if is_separated or is_red_group:
                return 1
            elif is_type_a or is_type_b or is_spam:
                return 0
            elif semantic_val.lower() == "ham":
                return 2
            else:
                return 3

        # ── 단일 복합 키 정렬 ────────────────────────────────────────────────
        # [Fix] 기존 3단계 분리 정렬(URL→메시지→랭크)은 URL 유무 구분이 메시지 정렬에
        # 덮어씌워지는 문제가 있었음. 단일 복합 튜플 키로 우선순위를 명확히 정의.
        #
        # 최종 정렬 순서 (기준 샘플 MMSC스팸추출_20260414_A.xlsx 기반):
        #   0순위 그룹(SPAM) URL있음 → 0순위 그룹(SPAM) URL없음
        #   1순위 그룹(RedGroup) URL있음 → 1순위 그룹(RedGroup) URL없음
        #   2순위 그룹(HAM) URL있음 → 2순위 그룹(HAM) URL없음
        #   각 소그룹 내에서는 메시지 가나다순(A→Z)
        def sort_key(row):
            rank    = rank_key(row)
            has_url = 0 if safe_get(row, url_col_idx) else 1  # URL있음=0(앞), URL없음=1(뒤)
            msg     = safe_get(row, msg_col_idx)
            return (rank, has_url, msg)

        data_rows.sort(key=sort_key)
        
        
        # 정렬된 데이터로 다시 쓰기
        for i, row_data in enumerate(data_rows):
            row_idx = i + 2  # 헤더가 1행이므로 2행부터
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _apply_formatting(self, ws, headers: list):
        """
        엑셀 서식 적용:
        - 메시지 컬럼: 너비 90, 자동줄바꿈, SPAM인 경우 황금색 강조
        - URL 컬럼: 너비 22
        - 구분 컬럼: 텍스트 중앙 정렬
        - 분류 컬럼: 텍스트 중앙 정렬
        - 메시지 길이 컬럼: 너비 10
        - Probability 컬럼: 텍스트 중앙 정렬, 너비 10
        - Reason 컬럼: 너비 90, 자동줄바꿈
        """
        # 컬럼 인덱스 찾기 (1-based)
        def get_col_idx(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None
        
        msg_col = get_col_idx("메시지")
        url_col = get_col_idx("URL")
        gubun_col = get_col_idx("구분")
        code_col = get_col_idx("분류")
        msg_len_col = get_col_idx("메시지 길이")
        prob_col = get_col_idx("Probability")
        semantic_col = get_col_idx("Semantic Class")
        reason_col = get_col_idx("Reason")
        
        # 컬럼 너비 설정
        if msg_col:
            ws.column_dimensions[get_column_letter(msg_col)].width = 90
        if url_col:
            ws.column_dimensions[get_column_letter(url_col)].width = 22
        if msg_len_col:
            ws.column_dimensions[get_column_letter(msg_len_col)].width = 10
        if prob_col:
            ws.column_dimensions[get_column_letter(prob_col)].width = 10
        if semantic_col:
            ws.column_dimensions[get_column_letter(semantic_col)].width = 15
        if reason_col:
            ws.column_dimensions[get_column_letter(reason_col)].width = 90
        
        # 정렬 스타일 정의
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_vcenter_align = Alignment(wrap_text=True, vertical='center')
        from openpyxl.styles import Border, Side, Font
        # 폰트 스타일 정의
        base_font = Font(size=10.5)
        msg_font = Font(size=10.5)
        
        # 강조 (Gold, Accent 4, Lighter 80%) = FFF2CC
        spam_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # 테두리: 얇은 실선(thin), 색상: 밝은 회색(BFBFBF: 배경1, 25% 더 어둡게)
        border_side = Side(style='thin', color='BFBFBF')
        box_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # 헤더(1행) 폰트 적용 (10.5) 및 테두리 적용
        header_font = Font(size=10, bold=True)
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            header_cell.font = header_font
            header_cell.border = box_border
        # 엑셀 자동 필터(정렬 드롭다운) 적용 (전체 데이터 영역)
        ws.auto_filter.ref = ws.dimensions

        for row_idx in range(2, ws.max_row + 1):
            # 행 단위 전체 셀에 테두리 및 기본 폰트 적용
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = box_border
                
                # 폰트 적용 (메시지는 10.5, 나머지는 10)
                if msg_col and col_idx == msg_col:
                    cell.font = msg_font
                else:
                    cell.font = base_font
                
            # 구분 컬럼 중앙 정렬
            if gubun_col:
                ws.cell(row=row_idx, column=gubun_col).alignment = center_align
            
            # 분류 컬럼 중앙 정렬
            if code_col:
                ws.cell(row=row_idx, column=code_col).alignment = center_align
            
            # Probability 컬럼 중앙 정렬
            if prob_col:
                ws.cell(row=row_idx, column=prob_col).alignment = center_align

            # 신규 컬럼 중앙 정렬
            if semantic_col:
                ws.cell(row=row_idx, column=semantic_col).alignment = center_align

            # 메시지 길이 및 URL 길이 중앙 정렬
            if msg_len_col:
                ws.cell(row=row_idx, column=msg_len_col).alignment = center_align
            url_len_col = get_col_idx("URL 길이")
            if url_len_col:
                ws.cell(row=row_idx, column=url_len_col).alignment = center_align
            
            # 메시지 컬럼: 세로 중앙 (자동줄바꿈 해제)
            if msg_col:
                ws.cell(row=row_idx, column=msg_col).alignment = Alignment(vertical='center', wrap_text=False)
            
            # Reason 컬럼: VCenter만 적용 (자동줄바꿈 해제)
            if reason_col:
                ws.cell(row=row_idx, column=reason_col).alignment = Alignment(vertical='center', wrap_text=False)
            
            # SPAM(o), Type_B, TEXT+URL분리인 경우 메시지 셀 채우기 적용
            if gubun_col and msg_col:
                gubun_val = ws.cell(row=row_idx, column=gubun_col).value
                reason_val = ws.cell(row=row_idx, column=reason_col).value if reason_col else ""
                semantic_val = ws.cell(row=row_idx, column=semantic_col).value if semantic_col else ""
                
                is_type_b = bool(semantic_val and str(semantic_val).startswith("Type_B")) or bool(
                    reason_val and "[FP Sentinel Override]" in str(reason_val)
                )
                is_separated = False
                if reason_val and "[텍스트 HAM + 악성 URL 분리 감지" in str(reason_val):
                    is_separated = True
                    
                red_col = get_col_idx("Red Group")
                is_red_group = (ws.cell(row=row_idx, column=red_col).value == "O") if red_col else False

                cell = ws.cell(row=row_idx, column=msg_col)
                vcenter_align = Alignment(vertical='center', wrap_text=False)
                if is_separated or is_red_group: # TEXT-HAM + URL-SPAM or Red Group
                    # 사용자 요청: 텍스트 HAM + 악성 URL 분리 감지 오버라이딩은 기존 Type B 처리 방식(FFCCCC) 유지
                    type_b_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.fill = type_b_fill
                    cell.alignment = vcenter_align
                elif gubun_val == "o" or is_type_b:  # 일반 SPAM (Type A) & 일반 Type B
                    # 사용자 요청: 일반 Type B는 Type A와 동일하게 황금색 강조
                    cell.fill = spam_fill
                    cell.alignment = vcenter_align
                else:
                    cell.alignment = vcenter_align


    def is_short_url(self, url: str) -> bool:
        """
        Check if the URL belongs to a known shortener service.
        """
        if not url: return False
        
        try:
            # Simple domain extraction for check
            # Remove protocol
            clean_url = re.sub(r'^https?://', '', url.lower())
            clean_url = re.sub(r'^www\.', '', clean_url)
            
            # Extract just the host part (before / or ? or :)
            host = clean_url.split('/')[0].split('?')[0].split(':')[0]
            
            return host in self.shortener_domains
        except:
            return False

    def _create_dedup_sheet(self, wb: Workbook, unique_urls: dict, unique_short_urls: dict = None, sheet_name: str = "URL중복 제거"):
        """
        Create 'URL중복 제거' (or 'TRAP.URL중복 제거') sheet with unique URLs.
        KISA: Columns A-C (일반 URL) + T-V (단축 URL) — 22컬럼
        TRAP: Columns A-C (모든 URL 통합) — 3컬럼 (샘플 기준)
        """
        is_trap = sheet_name.startswith("TRAP.")
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Style Definition
        header_font = Font(name='맑은 고딕', bold=True, size=10)
        base_font = Font(name='맑은 고딕', size=10.5)
        msg_font = Font(name='맑은 고딕', size=10.5)
        header_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        data_align = Alignment(vertical='center')
        # [Fix] 분류 컬럼 우측정렬 제거 (샘플 기준: 기본 세로 중앙정렬)
        # cls_align = Alignment(horizontal='right', vertical='center', indent=1)  ← 제거됨
        
        # 헤더 중복 방지 (create_template_workbook에서 이미 생성되었을 수 있음)
        headers_exist = ws.cell(row=1, column=1).value is not None
        
        if is_trap:
            # ── TRAP: 3컬럼 통합 형식 (샘플 기준) ──
            if not headers_exist:
                trap_headers = ["TRAP URL(중복제거)", "길이", "분류"]
                for ci, hv in enumerate(trap_headers, 1):
                    cell = ws.cell(row=1, column=ci, value=hv)
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.fill = header_fill
                ws.column_dimensions['A'].width = 25.6
                ws.column_dimensions['B'].width = 10.6
            
            # 데이터: 일반 URL + 단축 URL 통합 후 길이순 정렬
            all_urls = {}
            all_urls.update(unique_urls)
            if unique_short_urls:
                all_urls.update(unique_short_urls)
            
            sorted_all = sorted(all_urls.items(), key=lambda x: (x[1]['len'], str(x[0])))
            row_num = 2
            for url, info in sorted_all:
                ws.cell(row=row_num, column=1, value=self._sanitize_cell_value(url)).font = msg_font
                ws.cell(row=row_num, column=2, value=info['len']).font = base_font
                ws.cell(row=row_num, column=3, value=self._to_code_int(info['code'])).font = base_font
                for col_idx in range(1, 4):
                    ws.cell(row=row_num, column=col_idx).alignment = data_align
                row_num += 1
        else:
            # ── KISA: 22컬럼 형식 (A-C 일반, T-V 단축) ──
            if not headers_exist:
                headers = ["URL(중복제거)", "길이", "분류"]
                for _ in range(16):
                    headers.append("")
                headers.extend(["URL(단축URL)", "길이", "분류"])
                for ci, hv in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=ci, value=hv)
                    if hv:
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.fill = header_fill
                # 컬럼 너비 (샘플 기준)
                ws.column_dimensions[get_column_letter(1)].width = 30.3
                ws.column_dimensions[get_column_letter(3)].width = 8.6
                ws.column_dimensions[get_column_letter(20)].width = 29.6
            
            # Write Data (Normal URLs)
            row_num = 2
            sorted_urls = sorted(unique_urls.items(), key=lambda x: (x[1]['len'], str(x[0])))
            for url, info in sorted_urls:
                ws.cell(row=row_num, column=1, value=self._sanitize_cell_value(url)).font = msg_font
                ws.cell(row=row_num, column=2, value=info['len']).font = base_font
                ws.cell(row=row_num, column=3, value=self._to_code_int(info['code'])).font = base_font
                for col_idx in range(1, 4):
                    ws.cell(row=row_num, column=col_idx).alignment = data_align
                row_num += 1

            # Write Data (Short URLs)
            if unique_short_urls:
                row_num = 2
                sorted_short_urls = sorted(unique_short_urls.items(), key=lambda x: (x[1]['len'], str(x[0])))
                for url, info in sorted_short_urls:
                    ws.cell(row=row_num, column=20, value=self._sanitize_cell_value(url)).font = msg_font
                    ws.cell(row=row_num, column=21, value=info['len']).font = base_font
                    ws.cell(row=row_num, column=22, value=self._to_code_int(info['code'])).font = base_font
                    for col_idx in range(20, 23):
                        ws.cell(row=row_num, column=col_idx).alignment = data_align
                    row_num += 1

    def _create_blocklist_sheet(self, wb: Workbook, blocklist_data: list, sheet_name: str = "문자문장차단등록"):
        """
        Create '문자문장차단등록' (or 'TRAP.문자문장차단등록') sheet for extracted signatures.
        Columns: 메시지, 문자열, 길이, 문장열, 길이, 분류
        """
        if sheet_name in wb.sheetnames:
             ws = wb[sheet_name]
        else:
             ws = wb.create_sheet(sheet_name)
        
        # Headers (create_template_workbook에서 이미 생성되었을 수 있으므로 중복 방지)
        headers = ["메시지", "문자열", "길이", "문장열", "길이", "분류"]
        if not ws.cell(row=1, column=1).value:
            ws.append(headers)
        
        # Styling
        header_font = Font(bold=True, size=10)
        base_font = Font(size=10.5)
        msg_font = Font(size=10.5)
        msg_fill = PatternFill(start_color="D8EFD3", end_color="D8EFD3", fill_type="solid")
        
        header_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # 데이터 행 정렬
        data_align = Alignment(vertical='center')
        # [Fix] 문자열 컬럼(B) 전용: 가로 중앙정렬 추가 (샘플 기준)
        sig_align  = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
            
        # 컬럼 너비 조정 (픽셀 -> 엑셀 width 환산)
        ws.column_dimensions[get_column_letter(1)].width = 80.6   # 메시지
        ws.column_dimensions[get_column_letter(2)].width = 20.6   # 문자열
        ws.column_dimensions[get_column_letter(3)].width = 7.6    # 문자열 길이
        ws.column_dimensions[get_column_letter(4)].width = 37.6   # 문장열
        ws.column_dimensions[get_column_letter(5)].width = 7.6    # 문장열 길이
        ws.column_dimensions[get_column_letter(6)].width = 10     # 분류
            
        # Write Data
        # Deduplicate by signature 제거 (사용자 오더: 문자문장차단등록 시트는 모든 원본 데이터 보존)
        # Sort data: 문자열 (length <= 20) first, then 문장열 (length > 20).
        # Within each category, sort by length descending, and then by signature to group identical strings together.
        def sort_key(item):
            is_sentence = item['len'] > 20
            return (is_sentence, -item['len'], str(item['sig']))
            
        sorted_data = sorted(blocklist_data, key=sort_key)
        
        row_num = 2
        for item in sorted_data:
            msg = self._sanitize_cell_value(item['msg'])
            sig = self._sanitize_cell_value(item['sig'])
            length = item['len']
            code = self._to_code_int(item['code'])
            
            if length <= 20: # 문자열 부분
                row_data = [msg, sig, length, "", 0, code]
            else:            # 문장열 부분
                row_data = [msg, "", 0, sig, length, code]
                
            ws.append(row_data)
            
            # 데이터 셀 폰트 및 채우기 적용
            ws.cell(row=row_num, column=1).font = msg_font
            ws.cell(row=row_num, column=1).fill = msg_fill
            
            for c in range(2, 7):
                ws.cell(row=row_num, column=c).font = base_font
            
            # [Fix] 데이터 셀 정렬: 분류 우측정렬 제거, 문자열(B) 가로 중앙정렬 적용
            for c in range(1, 7):
                ws.cell(row=row_num, column=c).alignment = data_align
            ws.cell(row=row_num, column=2).alignment = sig_align   # 문자열 컬럼: 중앙정렬
            
            row_num += 1


    def process_file(self, file_path: str, output_path: str, processing_function, progress_callback=None, batch_size: int = 1):
        """
        Reads the Excel file, processes rows in batches, and appends results.
        Supports optional progress reporting.
        """
        try:
            # 1. Load Workbook
            wb = load_workbook(file_path)
            target_sheet_name = "육안분석(시뮬결과35_150)"
            
            if target_sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet_name}' not found.")
            
            ws = wb[target_sheet_name]
            
            # 2. Identify Headers
            headers = [cell.value for cell in ws[1]]
            try:
                msg_col_idx = headers.index("메시지") + 1 # 1-based index
            except ValueError:
                raise ValueError("'메시지' column not found.")
            
            try:
                url_col_idx = headers.index("URL") + 1 # 1-based index
            except ValueError:
                url_col_idx = -1
                
            # Add or Find Output Columns (Cols setup logic same as before)
            
            def get_col_idx(name, default_idx):
                try:
                    return headers.index(name) + 1
                except ValueError:
                    ws.cell(row=1, column=default_idx, value=name)
                    return default_idx

            gubun_col_idx = get_col_idx("구분", len(headers) + 1)
            code_col_idx = get_col_idx("분류", gubun_col_idx + 1)
            prob_col_idx = get_col_idx("Probability", code_col_idx + 1)
            semantic_col_idx = get_col_idx("Semantic Class", prob_col_idx + 1)
            reason_col_idx = get_col_idx("Reason", semantic_col_idx + 1)
            in_token_col_idx = get_col_idx("In_Token", reason_col_idx + 1)
            out_token_col_idx = get_col_idx("Out_Token", in_token_col_idx + 1)
            red_group_col_idx = get_col_idx("Red Group", out_token_col_idx + 1)

            # 3. Iterate Rows & Batch Processing
            total_rows = ws.max_row - 1 # Excluding header
            # 전체 메시지 큐 로드: ≤2000이면 전체, >2000이면 2000개씩 청크
            effective_batch_size = min(2000, total_rows) if total_rows > 0 else 1
            logger.info(f"Processing {total_rows} rows from Excel (Batch Size: {effective_batch_size})...")
            
            row_iterator = ws.iter_rows(min_row=2, max_row=ws.max_row)
            
            batch_buffer = [] # List of (vocab_row_idx, message_str, row_object)
            unique_urls = {} # URL Reduplication Store
            unique_short_urls = {} # Short URL Reduplication Store
            blocklist_data = [] # IBSE Blocklist Store
            
            def flush_batch(start_index=0):
                if not batch_buffer:
                    return
                
                # Extract messages
                messages = [item[1] for item in batch_buffer]
                
                # Extract URLs
                pre_parsed_urls = []
                for item in batch_buffer:
                    r_obj = item[2]
                    u = r_obj[url_col_idx - 1].value if url_col_idx > 0 else ""
                    pre_parsed_urls.append(str(u).strip() if u else "")
                
                # Call Processing Function (Expects List -> Returns List)
                try:
                    # Pass start_index (global offset) to processing function for correct UI mapping
                    # processing_function signature: processing_function(messages, start_index=0, total_count=0)
                    results = processing_function(messages, start_index=start_index, total_count=total_rows, pre_parsed_urls=pre_parsed_urls)
                except TypeError:
                     # Fallback for backward compatibility if function doesn't accept start_index yet
                     try:
                         results = processing_function(messages, start_index=start_index, total_count=total_rows)
                     except TypeError:
                         results = processing_function(messages)
                except Exception as e:
                    from ..main import CancellationException
                    if isinstance(e, CancellationException):
                        raise  # [Phase 4] 취소 예외는 상위로 재전파 (강제 종료 처리)
                    logger.error(f"Batch Processing Failed: {e}")
                    # Create error results
                    results = [{"is_spam": None, "reason": f"Error: {e}"} for _ in messages]
                
                # Map results back to rows
                for idx, result in enumerate(results):
                    if idx >= len(batch_buffer): break # Safety
                    
                    row_idx, _, _ = batch_buffer[idx] # Current Row Element
                    
                    # [User Request] Skip Excel Write but SHOW in UI Report
                    if result.get("exclude_from_excel"):
                        logger.debug(f"Row {row_idx}: Skipped from Excel (exclude_from_excel=True), but showing in UI")
                        if progress_callback:
                            progress_callback({
                                "current": row_idx - 1,
                                "total": total_rows,
                                "message": batch_buffer[idx][1],
                                "excel_row_number": row_idx,
                                "result": result
                            })
                        continue

                    # Write Result
                    semantic_val = result.get("semantic_class", "")
                    is_type_b = str(semantic_val).startswith("Type_B")
                    
                    is_red_group = bool(result.get("red_group"))
                    if is_separated or is_red_group:
                        # 오버라이딩 된 경우 기존 Type B 처리 방식 (구분 공란)
                        gubun_val = ""
                        raw_val = str(result.get("classification_code", ""))
                        import re
                        match = re.search(r'\d+', raw_val)
                        code_val = match.group(0) if match else raw_val
                    elif is_type_b or result.get("is_spam") is True:
                        # 일반 Type B 및 Type A는 구분 "o"
                        gubun_val = "o"
                        raw_val = str(result.get("classification_code", ""))
                        import re
                        match = re.search(r'\d+', raw_val)
                        code_val = match.group(0) if match else raw_val
                    elif result.get("is_spam") is False:
                        gubun_val = ""
                        code_val = ""
                    else:
                        gubun_val = "UNKNOWN"
                        raw_val = str(result.get("classification_code") or "")
                        import re
                        match = re.search(r'\d+', raw_val)
                        code_val = match.group(0) if match else raw_val

                    extracted_url_code = ""
                    if result.get("malicious_url_extracted"):
                        raw_ext_code = str(result.get("url_spam_code") or "")
                        m_ext = re.search(r'\d+', raw_ext_code)
                        extracted_url_code = m_ext.group(0) if m_ext else raw_ext_code
                    prob_val = result.get("spam_probability", 0.0)
                    semantic_val = result.get("semantic_class", "")
                    reason_val = result.get("reason", "")
                    in_token_val = result.get("input_tokens", 0)
                    out_token_val = result.get("output_tokens", 0)
                    
                    ws.cell(row=row_idx, column=gubun_col_idx, value=gubun_val)
                    ws.cell(row=row_idx, column=code_col_idx, value=self._to_code_int(code_val))
                    ws.cell(row=row_idx, column=prob_col_idx, value=prob_val)
                    ws.cell(row=row_idx, column=semantic_col_idx, value=semantic_val)
                    ws.cell(row=row_idx, column=reason_col_idx, value=reason_val)
                    ws.cell(row=row_idx, column=in_token_col_idx, value=in_token_val)
                    ws.cell(row=row_idx, column=out_token_col_idx, value=out_token_val)
                    ws.cell(row=row_idx, column=red_group_col_idx, value="O" if result.get("red_group") else "")

                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        _msg = ws.cell(row=row_idx, column=msg_col_idx).value
                        if _msg:
                            finance_ws.append([self._sanitize_cell_value(_msg)])

                    # --- URL Collection Logic ---
                    # Only collect URL from the input column if SPAM or extracted from HAM
                    if result.get("is_spam") is True or result.get("malicious_url_extracted"):
                        if result.get("drop_url"):
                            try:
                                url_col_idx = headers.index("URL") + 1
                                ws.cell(row=row_idx, column=url_col_idx, value="")
                            except ValueError:
                                pass
                            try:
                                string_col_idx = headers.index("문자열") + 1
                                ws.cell(row=row_idx, column=string_col_idx, value="")
                            except ValueError:
                                pass
                        else:
                            raw_url_code = str(result.get("classification_code") or "")
                            _m = re.search(r'\d+', raw_url_code)
                            url_dedup_code = _m.group(0) if _m else raw_url_code
                            if not result.get("is_spam"):
                                url_dedup_code = extracted_url_code
                            
                            # Use the directly mapped input URL, not regex from message
                            url_val = ws.cell(row=row_idx, column=get_col_idx("URL", len(headers) + 1)).value
                            url_val_str = str(url_val).strip() if url_val else ""
                            
                            is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(result.get("reason", ""))
                            if is_separated and not url_val_str:
                                # [User Request] 분홍색 메시지(URL 분리감지)는 URL 필드가 없어도 강제로 채움
                                url_val_str = result.get("message_extracted_url", "")
                                url_col_idx = get_col_idx("URL", len(headers) + 1)
                                ws.cell(row=row_idx, column=url_col_idx, value=self._sanitize_cell_value(url_val_str))
                                
                            urls = [url_val_str] if url_val_str else []
                        
                        for url in urls:
                            # Clean URL
                            url = str(url).strip().rstrip('.,;:!?)]}"\'')
                            
                            if not self.is_short_url(url):
                                 # Only non-short URLs
                                 if url not in unique_urls:
                                     unique_urls[url] = {
                                         "len": self._lenb(url),
                                         "code": url_dedup_code,
                                         "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                     }
                            else:
                                 # Short URLs
                                 if url not in unique_short_urls:
                                     unique_short_urls[url] = {
                                         "len": self._lenb(url),
                                         "code": url_dedup_code,
                                         "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                     }

                    # --- IBSE Collection Logic ---
                    if result.get("ibse_signature"):
                        # User requested "공백제거된 메시지"
                        
                        clean_msg = re.sub(r'[ \t\r\n\f\v]+', '', str(batch_buffer[idx][1]))
                        
                        raw_ibse_code = str(result.get("classification_code") or "")
                        m_ibse = re.search(r'\d+', raw_ibse_code)
                        ibse_code = m_ibse.group(0) if m_ibse else raw_ibse_code
                        
                        blocklist_data.append({
                            "msg": clean_msg, 
                            "sig": result.get("ibse_signature"),
                            "len": result.get("ibse_len", 0),
                            "code": ibse_code
                        })

                    # Callback (Progress) 
                    if progress_callback:
                        progress_callback({
                            "current": row_idx - 1, # approx
                            "total": total_rows,
                            "message": batch_buffer[idx][1],
                            "excel_row_number": row_idx,  # Actual Excel row number
                            "result": result
                        })

                # Auto-Save after each batch to prevent data loss
                # Note: We don't write dedup sheet until the very end because we need full set.
                try:
                    wb.save(output_path)
                except Exception as save_err:
                    logger.error(f"Auto-save failed: {save_err}")

                batch_buffer.clear()

            # Loop
            for i, row in enumerate(row_iterator):
                msg_cell = row[msg_col_idx - 1]
                message = msg_cell.value
                current_row_idx = row[0].row 
                
                if not message:
                    message_str = ""
                else:
                    message_str = str(message)
                
                batch_buffer.append((current_row_idx, message_str, row))
                
                if len(batch_buffer) >= effective_batch_size:
                    logger.info(f"Processing Batch (Row {i+1}/{total_rows})...")
                    # Calculate start_index based on current position
                    current_start_index = i - len(batch_buffer) + 1
                    flush_batch(start_index=current_start_index)
            
            # Process remaining
            if batch_buffer:
                logger.info(f"Processing Remaining Batch...")
                # Calculate start_index for remaining items
                final_start_index = total_rows - len(batch_buffer)
                flush_batch(start_index=final_start_index)

            # 3.5 구분 컬럼 기준 정렬 (SPAM 상단, HAM 하단)
            self._sort_sheet_by_gubun(ws, gubun_col_idx, reason_col_idx)
            
            # 3.6 서식 적용
            self._apply_formatting(ws, headers)

            # 4. Create Dedup Sheet (After all rows processed)
            self._create_dedup_sheet(wb, unique_urls, unique_short_urls)

            # 6. Create Blocklist Sheet
            self._create_blocklist_sheet(wb, blocklist_data)

            # 5. Save
            wb.save(output_path)
            return {"success": True, "total_rows": total_rows}
            
        except Exception as e:
            logger.error(f"Error processing Excel: {e}")
            raise e

    def process_kisa_txt(self, file_path: str, output_dir: str, processing_function, progress_callback=None, batch_size: int = 10, original_filename: str = None, manager=None, client_id: str = None, is_trap: bool = False, override_output_path: str = None, index_offset: int = 0, global_total_rows: int = 0):
        """
        Process KISA format TXT file: [Body] <TAB> [URL]
        """
        try:
            if override_output_path:
                output_path = override_output_path
                output_filename = os.path.basename(output_path)
            else:
                # 1. Parse Input Filename to determine Output Filename
                # kisa_20260103_A_result_hamMsg_url.txt -> MMSC스팸추출_20260103_A.xlsx
                input_filename = original_filename if original_filename else os.path.basename(file_path)
                
                # Try to extract 'yyyymmdd_A' pattern
                match = re.search(r'(?:kisa_|trap_)(\d{8}_[A-Za-z0-9]+)', input_filename)
                
                if match:
                    extracted_part = match.group(1)
                else:
                    # Fallback: Just date
                    date_match = re.search(r'\d{8}', input_filename)
                    if date_match:
                        extracted_part = date_match.group(0)
                    else:
                        extracted_part = datetime.now().strftime("%Y%m%d")
                
                base_filename = f"MMSC스팸추출_{extracted_part}"
                ext = ".xlsx"
                output_filename = f"{base_filename}{ext}"
                output_path = os.path.join(output_dir, output_filename)
                
                # Check for duplicates/locks and increment counter
                counter = 1
                while os.path.exists(output_path):
                    output_filename = f"{base_filename} ({counter}){ext}"
                    output_path = os.path.join(output_dir, output_filename)
                    counter += 1
                
                os.makedirs(output_dir, exist_ok=True)
                
            # 2. Read Text File with Encoding Detection (UTF-8 prior to CP949)
            raw_data = b""
            try:
                with open(file_path, 'rb') as f:
                    raw_data = f.read()
            except Exception as e:
                logger.error(f"Failed to read file for encoding detection: {e}")
                raise e

            encodings_to_try = ['utf-8', 'cp949', 'euc-kr', 'latin1']
            
            # Remove duplicates while preserving order
            encodings_to_try = list(dict.fromkeys(encodings_to_try))

            lines = []
            success_encoding = None

            for enc in encodings_to_try:
                try:
                    logger.debug(f"Attempting decoding with {enc}...")
                    decoded_text = raw_data.decode(enc)
                    lines = decoded_text.splitlines()
                    success_encoding = enc
                    logger.info(f"Successfully decoded using {enc}")
                    break
                except UnicodeDecodeError:
                    # logger.debug(f"Failed decoding with {enc}")
                    continue
            
            if not success_encoding:
                # Last resort: CP949 with replace (KISA standard)
                logger.warning("All decoding attempts failed. Fallback to CP949 (replace).")
                decoded_text = raw_data.decode('cp949', errors='replace')
                lines = decoded_text.splitlines()
                success_encoding = 'cp949-replace'

            logger.info(f"Processed KISA text file using {success_encoding} encoding. Total lines: {len(lines)}")

            rows = []
            for line in lines:
                line = line.strip()
                if not line: continue
                parts = line.split('\t')
                    
                msg_body = parts[0] if len(parts) > 0 else ""
                # URL is normally in parts[1] if exists
                url_in_file = parts[1] if len(parts) > 1 else ""
                
                rows.append({"message": msg_body, "url": url_in_file, "original_line": line})

            total_rows = len(rows)
            # 전체 메시지 큐 로드: ≤2000이면 전체, >2000이면 2000개씩 청크
            effective_batch_size = min(2000, total_rows) if total_rows > 0 else 1
            logger.info(f"Processing {total_rows} rows from KISA TXT (Batch Size: {effective_batch_size})...")

            # 3. Load or Create Template Excel & Setup target sheet
            if os.path.exists(output_path):
                wb = load_workbook(output_path)
            else:
                self.create_template_workbook(output_path)
                wb = load_workbook(output_path)
                
            target_sheet_name = "TRAP.육안분석(시뮬결과35_150)" if is_trap else "육안분석(시뮬결과35_150)"
            if target_sheet_name not in wb.sheetnames:
                wb.create_sheet(target_sheet_name)
            ws = wb[target_sheet_name]
            
            # 헤더는 템플릿 생성 시 이미 적용되었거나, 없으면 나중에 _apply_formatting 시 추가될 수 있지만,
            # 안전하게 기존 1행이 비어 있을 때만 추가합니다.
            if ws.max_row <= 1:
                headers = ["메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이", "Probability", "Semantic Class", "Reason", "Red Group"]
                if not ws.cell(row=1, column=1).value:
                    for c_idx, h_val in enumerate(headers, start=1):
                        ws.cell(row=1, column=c_idx, value=h_val)
                    header_font = Font(bold=True, size=10)
                    header_align = Alignment(horizontal='center', vertical='center')
                    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.fill = header_fill

            sim_sheet_name = "TRAP.시뮬결과전체" if is_trap else "시뮬결과전체"
            if sim_sheet_name not in wb.sheetnames:
                wb.create_sheet(sim_sheet_name)
            sim_ws = wb[sim_sheet_name]
            
            # "시뮬결과전체" 시트는 입력 텍스트를 로드한 것과 동일해야 함 (표준 서식: 메시지, URL, 구분, 메시지길이, URL길이)
            if sim_ws.max_row <= 1 and not sim_ws.cell(row=1, column=1).value:
                sim_headers = ["메시지", "URL", "구분", "메시지길이", "URL길이"]
                for c_idx, h_val in enumerate(sim_headers, start=1):
                    sim_ws.cell(row=1, column=c_idx, value=h_val)
                for cell in sim_ws[1]:
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                
                # 열 너비 조정 (샘플 서식 반영)
                sim_ws.column_dimensions['A'].width = 100  # 메시지
                sim_ws.column_dimensions['B'].width = 50   # URL
                sim_ws.column_dimensions['C'].width = 10   # 구분
                sim_ws.column_dimensions['D'].width = 15   # 메시지길이
                sim_ws.column_dimensions['E'].width = 15   # URL길이

            start_row = sim_ws.max_row + 1 if sim_ws.cell(row=1, column=1).value else 1
            for idx, r_dict in enumerate(rows):
                msg = r_dict["message"]
                url = r_dict.get("url", "")
                
                msg_cell = sim_ws.cell(row=start_row + idx, column=1, value=self._sanitize_cell_value(msg))
                msg_cell.alignment = Alignment(vertical='center', wrap_text=False)
                
                sim_ws.cell(row=start_row + idx, column=2, value=self._sanitize_cell_value(url))
                # column 3 (구분) -> 빈칸
                sim_ws.cell(row=start_row + idx, column=4, value=self._lenb(msg))
                sim_ws.cell(row=start_row + idx, column=5, value=self._lenb(url))

            # 4. Batch Processing
            batch_buffer = []
            unique_urls = {} # URL Reduplication Store
            unique_short_urls = {} # Short URL Reduplication Store
            blocklist_data = [] # IBSE Blocklist Store
            stats = {"spam_count": 0}

            def flush_batch(start_idx):
                if not batch_buffer:
                    return

                # ✅ 취소 확인
                if manager and client_id and manager.is_cancelled(client_id):
                    from ..main import CancellationException
                    raise CancellationException("Processing cancelled by user")
                
                messages = [item["message"] for item in batch_buffer]
                # [Batch KISA TXT] 파일에서 파싱한 URL 전달 (본문 추출 대신 사용)
                pre_parsed_urls = [item.get("url", "") for item in batch_buffer]
                
                try:
                    # Pass start_index, pre_parsed_urls (KISA TXT), and is_trap to processing_function
                    # Use global_total_rows for smooth progress bar handling
                    _start = index_offset + start_idx
                    _total = global_total_rows if (global_total_rows is not None and global_total_rows > 0) else total_rows
                    
                    results = processing_function(messages, start_index=_start, total_count=_total, pre_parsed_urls=pre_parsed_urls, is_trap=is_trap)
                except TypeError:
                    # Fallback: 이전 시그니처 호환 (Excel 등 pre_parsed_urls 미지원 시에도 억지로라도 is_trap을 보내볼 수 있지만 안전하게 fallback은 무시)
                    try:
                         results = processing_function(messages, start_index=_start, total_count=_total, is_trap=is_trap)
                    except TypeError:
                         results = processing_function(messages, start_index=_start, total_count=_total)
                except Exception as e:
                    from ..main import CancellationException
                    if isinstance(e, CancellationException):
                        raise  # [Phase 4] 취소 예외는 상위로 재전파 (강제 종료 처리)
                    logger.error(f"Batch Processing Failed: {e}")
                    results = [{"is_spam": None, "reason": f"Error: {e}"} for _ in messages]

                # Populate Excel Rows
                for i, result in enumerate(results):
                    # [User Request] Skip Excel Write but SHOW in UI Report
                    if result.get("exclude_from_excel"):
                        logger.debug(f"Row {start_idx + i + 1}: Skipped from Excel (exclude_from_excel=True), but showing in UI")
                        continue


                    original_data = batch_buffer[i]
                    msg_val = original_data["message"]
                    url_val = original_data["url"]
                    
                    # Logic
                    is_spam = result.get("is_spam")
                    semantic_class = result.get("semantic_class", "")
                    is_type_b = str(semantic_class).startswith("Type_B")
                    
                    reason_val = result.get("reason", "")
                    is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(reason_val)
                    
                    is_red_group = bool(result.get("red_group"))
                    if is_separated or is_red_group:
                        # 텍스트 HAM + 악성 URL 분리 감지 이거나 Red Group 인 경우
                        # 사용자 요청: 구분 열은 "빈칸"으로 유지하되 통계 개수에는 합산
                        gubun_val = ""
                        stats["spam_count"] += 1
                        raw_code = str(result.get("classification_code", ""))
                        match = re.search(r'\d+', raw_code)
                        code_val = match.group(0) if match else raw_code
                    elif is_type_b or is_spam is True:
                        # 일반 Type B 및 Type A는 구분 "o" 통일
                        gubun_val = "o"
                        stats["spam_count"] += 1
                        raw_code = str(result.get("classification_code", ""))
                        match = re.search(r'\d+', raw_code)
                        code_val = match.group(0) if match else raw_code
                    elif is_type_b or is_spam is True:
                        # 일반 Type B 및 Type A는 구분 "o" 통일
                        gubun_val = "o"
                        stats["spam_count"] += 1
                        raw_code = str(result.get("classification_code", ""))
                        match = re.search(r'\d+', raw_code)
                        code_val = match.group(0) if match else raw_code
                    else:
                        gubun_val = ""
                        code_val = ""

                    extracted_url_code = ""
                    if result.get("malicious_url_extracted"):
                        raw_ext_code = str(result.get("url_spam_code", ""))
                        m_ext = re.search(r'\d+', raw_ext_code)
                        extracted_url_code = m_ext.group(0) if m_ext else raw_ext_code
                        
                    # Probability (e.g. 98%)
                    prob_float = result.get("spam_probability", 0.0)
                    prob_val = f"{int(prob_float * 100)}%"
                    
                    semantic_val = result.get("semantic_class", "")
                    reason_val = result.get("reason", "")
                    
                    # Lengths
                    msg_len = self._lenb(msg_val)
                    url_len = self._lenb(url_val)
                    
                    # drop_url Flag Check
                    if result.get("drop_url"):
                        url_val = ""
                        url_len = 0
                    
                    is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(reason_val)
                    if is_separated and not url_val:
                        # [User Request] 분홍색 분리감지 케이스만 수동으로 URL 컬럼에 강제 오버라이딩
                        url_val = result.get("message_extracted_url", "")
                        url_len = self._lenb(url_val)
                        
                    # [최종 확인] 엑셀에 기록될 최종 url_val이 Path나 Query 없는 단독 도메인이라면 무조건 비우기 (User Request)
                    if url_val:
                        import urllib.parse
                        filtered_urls = []
                        for u in url_val.split(","):
                            u = u.strip()
                            if not u: continue
                            test_u = u if "://" in u else "http://" + u
                            try:
                                parsed_u = urllib.parse.urlparse(test_u)
                                # path가 없거나 "/"뿐이고, query도 없으면 단독 도메인 -> 제외
                                if (not parsed_u.path or parsed_u.path == "/") and not parsed_u.query:
                                    # [FIX] 명백히 스팸 혐의(Red Group, Spam 증거)로 선정된 도메인이면 보존
                                    if not (result.get("is_spam") or result.get("red_group") or result.get("malicious_url_extracted")):
                                        continue
                                # 본문 추출 유도를 위해 파손된 단축 URL로 판단된 형태(괄호, 별표 등 포함) 배제 
                                # (한글은 합법적인 커스텀 URL 슬러그일 수 있으므로 허용)
                                if bool(re.search(r'[\[\]\*\(\)\{\}\<\>]', parsed_u.path)):
                                    continue
                                filtered_urls.append(u)
                            except Exception:
                                filtered_urls.append(u)
                        
                        url_val = ", ".join(filtered_urls)
                        url_len = self._lenb(url_val)
                    
                    # Write Row
                    ws.append([
                        self._sanitize_cell_value(msg_val), 
                        self._sanitize_cell_value(url_val), 
                        self._sanitize_cell_value(gubun_val), 
                        self._to_code_int(code_val), 
                        msg_len, 
                        url_len, 
                        self._sanitize_cell_value(prob_val),
                        self._sanitize_cell_value(semantic_val),
                        self._sanitize_cell_value(reason_val),
                        self._sanitize_cell_value("O" if result.get("red_group") else "")
                    ])
                    
                    if str(code_val) == "3":
                        finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                        if msg_val:
                            finance_ws.append([self._sanitize_cell_value(msg_val)])
                    
                    # --- URL Collection Logic ---
                    # Only collect URLs from SPAM messages or extracted from HAM
                    # drop_url이 True인 경우 (위장 URL, 가비지 URL 등) 중복제거 시트에서도 완벽히 배제
                    if (result.get("is_spam") is True or result.get("malicious_url_extracted") or result.get("red_group")) and not result.get("drop_url"):
                        target_url_raw = url_val
                        # --- [UI vs KISA Export Separation] ---
                        # 도메인 난독화로 추출된 URL은 엑셀(UI 표시용)에는 로깅하되,
                        # KISA 전송용 URL 텍스트파일(차단기 IP 연동용)에서는 제외해야 함
                        is_obfuscated = "[FP Sentinel Override] 도메인 난독화" in str(result.get("reason", ""))
                        
                        if target_url_raw and not is_obfuscated:
                            for raw_u in target_url_raw.split(","):
                                target_url = raw_u.strip().rstrip('.,;:!?)}"\'')
                                if not target_url: continue
                                
                                # Additional Safety
                                if not re.search(r'[^\x00-\x7F]', target_url): # If pure ASCII (simple check) => Good
                                     pass 
                                else:
                                     pass
        
                                raw_url_code = str(result.get("classification_code", ""))
                                _m = re.search(r'\d+', raw_url_code)
                                url_dedup_code = _m.group(0) if _m else raw_url_code
                                
                                # [FIX 1] Red Group일 경우 UI가 직접 `malicious_url_extracted`를 제어하지 못하더라도 원본 코드 유지
                                if not result.get("is_spam") and not result.get("red_group"):
                                    url_dedup_code = extracted_url_code
                                    
                                if not self.is_short_url(target_url):
                                     # 40바이트 제한 반영
                                     if target_url not in unique_urls and self._lenb(target_url) <= 40:
                                         unique_urls[target_url] = {
                                             "len": self._lenb(target_url),
                                             "code": url_dedup_code,
                                             "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                         }
                                else:
                                     # 40바이트 제한 반영
                                     if target_url not in unique_short_urls and self._lenb(target_url) <= 40:
                                         unique_short_urls[target_url] = {
                                             "len": self._lenb(target_url),
                                             "code": url_dedup_code,
                                             "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                         }


                    # 1) AI가 명시적으로 ibse_signature를 추출한 경우에만 컬렉션에 추가
                    # (일반 스팸 본문을 억지로 잘라서 시그니처로 만드는 Fallback 제거)
                    
                    if result.get("is_spam") or result.get("red_group"):
                        if result.get("ibse_signature") and str(result.get("ibse_signature")).strip().lower() not in ["none", "unextractable"]:
                            clean_sig = str(result.get("ibse_signature")).replace(" ", "").replace("\n", "").replace("\r", "")
                        
                            raw_ibse_code = str(result.get("classification_code", ""))
                            m_ibse = re.search(r'\d+', raw_ibse_code)
                            ibse_code = m_ibse.group(0) if m_ibse else raw_ibse_code
                            
                            sig_len_raw = result.get("ibse_len")
                            sig_len = int(sig_len_raw) if sig_len_raw is not None else self._lenb(clean_sig)
                            
                            # 데드존(21~38바이트) 및 최소 글자(9바이트 미만) 제외 룰 유지
                            if 20 < sig_len < 39 or sig_len < 9:
                                pass
                            else:
                                blocklist_data.append({
                                    "msg": re.sub(r'[ \t\r\n\f\v]+', '', msg_val),
                                    "sig": clean_sig,
                                    "len": sig_len,
                                    "code": ibse_code
                                })

                try:
                    wb.save(output_path)
                except Exception as e:
                    logger.error(f"Auto-save error: {e}")

                batch_buffer.clear()

            # Iterate
            for i, row_data in enumerate(rows):
                batch_buffer.append(row_data)
                
                if len(batch_buffer) >= effective_batch_size:
                    flush_batch(i - effective_batch_size + 1)
            
            # Remaining
            if batch_buffer:
                flush_batch(total_rows - len(batch_buffer))
            
            # 4.5 타입 및 자동 필터 연계용 그룹핑 정렬 (Type A -> Type B 서브분류 -> HAM 순서로 모아주기)
            self._sort_sheet_by_type(ws, headers)
            
            # 4.6 서식 적용
            self._apply_formatting(ws, headers)
            
            # 5. Create Dedup Sheet
            dedup_sheet_name = "TRAP.URL중복 제거" if is_trap else "URL중복 제거"
            self._create_dedup_sheet(wb, unique_urls, unique_short_urls, sheet_name=dedup_sheet_name)
            
            # 6. Create Blocklist Sheet
            blocklist_sheet_name = "TRAP.문자문장차단등록" if is_trap else "문자문장차단등록"
            self._create_blocklist_sheet(wb, blocklist_data, sheet_name=blocklist_sheet_name)
            
            sheet_name_str = "TRAP.문자열 중복제거" if is_trap else "문자열중복제거"
            sheet_name_sen = "TRAP.문장 중복제거" if is_trap else "문장중복제거"
            str_cnt, sen_cnt = self._create_split_dedup_sheets(wb, blocklist_data, sheet_name_str, sheet_name_sen)
            
            # 7. Update Summary Table
            url_cnt = len(unique_urls) + len(unique_short_urls)
            actual_spam_cnt = stats["spam_count"]
            self._update_summary_table(wb, is_trap, output_filename, actual_spam_cnt, url_cnt, str_cnt, sen_cnt)
                
            wb.save(output_path)
            # return the actual path used so we can reuse it
            return {"success": True, "output_path": output_path, "filename": output_filename, "total_rows": total_rows}

        except Exception as e:
            logger.error(f"Error processing KISA TXT: {e}")
            raise e

    def _update_summary_table(self, wb: Workbook, is_trap: bool, filename: str, spam_cnt: int, url_cnt: int, str_cnt: int, sen_cnt: int):
        """
        Updates the summary statistics table on 'TRAP.문장 중복제거' sheet.
        """
        target_sheet = "TRAP.문장 중복제거"
        if target_sheet not in wb.sheetnames:
            ws = wb.create_sheet(target_sheet)
        else:
            ws = wb[target_sheet]
            
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        # 헤더가 비어있으면 초기 뼈대 세팅
        if not ws["E2"].value:
            # 1) 헤더 및 날짜명칭
            match = re.search(r'(\d{4})(\d{2})(\d{2})_([A-Za-z0-9]+)', filename)
            if match:
                m, d, group = match.group(2), match.group(3), match.group(4)
                try:
                    dt = datetime(int(match.group(1)), int(m), int(d))
                    weekdays = ["월", "화", "수", "목", "금", "토", "일"]
                    wd = weekdays[dt.weekday()]
                    date_str = f"● {m}월 {d}일 ({wd})_{group}"
                except ValueError:
                    date_str = f"● {m}월 {d}일_{group}"
            else:
                date_str = "● 요약"
            
            ws["E1"] = date_str
            ws["E1"].font = Font(bold=True, size=10)
            
            headers = ["구분", "스팸태깅", "URL", "문자열", "문장"]
            for col_idx, h in enumerate(headers, start=5): # E=5
                cell = ws.cell(row=2, column=col_idx, value=h)
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            ws["E3"] = "SPAM"
            ws["E4"] = "TRAP"
            ws["E3"].font = Font(bold=True, size=10)
            ws["E4"].font = Font(bold=True, size=10)
            ws["E3"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws["E4"].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            ws["E3"].alignment = Alignment(horizontal='center', vertical='center')
            ws["E4"].alignment = Alignment(horizontal='center', vertical='center')
            
            # 셀 초기화 (빈값은 0으로 표시, 정렬: 중앙 — 샘플 기준)
            for r in [3, 4]:
                for c in range(6, 10):
                    ws.cell(row=r, column=c, value=0).alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 업데이트 (SPAM 이면 3행, TRAP 이면 4행)
        row_idx = 4 if is_trap else 3
        # 기존 값이 "숫자"인 경우 덮어쓰지 않고 최신값 유지
        if ws.cell(row=row_idx, column=6).value in (0, "-", None):
            ws.cell(row=row_idx, column=6, value=spam_cnt).alignment = Alignment(horizontal='center', vertical='center')
        if ws.cell(row=row_idx, column=7).value in (0, "-", None):
            ws.cell(row=row_idx, column=7, value=url_cnt).alignment = Alignment(horizontal='center', vertical='center')
        if ws.cell(row=row_idx, column=8).value in (0, "-", None):
            ws.cell(row=row_idx, column=8, value=str_cnt).alignment = Alignment(horizontal='center', vertical='center')
        if ws.cell(row=row_idx, column=9).value in (0, "-", None):
            ws.cell(row=row_idx, column=9, value=sen_cnt).alignment = Alignment(horizontal='center', vertical='center')
        
        # 테두리 적용 (E2:I4)
        for r in range(2, 5):
            for c in range(5, 10):
                ws.cell(row=r, column=c).border = thin_border

        # 너비 조정 (샘플 기준)
        ws.column_dimensions['E'].width = 12.6
        ws.column_dimensions['F'].width = 10.6
        ws.column_dimensions['G'].width = 10.6
        ws.column_dimensions['H'].width = 10.6
        ws.column_dimensions['I'].width = 10.6

    def _create_split_dedup_sheets(self, wb: Workbook, blocklist_data: list, sheet_name_str: str, sheet_name_sen: str):
        if sheet_name_str not in wb.sheetnames:
            ws_str = wb.create_sheet(sheet_name_str)
        else:
            ws_str = wb[sheet_name_str]

        if sheet_name_sen not in wb.sheetnames:
            ws_sen = wb.create_sheet(sheet_name_sen)
        else:
            ws_sen = wb[sheet_name_sen]

        # 헤더 설정 (KISA/TRAP별 다른 헤더명, create_template_workbook 중복 방지)
        if not ws_str["A1"].value:
            is_trap_sheets = sheet_name_str.startswith("TRAP.")
            if is_trap_sheets:
                headers_str = ["TRAP 문자열(중복제거)", "길이", "분류"]
                headers_sen = ["TRAP 문장(중복제거)", "길이", "분류"]
            else:
                headers_str = ["문자열(중복제거)", "길이", "분류"]
                headers_sen = ["문장(중복제거)", "길이", "분류"]
            for col_idx in range(3):
                cell_str = ws_str.cell(row=1, column=col_idx + 1, value=headers_str[col_idx])
                cell_sen = ws_sen.cell(row=1, column=col_idx + 1, value=headers_sen[col_idx])
                for c in [cell_str, cell_sen]:
                    c.font = Font(bold=True, size=10)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Deduplicate
        unique_str = {}
        unique_sen = {}

        for item in blocklist_data:
            sig = item['sig']
            if not sig: continue
            code = item['code']
            length = item['len']
            
            if length <= 20:
                if sig not in unique_str:
                    unique_str[sig] = {"len": length, "code": code}
            else:
                if sig not in unique_sen:
                    unique_sen[sig] = {"len": length, "code": code}

        # [Fix] 데이터 행 폰트 및 세로 중앙정렬 적용
        _data_font  = Font(name='맑은 고딕', size=10.5)
        _data_align = Alignment(vertical='center')

        # 문자열(20바이트 이하): 길이가 작은 순서로 오름차순, 그다음 가나다순
        sorted_str = sorted(unique_str.items(), key=lambda x: (x[1].get('len') or 0, str(x[0])))
        start_row_str = ws_str.max_row + 1 if ws_str.cell(row=1,column=1).value else 1
        for row_idx, (sig, info) in enumerate(sorted_str, start=start_row_str):
            for col, val in [(1, self._sanitize_cell_value(sig)), (2, info['len']), (3, self._to_code_int(info['code']))]:
                cell = ws_str.cell(row=row_idx, column=col, value=val)
                cell.font  = _data_font
                cell.alignment = _data_align

        # 문장열(21바이트 이상): 길이가 긴 순서로 내림차순, 그다음 가나다순
        sorted_sen = sorted(unique_sen.items(), key=lambda x: (-(x[1].get('len') or 0), str(x[0])))
        start_row_sen = ws_sen.max_row + 1 if ws_sen.cell(row=1,column=1).value else 1
        for row_idx, (sig, info) in enumerate(sorted_sen, start=start_row_sen):
            for col, val in [(1, self._sanitize_cell_value(sig)), (2, info['len']), (3, self._to_code_int(info['code']))]:
                cell = ws_sen.cell(row=row_idx, column=col, value=val)
                cell.font  = _data_font
                cell.alignment = _data_align

        # 컬럼 너비 (샘플 기준, 문자열/문장 시트별 차별화)
        is_trap_sheets = sheet_name_str.startswith("TRAP.")
        if is_trap_sheets:
            ws_str.column_dimensions['A'].width = 25.6
            ws_str.column_dimensions['B'].width = 10.6
            ws_sen.column_dimensions['A'].width = 40.6
            ws_sen.column_dimensions['B'].width = 10.6
        else:
            ws_str.column_dimensions['A'].width = 25.9
            ws_sen.column_dimensions['A'].width = 46.1

        return len(unique_str), len(unique_sen)

    def generate_excel_from_json(self, logs: list, output_path: str, is_trap: bool, original_filename: str = None) -> dict:
        """
        Re-generate the Excel file entirely from the UI's JSON state (logs).
        Splits data automatically into KISA and TRAP sheets based on individual log item's 'is_trap' flag.
        """
        # 1. 템플릿 생성 및 로딩 (14개 기본 시트 보존)
        self.create_template_workbook(output_path)
        wb = load_workbook(output_path)
        
        logs_kisa = [l for l in logs if not l.get("is_trap")]
        logs_trap = [l for l in logs if l.get("is_trap")]
        
        if logs_kisa:
            self._populate_workbook_with_logs(wb, logs_kisa, False, original_filename)
            
        if logs_trap:
            self._populate_workbook_with_logs(wb, logs_trap, True, original_filename)
        else:
            # TRAP 데이터가 없을 때도 TRAP 시트 서식/통계 테이블을 완성
            # (create_template_workbook이 껍데기만 만들고, 후처리는 _populate_workbook_with_logs에서만 호출되기 때문)
            self._create_dedup_sheet(wb, {}, {}, sheet_name="TRAP.URL중복 제거")
            self._create_blocklist_sheet(wb, [], sheet_name="TRAP.문자문장차단등록")
            self._create_split_dedup_sheets(wb, [], "TRAP.문자열 중복제거", "TRAP.문장 중복제거")
            self._update_summary_table(wb, is_trap=True, filename=original_filename or "generated.xlsx", spam_cnt=0, url_cnt=0, str_cnt=0, sen_cnt=0)
            
        wb.save(output_path)
        return {"success": True, "output_path": output_path, "filename": original_filename, "total_rows": len(logs)}

    def _populate_workbook_with_logs(self, wb, logs: list, is_trap: bool, original_filename: str = None):
        # 2. 메인 시트 접근
        main_sheet_name = "TRAP.육안분석(시뮬결과35_150)" if is_trap else "육안분석(시뮬결과35_150)"
        ws = wb[main_sheet_name]
        
        sim_sheet_name = "TRAP.시뮬결과전체" if is_trap else "시뮬결과전체"
        sim_ws = wb[sim_sheet_name]
        
        # 육안분석 시트 동적 헤더 처리
        headers = [cell.value for cell in ws[1]]
        def get_col_idx(name, default_idx):
            try:
                return headers.index(name) + 1
            except ValueError:
                ws.cell(row=1, column=default_idx, value=name)
                headers.append(name)
                return default_idx
                
        # 기본 템플릿: "메시지", "URL", "구분", "분류", "메시지 길이", "URL 길이", "Probability", "Semantic Class", "Reason", "Red Group"
        in_token_col_idx = get_col_idx("In_Token", len(headers) + 1)
        out_token_col_idx = get_col_idx("Out_Token", len(headers) + 1)
        
        # 시뮬결과전체 시트 초기화 (템플릿이 안 만들었을 수도 있음)
        if sim_ws.max_row <= 1 and not sim_ws.cell(row=1, column=1).value:
            sim_headers = ["메시지", "URL", "구분", "메시지길이", "URL길이"]
            for c_idx, h_val in enumerate(sim_headers, start=1):
                sim_ws.cell(row=1, column=c_idx, value=h_val)
            for cell in sim_ws[1]:
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            sim_ws.column_dimensions['A'].width = 100
            sim_ws.column_dimensions['B'].width = 50
            sim_ws.column_dimensions['C'].width = 10
            sim_ws.column_dimensions['D'].width = 15
            sim_ws.column_dimensions['E'].width = 15

        unique_urls = {}
        unique_short_urls = {}
        blocklist_data = []
        stats = {"spam_count": 0}
        
        sim_start_row = sim_ws.max_row + 1 if sim_ws.cell(row=1, column=1).value else 1
        ws_start_row = 2
        
        for idx, log_item in enumerate(logs):
            result = log_item.get("result", {})
            msg_val = log_item.get("message", "") # Fallback
            if not msg_val:
                msg_val = result.get("message", "")
                
            # 원본 URL (시뮬결과전체 기록용)
            req_obj = log_item.get("request", {})
            original_url = req_obj.get("url", "")
            
            # --- 시뮬결과전체 시트 기록 ---
            msg_cell = sim_ws.cell(row=sim_start_row + idx, column=1, value=self._sanitize_cell_value(msg_val))
            msg_cell.alignment = Alignment(vertical='center', wrap_text=False)
            sim_ws.cell(row=sim_start_row + idx, column=2, value=self._sanitize_cell_value(original_url))
            sim_ws.cell(row=sim_start_row + idx, column=4, value=self._lenb(msg_val))
            sim_ws.cell(row=sim_start_row + idx, column=5, value=self._lenb(original_url))
            
            if result.get("exclude_from_excel"):
                continue
            req_obj = log_item.get("request", {})
            input_url_val = req_obj.get("url", "")
            
            # [사용자 요청] 기본적으로 무조건 입력 파일의 원본 URL을 사용한다.
            url_val = input_url_val
            
            is_red_group = bool(result.get("red_group"))
            is_spam = result.get("is_spam")
            semantic_class = result.get("semantic_class", "")
            reason_val = result.get("reason", "")
            
            is_type_b = str(semantic_class).startswith("Type_B")
            is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(reason_val)
            
            if result.get("drop_url"):
                url_val = ""
                
            if is_separated and not url_val:
                url_val = result.get("message_extracted_url", "")
                
            # [최종 확인] 엑셀에 기록될 최종 url_val이 Path나 Query 없는 단독 도메인이라면 무조건 비우기 (User Request)
            if url_val:
                import urllib.parse
                filtered_urls = []
                for u in url_val.split(","):
                    u = u.strip()
                    if not u: continue
                    test_u = u if "://" in u else "http://" + u
                    try:
                        parsed_u = urllib.parse.urlparse(test_u)
                        if (not parsed_u.path or parsed_u.path == "/") and not parsed_u.query:
                            if not (result.get("is_spam") or result.get("red_group") or result.get("malicious_url_extracted")):
                                continue
                        import re
                        if bool(re.search(r'[\[\]\*\(\)\{\}\<\>]', parsed_u.path)):
                            continue
                        filtered_urls.append(u)
                    except Exception:
                        filtered_urls.append(u)
                
                url_val = ", ".join(filtered_urls)
                
            if is_separated or is_red_group:
                # 사용자 요청: 열은 비우고 카운트는 올림
                gubun_val = ""
                stats["spam_count"] += 1
                raw_code = str(result.get("classification_code", ""))
                import re
                match = re.search(r'\d+', raw_code)
                code_val = match.group(0) if match else raw_code
            elif is_type_b or is_spam is True:
                gubun_val = "o"
                stats["spam_count"] += 1
                raw_code = str(result.get("classification_code", ""))
                import re
                match = re.search(r'\d+', raw_code)
                code_val = match.group(0) if match else raw_code
            elif is_type_b or is_spam is True:
                gubun_val = "o"
                stats["spam_count"] += 1
                raw_code = str(result.get("classification_code", ""))
                import re
                match = re.search(r'\d+', raw_code)
                code_val = match.group(0) if match else raw_code
            else:
                gubun_val = ""
                code_val = ""

            extracted_url_code = ""
            if result.get("malicious_url_extracted"):
                raw_ext_code = str(result.get("url_spam_code", ""))
                m_ext = re.search(r'\d+', raw_ext_code)
                extracted_url_code = m_ext.group(0) if m_ext else raw_ext_code

            prob_float = result.get("spam_probability", 0.0)
            prob_val = f"{int(float(prob_float) * 100)}%"
            
            # Lengths
            msg_len = self._lenb(msg_val)
            url_len = self._lenb(url_val)
            
            if result.get("drop_url"):
                url_val = ""
                url_len = 0
                
            if is_separated and not url_val:
                url_val = result.get("message_extracted_url", "")
                url_len = self._lenb(url_val)
                
            # [사용자 요청 복구] 육안분석 시트의 URL 컬럼은 정상 요소(HAM)이건 스팸이건 모두 원본 입력 그대로 유지합니다.

            in_token_val = result.get("input_tokens", 0)
            out_token_val = result.get("output_tokens", 0)

            # Write Row to 육안분석
            ws.cell(row=ws_start_row, column=1, value=self._sanitize_cell_value(msg_val))
            ws.cell(row=ws_start_row, column=2, value=self._sanitize_cell_value(url_val))
            ws.cell(row=ws_start_row, column=3, value=self._sanitize_cell_value(gubun_val))
            ws.cell(row=ws_start_row, column=4, value=self._to_code_int(code_val))
            ws.cell(row=ws_start_row, column=5, value=msg_len)
            ws.cell(row=ws_start_row, column=6, value=url_len)
            ws.cell(row=ws_start_row, column=7, value=self._sanitize_cell_value(prob_val))
            ws.cell(row=ws_start_row, column=8, value=self._sanitize_cell_value(semantic_class))
            ws.cell(row=ws_start_row, column=9, value=self._sanitize_cell_value(reason_val))
            ws.cell(row=ws_start_row, column=10, value=self._sanitize_cell_value("O" if result.get("red_group") else ""))
            ws.cell(row=ws_start_row, column=in_token_col_idx, value=in_token_val)
            ws.cell(row=ws_start_row, column=out_token_col_idx, value=out_token_val)
            ws_start_row += 1
            
            if str(code_val) == "3":
                finance_ws = wb["금융.SPAM"] if "금융.SPAM" in wb.sheetnames else wb.create_sheet("금융.SPAM")
                if msg_val:
                    finance_ws.append([self._sanitize_cell_value(msg_val)])
                    
            # --- URL Collection Logic ---
            if (result.get("is_spam") is True or result.get("malicious_url_extracted") or result.get("red_group")) and not result.get("drop_url"):
                target_url = url_val
                is_obfuscated = "[FP Sentinel Override] 도메인 난독화" in str(result.get("reason", ""))
                
                if target_url and not is_obfuscated:
                    import urllib.parse
                    import re
                    
                    for raw_u in target_url.split(","):
                        u = raw_u.strip().rstrip('.,;:!?)}"\'')
                        if not u: continue
                        
                        target_u = u
                        
                        raw_url_code = str(result.get("classification_code", ""))
                        _m = re.search(r'\d+', raw_url_code)
                        url_dedup_code = _m.group(0) if _m else raw_url_code
                        
                        # [FIX 1] Red Group일 경우 UI가 직접 `malicious_url_extracted`를 제어하지 못하더라도 원본 코드 유지
                        if not result.get("is_spam") and not result.get("red_group"):
                            url_dedup_code = extracted_url_code
                            
                        # 단일 도메인 40바이트 제한 평가 로직 적용 (여러 URL 생성 시 개별적으로 평가)
                        if not self.is_short_url(target_u):
                            if target_u not in unique_urls and self._lenb(target_u) <= 40:
                                unique_urls[target_u] = {
                                    "len": self._lenb(target_u),
                                    "code": url_dedup_code,
                                    "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                }
                        else:
                            if target_u not in unique_short_urls and self._lenb(target_u) <= 40:
                                unique_short_urls[target_u] = {
                                    "len": self._lenb(target_u),
                                    "code": url_dedup_code,
                                    "malicious_url_extracted": result.get("malicious_url_extracted", False)
                                }
                            
            # --- IBSE Collection Logic ---
            ibse_sig = result.get("ibse_signature")
            if result.get("is_spam") or result.get("red_group"):
                if ibse_sig and str(ibse_sig).strip().lower() not in ["none", "unextractable"]:
                    clean_sig = str(ibse_sig).replace(" ", "").replace("\n", "").replace("\r", "")
                
                    raw_ibse_code = str(result.get("classification_code", ""))
                    m_ibse = re.search(r'\d+', raw_ibse_code)
                    ibse_code = m_ibse.group(0) if m_ibse else raw_ibse_code
                    
                    sig_len_raw = result.get("ibse_len")
                    sig_len = int(sig_len_raw) if sig_len_raw is not None else self._lenb(clean_sig)
                    
                    if 20 < sig_len < 39 or sig_len < 9:
                        pass
                    else:
                        import re
                        blocklist_data.append({
                            "msg": re.sub(r'[ \t\r\n\f\v]+', '', msg_val),
                            "sig": clean_sig,
                            "len": sig_len,
                            "code": ibse_code
                        })

        # Sort & Format
        self._sort_sheet_by_type(ws, headers)
        self._apply_formatting(ws, headers)
        
        # Create sheets (these will just update existing ones now!)
        dedup_sheet_name = "TRAP.URL중복 제거" if is_trap else "URL중복 제거"
        self._create_dedup_sheet(wb, unique_urls, unique_short_urls, sheet_name=dedup_sheet_name)
        
        blocklist_sheet_name = "TRAP.문자문장차단등록" if is_trap else "문자문장차단등록"
        self._create_blocklist_sheet(wb, blocklist_data, sheet_name=blocklist_sheet_name)
        
        sheet_name_str = "TRAP.문자열 중복제거" if is_trap else "문자열중복제거"
        sheet_name_sen = "TRAP.문장 중복제거" if is_trap else "문장중복제거"
        str_cnt, sen_cnt = self._create_split_dedup_sheets(wb, blocklist_data, sheet_name_str, sheet_name_sen)
        
        url_cnt = len(unique_urls) + len(unique_short_urls)
        actual_spam_cnt = stats["spam_count"]
        self._update_summary_table(wb, is_trap, original_filename or "generated.xlsx", actual_spam_cnt, url_cnt, str_cnt, sen_cnt)
        
