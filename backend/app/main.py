from fastapi import FastAPI, UploadFile, File, HTTPException, WebSocket, WebSocketDisconnect, Form, BackgroundTasks, Body
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import shutil
import os
import asyncio
import sys
import logging
import warnings
import json
import time
import re
import unicodedata

# **CRITICAL**: 다른 앱 모듈 import 전에 로깅 설정 먼저 초기화
from dotenv import load_dotenv, set_key
load_dotenv(override=True)  # .env 파일 로드

from app.core.logging_config import setup_logging, get_logger, batch_id_context, set_log_level, set_console_enabled
from app.core.models import LLM_MODELS, CONFIG_METADATA

# 로깅 시스템 초기화 (앱 모듈 import 전에 반드시 호출!)
# 환경변수: LOG_LEVEL_CONSOLE, LOG_LEVEL_FILE, LOG_JSON_ENABLED
setup_logging()

logger = get_logger(__name__)

# Suppress noisy warnings and verbose SDK logs
warnings.filterwarnings("ignore")
logging.getLogger("google_genai").setLevel(logging.WARNING)
logging.getLogger("google.generativeai").setLevel(logging.WARNING)

# **CRITICAL FIX**: Force ProactorEventLoop on Windows for Playwright/Subprocess support
if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import uuid
from typing import List, Dict, Any

# 이제 앱 모듈들을 import (로깅 설정 완료 후)
from app.services.rule_service import RuleBasedFilter
from app.agents.content_agent.agent import ContentAnalysisAgent
from app.agents.url_agent.agent import UrlAnalysisAgent
from app.utils.excel_handler import ExcelHandler
from app.core.constants import SPAM_CODE_MAP
from app.core.llm_manager import key_manager

# Custom Exception for Cancellation
class CancellationException(Exception):
    """처리 취소 예외"""
    pass

# [Phase 4] Ctrl+C(SIGINT) 시 배치 조기 종료용 플래그 (시그널 핸들러에서 설정)
shutdown_requested = False

def _signal_handler(signum, frame):
    """SIGINT(Ctrl+C) 시 shutdown 플래그 설정 후 KeyboardInterrupt로 프로세스 종료."""
    global shutdown_requested
    shutdown_requested = True
    logger.info("Shutdown requested (Ctrl+C)")
    raise KeyboardInterrupt()  # 프로세스 종료 (기본 동작 복원)

# Windows에서 SIGINT 등록 (Ctrl+C)
try:
    import signal
    signal.signal(signal.SIGINT, _signal_handler)
except Exception as e:
    logger.warning(f"Could not register SIGINT handler: {e}")

app = FastAPI()


# Global Executor to prevent Thread Leak
global_executor = None

# WebSocket Connection Manager
class ConnectionManager:
    def __init__(self):
        # Map client_id to WebSocket connection
        self.active_connections: Dict[str, WebSocket] = {}
        # Map client_id to list of buffered messages (offline queue)
        self.message_queue: Dict[str, List[dict]] = {}
        # Cancellation flags for processing
        self.cancellation_flags: Dict[str, bool] = {}
        # HITL Events
        import threading
        self.hitl_events: Dict[str, threading.Event] = {}
        self.hitl_responses: Dict[str, dict] = {}

    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket
        logger.info(f"WebSocket Connected: Client {client_id}")
        
        # Replay buffered messages
        if client_id in self.message_queue:
            queue = self.message_queue[client_id]
            if queue:
                logger.info(f"Replaying {len(queue)} buffered messages for {client_id}")
                for msg in queue:
                    try:
                        await websocket.send_json(msg)
                    except Exception as e:
                        logger.error(f"Error replaying message: {e}")
                self.message_queue[client_id] = [] # Clear queue

    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]
            logger.info(f"WebSocket Disconnected: Client {client_id}")

    # HITL Methods (Thread-Safe Event Handling)
    def register_hitl_request(self, client_id: str):
        import threading
        event = threading.Event()
        self.hitl_events[client_id] = event
        return event

    def resolve_hitl_request(self, client_id: str, response_data: dict):
        if client_id in self.hitl_events:
            self.hitl_responses[client_id] = response_data
            self.hitl_events[client_id].set()

    def get_hitl_response(self, client_id: str) -> dict:
        return self.hitl_responses.get(client_id, {})

    def cleanup_hitl(self, client_id: str):
        self.hitl_events.pop(client_id, None)
        self.hitl_responses.pop(client_id, None)

    async def send_personal_message(self, message: dict, client_id: str):
        if client_id in self.active_connections:
            try:
                # logger.info(f"Sending to {client_id}: {message}") 
                await self.active_connections[client_id].send_json(message)
            except Exception as e:
                logger.warning(f"Failed to send message to active client {client_id}: {e}")
                # Fallback to buffer if send fails
                self._buffer_message(client_id, message)
        else:
            # Client offline, buffer message
            self._buffer_message(client_id, message)

    def _buffer_message(self, client_id: str, message: dict):
        if client_id not in self.message_queue:
            self.message_queue[client_id] = []
        
        self.message_queue[client_id].append(message)
    
    # Cancellation Methods
    def request_cancellation(self, client_id: str):
        """클라이언트의 처리 취소 요청"""
        self.cancellation_flags[client_id] = True
        logger.info(f"Cancellation requested for client {client_id}")
    
    def is_cancelled(self, client_id: str) -> bool:
        """취소 요청 확인"""
        return self.cancellation_flags.get(client_id, False)
    
    def clear_cancellation(self, client_id: str):
        """취소 플래그 초기화"""
        self.cancellation_flags.pop(client_id, None)

manager = ConnectionManager()


def _log_env_config():
    """서버 기동 시 .env 설정값 출력 (API 키 제외)"""
    # 민감 정보 패턴 - 이 키들은 값 출력 생략
    SENSITIVE_PATTERNS = ("KEY", "SECRET", "PASSWORD", "TOKEN")
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    if not os.path.exists(env_path):
        return
    try:
        lines = []
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" in line:
                    key, _, val = line.partition("=")
                    key = key.strip()
                    val = val.strip().strip('"\'')
                    # 민감 키 스킵
                    key_upper = key.upper()
                    if any(p in key_upper for p in SENSITIVE_PATTERNS):
                        lines.append(f"  {key}=***")
                    else:
                        lines.append(f"  {key}={val}")
        if lines:
            logger.info("📋 [Config] .env loaded:\n" + "\n".join(lines))
    except Exception as e:
        logger.warning(f"Could not read .env for config log: {e}")


@app.on_event("startup")
async def startup_event():
    # [Config] .env 값 출력 (API 키 제외)
    _log_env_config()

    # [Performance Fix] Initialize Global ThreadPoolExecutor ONCE
    # Preventing thread leak (4320 threads issue)
    global global_executor
    workers = int(os.getenv("MAX_THREAD_WORKERS", 70))
    from concurrent.futures import ThreadPoolExecutor
    global_executor = ThreadPoolExecutor(max_workers=workers)
    
    # Set as default for the main loop
    loop = asyncio.get_event_loop()
    loop.set_default_executor(global_executor)

    # [Performance Optimization] Warm-up Heavy Components
    # Eliminate 30s+ delay on first request by initializing resources here
    logger.info("🔥 [Startup] Warming up AI Components (Vector DB & LLM)...")
    import time
    t0 = time.time()
    
    try:
        # 1. Warm-up SpamRagService (loads Chroma, OpenAIEmbeddings)
        from app.services.spam_rag_service import get_spam_rag_service
        rag_service = get_spam_rag_service()
        # Force initialization of lazy properties
        _ = rag_service._get_db() 
        logger.info(f"   -> SpamRagService Warmed up.")

        # 2. Warm-up ContentAnalysisAgent (loads Guide Chroma)
        from app.agents.content_agent.agent import ContentAnalysisAgent
        content_agent = ContentAnalysisAgent()
        # Force initialization
        _ = content_agent._get_vector_db() 
        logger.info(f"   -> ContentAnalysisAgent Warmed up.")
        
    except Exception as e:
        logger.warning(f"⚠️ [Startup] Warm-up failed (non-critical): {e}")

    logger.info(f"✅ Server Application Started! (Startup took {time.time() - t0:.2f}s) | Global ThreadPool: {workers} workers.")

@app.on_event("shutdown")
async def shutdown_event():
    logger.info("Server shutting down...")
    global global_executor
    if global_executor:
        global_executor.shutdown(wait=False)
        logger.info("Global ThreadPoolExecutor shutdown.")

@app.get("/health")
async def health_check():
    logger.info("✅ Health Check Endpoint Reached!")
    return {"status": "ok"}

# ========== 로그 레벨 런타임 변경 API ==========
from app.core.logging_config import get_log_levels, set_log_level, set_console_enabled

@app.get("/api/log-level")
async def get_current_log_level():
    """현재 로그 레벨 조회"""
    return get_log_levels()

class LogLevelChange(BaseModel):
    target: str  # "console" 또는 "file"
    level: str   # "DEBUG", "INFO", "WARNING", "ERROR"

class ConsoleToggle(BaseModel):
    enabled: bool  # True=ON, False=OFF

@app.post("/api/log-level")
async def change_log_level(request: LogLevelChange):
    """런타임에 로그 레벨 변경 (서버 재시작 없이)"""
    result = set_log_level(request.target, request.level)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    return result

@app.post("/api/log-console")
async def toggle_console_log(request: ConsoleToggle):
    """콘솔 로그 출력 ON/OFF"""
    return set_console_enabled(request.enabled)

# ========== 런타임 환경설정 관리 API ==========

@app.get("/api/models")
async def get_supported_models():
    """지원하는 LLM 모델 리스트 조회"""
    return LLM_MODELS

@app.get("/api/config")
async def get_current_config():
    """현재 시스템 설정값 조회 (마스킹 처리)"""
    config_values = {}
    SENSITIVE_KEYS = ["KEY", "SECRET", "PASSWORD", "TOKEN"]
    
    for item in CONFIG_METADATA:
        key = item["key"]
        val = os.getenv(key, "")
        
        # 민감 정보 마스킹
        if any(sk in key.upper() for sk in SENSITIVE_KEYS) and val:
            config_values[key] = "***"
        else:
            config_values[key] = val
            
    return {
        "metadata": CONFIG_METADATA,
        "values": config_values
    }

class ConfigUpdate(BaseModel):
    settings: Dict[str, Any]

@app.post("/api/config")
async def update_config(request: ConfigUpdate):
    """런타임에 설정값 변경 및 .env 저장"""
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    updated_keys = []
    
    try:
        for key, value in request.settings.items():
            # 1. 메모리(os.environ) 업데이트
            str_val = str(value)
            os.environ[key] = str_val
            
            # 2. .env 파일 영구 저장 (민감정보 제외 - 민감정보는 별도 처리 권장되나 여기서는 일단 저장 지원)
            # 유효성 검증 로직 추가 가능
            set_key(env_path, key, str_val)
            updated_keys.append(key)
            
        logger.info(f"⚙️ [Config] Updated keys: {', '.join(updated_keys)}")
        
        # 3. 로그 레벨 관련 설정 즉시 반영
        from app.core.logging_config import set_log_level, set_console_enabled
        if "LOG_LEVEL_CONSOLE" in updated_keys:
            set_log_level("console", os.environ["LOG_LEVEL_CONSOLE"])
        if "LOG_LEVEL_FILE" in updated_keys:
            set_log_level("file", os.environ["LOG_LEVEL_FILE"])
        if "LOG_CONSOLE_ENABLED" in updated_keys:
            set_console_enabled(os.environ["LOG_CONSOLE_ENABLED"] == "1")
        
        # 4. 필터 임계값 관련 설정 즉시 반영
        if "MIN_MESSAGE_LENGTH" in updated_keys or "ALPHANUMERIC_OBFUSCATION_RATIO_THRESHOLD" in updated_keys:
            rule_filter.update_thresholds()
        
        # 5. 변경 사항 즉시 로그 출력 (마스킹 적용)
        _log_env_config()
        
        return {"success": True, "updated": updated_keys}
    except Exception as e:
        logger.error(f"Config update failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/quota-status")
async def get_quota_status():
    """LLM 공급자별 Quota Exhausted 상태 조회 (설정 UI용)"""
    from app.core.llm_manager import key_manager
    return {"success": True, "quota_status": key_manager.get_quota_status()}


class ResetQuotaRequest(BaseModel):
    provider: str | None = None  # None이면 전체 리셋


@app.post("/api/config/reset-quota")
async def reset_quota(request: ResetQuotaRequest | None = Body(default=None)):
    """Quota Exhausted 플래그 수동 리셋 (설정 UI 버튼용)"""
    from app.core.llm_manager import key_manager
    provider = request.provider if request else None
    result = key_manager.reset_quota_exhausted(provider)
    return {"success": True, **result}


class SetKeyIndexRequest(BaseModel):
    indices: dict[str, int]  # { "GEMINI": 1, "OPENAI": 0, ... }


@app.post("/api/config/set-key-index")
async def set_key_index(request: SetKeyIndexRequest = Body(...)):
    """특정 LLM 공급자의 사용 키 인덱스를 수동 지정 (설정 UI용)"""
    from app.core.llm_manager import key_manager
    results = {}
    for provider, idx in request.indices.items():
        success = key_manager.set_current_index(provider, idx)
        results[provider] = {"success": success, "index": idx}
        
    return {"success": True, "results": results}


# ========== Spam RAG API (Reference Examples) ==========
from app.services.spam_rag_service import get_spam_rag_service

class SpamRagCreate(BaseModel):
    message: str
    label: str = "SPAM"
    code: str
    category: str
    reason: str

class SpamRagUpdate(BaseModel):
    message: str = None
    label: str = None
    code: str = None
    category: str = None
    reason: str = None

@app.get("/api/spam-rag")
async def get_spam_rag_examples():
    """참조 예시 전체 조회"""
    try:
        service = get_spam_rag_service()
        examples = service.get_all_examples()
        return {"success": True, "data": examples, "total": len(examples)}
    except Exception as e:
        logger.error(f"Spam RAG GET Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/spam-rag/stats")
async def get_spam_rag_stats():
    """참조 예시 통계 조회"""
    try:
        service = get_spam_rag_service()
        stats = service.get_stats()
        return {"success": True, "data": stats}
    except Exception as e:
        logger.error(f"Spam RAG Stats Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/spam-rag/search")
async def search_spam_rag_examples(query: str, k: int = 3):
    """유사 참조 예시 검색 (Intent-based, Threshold-filtered)"""
    try:
        # 1. Generate Intent Summary for the query (to align with stored vectors)
        from app.agents.content_agent.agent import ContentAnalysisAgent
        agent = ContentAnalysisAgent()
        intent_summary = await agent.agenerate_intent_summary(query)
        
        logger.info(f"RAG Search Query: '{query}' -> Intent: '{intent_summary}'")

        # 2. Search using the Intent Summary
        service = get_spam_rag_service()
        results = service.search_similar(intent_summary, k=k)
        
        # 3. Filter by RAG_DISTANCE_THRESHOLD (align with LLM prompt injection logic)
        # 3-small models tend to have higher cosine distances than ada-002.
        distance_threshold = float(os.getenv("RAG_DISTANCE_THRESHOLD", "0.95"))
        hits = results.get("hits", [])
        filtered_hits = [
            hit for hit in hits 
            if hit.get("distance", 999) <= distance_threshold
        ]
        
        logger.info(f"RAG Search: {len(hits)} raw results, {len(filtered_hits)} within threshold ({distance_threshold})")
        
        return {
            "success": True, 
            "data": {"hits": filtered_hits, "stats": results.get("stats", {})}, 
            "total": len(filtered_hits)
        }
    except Exception as e:
        # Check if the error is due to Quota Exhaustion
        error_msg = str(e)
        if "quota" in error_msg.lower() or "429" in error_msg.lower() or "QuotaExhaustedNoRetryError" in error_msg:
            logger.warning(f"Spam RAG Search Quote Exhausted during intent generation: {e}")
            # Instead of failing with 500, return an empty result set gracefully for the UI
            return {
                "success": True, 
                "data": {"hits": [], "stats": {}}, 
                "total": 0,
                "warning": "LLM Quota Exhausted: Could not generate intent for search."
            }
        
        logger.error(f"Spam RAG Search Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/spam-rag/{example_id}")
async def get_spam_rag_example(example_id: str):
    """특정 참조 예시 조회"""
    try:
        service = get_spam_rag_service()
        example = service.get_example_by_id(example_id)
        if not example:
            raise HTTPException(status_code=404, detail="Example not found")
        return {"success": True, "data": example}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Spam RAG GET Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/spam-rag")
async def create_spam_rag_example(example: SpamRagCreate):
    """
    Spam RAG: 새로운 참조 예시 추가 (Intent-based)
    1. ContentAnalysisAgent를 통해 의도 요약 생성
    2. Intent Summary를 임베딩 저장, 원본은 메타데이터 저장
    """
    try:
        # 1. Generate Intent Summary
        # Import inside to ensure availability or avoid circular deps if any (though unlikely here)
        from app.agents.content_agent.agent import ContentAnalysisAgent
        agent = ContentAnalysisAgent()
        intent_summary = await agent.agenerate_intent_summary(example.message)
        logger.info(f"Generated Intent Summary for RAG: '{intent_summary}' (Original: {example.message[:20]}...)")

        # 2. Save to RAG (Intent Summary based)
        service = get_spam_rag_service()
        result = service.add_example(
            intent_summary=intent_summary,  # Embedding Target
            original_message=example.message, # Metadata
            label=example.label,
            code=example.code,
            category=example.category,
            reason=example.reason
        )
        logger.info(f"Spam RAG Example Created: {result['id']}")
        return {"success": True, "data": result}
    except ValueError as e:
        # 중복 에러 처리 (409 Conflict)
        logger.warning(f"Spam RAG Create Conflict: {e}")
        raise HTTPException(status_code=409, detail=str(e))
    except Exception as e:
        logger.error(f"Spam RAG Create Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/spam-rag/{example_id}")
async def update_spam_rag_example(example_id: str, example: SpamRagUpdate):
    """참조 예시 수정"""
    try:
        service = get_spam_rag_service()
        result = service.update_example(
            example_id=example_id,
            message=example.message,
            label=example.label,
            code=example.code,
            category=example.category,
            reason=example.reason
        )
        if not result:
            raise HTTPException(status_code=404, detail="Example not found")
        logger.info(f"Spam RAG Example Updated: {example_id}")
        return {"success": True, "data": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Spam RAG Update Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/spam-rag/{example_id}")
async def delete_spam_rag_example(example_id: str):
    """참조 예시 삭제"""
    try:
        service = get_spam_rag_service()
        success = service.delete_example(example_id)
        if not success:
            raise HTTPException(status_code=404, detail="Example not found or delete failed")
        logger.info(f"Spam RAG Example Deleted: {example_id}")
        return {"success": True, "message": "Deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Spam RAG Delete Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# CORS config
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Directories
UPLOAD_DIR = "../data/uploads"
OUTPUT_DIR = "../data/outputs"
REPORTS_DIR = "../data/reports"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

# ========== Report Management API ==========
class ReportSaveRequest(BaseModel):
    report_name: str
    source_filename: str
    logs: List[dict]

@app.post("/api/reports/save")
async def save_report(request: ReportSaveRequest):
    """현재 로그를 리포트 파일로 저장"""
    import json
    from datetime import datetime
    try:
        timestamp = datetime.now().isoformat()
        report_data = {
            "report_name": request.report_name,
            "source_filename": request.source_filename,
            "timestamp": timestamp,
            "logs": request.logs
        }
        
        # 파일명 생성: report_name_timestamp.json
        safe_name = "".join([c for c in request.report_name if c.isalnum() or c in (' ', '_', '-')]).strip()
        if not safe_name: safe_name = "report"
        filename = f"{safe_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(REPORTS_DIR, filename)
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
            
        logger.info(f"Report saved: {filename}")
        return {"success": True, "filename": filename}
    except Exception as e:
        logger.error(f"Report Save Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports")
async def list_reports():
    """저장된 리포트 목록 조회"""
    try:
        if not os.path.exists(REPORTS_DIR):
            return {"success": True, "reports": []}
        files = [f for f in os.listdir(REPORTS_DIR) if f.endswith(".json")]
        reports = []
        for f in files:
             filepath = os.path.join(REPORTS_DIR, f)
             reports.append({
                 "filename": f,
                 "mtime": os.path.getmtime(filepath)
             })
        # 최신 수정순 정렬
        reports.sort(key=lambda x: x["mtime"], reverse=True)
        return {"success": True, "reports": reports}
    except Exception as e:
        logger.error(f"List Reports Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/reports/{filename}")
async def get_report(filename: str):
    """특정 리포트 내용 로드"""
    import json
    try:
        filepath = os.path.join(REPORTS_DIR, filename)
        if not os.path.exists(filepath):
            raise HTTPException(status_code=404, detail="Report not found")
            
        with open(filepath, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {"success": True, "data": data}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get Report Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Initialize Services & Agents
from app.agents.ibse_agent.service import IBSEAgentService

# Initialize Services & Agents
rule_filter = RuleBasedFilter()
rag_filter = ContentAnalysisAgent()
url_filter = UrlAnalysisAgent()
excel_handler = ExcelHandler()
ibse_service = IBSEAgentService()

def process_message(message: str) -> dict:
    """
    Orchestrates the 3-Stage Filtering Pipeline
    """
    # Stage 1: Rule-Based
    # logger.info("    [Stage 1] Rule Checking...")
    s1_result = rule_filter.check(message)
    
    # If HAM confirmed by Rule (e.g., Non-Korean message)
    if s1_result["is_spam"] is False:
        s1_code = s1_result.get("classification_code")
        s1_reason = s1_result.get("reason", "Rule-based HAM")
        logger.info(f"    -> Rule-based HAM (Code: {s1_code})")
        logger.info(f"\n[DEBUG RESULT]\nMessage: {message}\nIs Spam: False\nCode: {s1_code}\nReason: {s1_reason}\n")
        return {
            "is_spam": False,
            "classification_code": s1_code,
            "spam_probability": 0.0,
            "reason": s1_reason,
            "input_tokens": 0,
            "output_tokens": 0,
            "exclude_from_excel": s1_result.get("exclude_from_excel", False)
        }

    # Stage 2: RAG + LLM + Antigravity
    # We pass s1_result to include detected patterns in prompt
    logger.info("    [Stage 2] AI Analysis (RAG + LLM)...")
    s2_result = rag_filter.check(message, s1_result)
    
    final_is_spam = s2_result["is_spam"]
    final_code = s2_result["classification_code"]
    final_reason = s2_result.get("reason", "No reason provided")
    final_prob = s2_result.get("spam_probability", 0.0)
    
    # Extract tokens (default to 0 if missing)
    input_tokens = s2_result.get("input_tokens", 0)
    output_tokens = s2_result.get("output_tokens", 0)

    # Common return dictionary builder
    def build_result(is_spam, code, reason):
        return {
            "is_spam": is_spam,
            "classification_code": code,
            "spam_probability": final_prob,
            "reason": reason,
            "input_tokens": input_tokens,
            "output_tokens": output_tokens
        }

    # [Early Exit] Quota Exhausted/Exceeded 발생 시 후속 작업(URL, IBSE) 생략
    if "quota exhausted" in final_reason.lower() or "quota exceeded" in final_reason.lower() or "429" in final_reason:
        logger.warning(f"    -> LLM Quota Exhausted/Exceeded detected. Bypassing URL and IBSE stages.")
        return build_result(final_is_spam, final_code, final_reason)

    # Logic Update:
    # If Stage 2 is SPAM -> Finalize as SPAM (skip Stage 3)
    if final_is_spam:
        logger.info(f"    -> Identified as SPAM (Code: {final_code})")
        logger.info(f"\n[DEBUG RESULT]\nMessage: {message}\nIs Spam: {final_is_spam}\nProbability: {final_prob}\nCode: {final_code}\nReason: {final_reason}\nTokens: In={input_tokens}, Out={output_tokens}\n")
        return build_result(True, final_code, final_reason)
    
    # Step 3: URL Deep Dive (Conditional)
    # Check URL if Content is Spam OR if Content is Safe but URL exists
    import re
    # url_pattern = re.compile(r'(https?://\S+|www\.\S+|[a-zA-Z0-9-]+\.[a-zA-Z]{2,})')
    url_pattern = re.compile(r'(?:https?://|www\.)\S+|[a-zA-Z0-9\uac00-\ud7a3\u3131-\u3163-]+\.[a-zA-Z가-힣]{2,}')
    
    # 원본 메시지에서 URL 체크
    has_url = bool(url_pattern.search(message))
    
    # 난독화 디코딩된 텍스트에서도 URL 체크
    decoded_text = s1_result.get("decoded_text")
    if decoded_text and not has_url:
        has_url = bool(url_pattern.search(decoded_text))
    
    # s1에서 이미 추출한 decoded_urls가 있으면 URL 존재
    if s1_result.get("decoded_urls"):
        has_url = True
    
    if has_url:
        logger.info("    [Stage 3] URL Deep Dive...")
        try:
            isaa_result = url_filter.check(message, decoded_text=decoded_text)
                    
            url_is_spam = isaa_result.get("is_spam")
            reason_lower = isaa_result.get("reason", "").lower()
            is_inconclusive = any(x in reason_lower for x in ["error", "inconclusive", "insufficient", "image only"])
            
            if is_inconclusive or url_is_spam is None:
                 # URL 심층 분석 불가 (404, 내용 없음, 오류 등) -> 본문 분석 결과(Content Agent)를 전적으로 따름
                 logger.info("    -> URL Analysis Inconclusive/Error. Falling back to Content Agent verdict.")
                 final_reason += f" | [URL: Inaccessible/Empty - Using Content Verdict]"
            elif url_is_spam:
                 # 1차: 실질적 유해성(Harmful Intent) 여부 확인
                 url_reason = isaa_result.get("reason", "").lower()
                 url_code = isaa_result.get('classification_code')
                 
                 # 단순 문맥 불일치(Inconsistency)나 사칭(Impersonation)만으로는 HAM을 SPAM으로 뒤집지 않음.
                 # 실질적 유해성(도박, 성인, 불법, 사기 등)이 언급되거나 관련 코드가 있는 경우에만 SPAM 전환.
                 harm_keywords = ["gambling", "adult", "phishing", "malicious", "fraud", "scam", "illegal", "유해", "도박", "성인", "피싱", "사기"]
                 has_harmful_intent = any(k in url_reason for k in harm_keywords) or (url_code and str(url_code) != "0")
                 
                 if not final_is_spam and not has_harmful_intent:
                      # 본문은 HAM인데 URL은 단순 불일치/낚시성인 경우 -> HAM 유지 (의도 중심)
                      logger.info("    -> URL shows inconsistency but no harmful intent. Maintaining HAM.")
                      final_reason += " | [URL: Inconsistent but no clear harm (HAM maintained)]"
                 else:
                      # 실제 유해성이 확인되었거나 본문이 이미 SPAM인 경우 -> SPAM 확정
                      logger.info("    -> Suspicious URL with Harmful Intent Confirmed! Finalizing as SPAM.")
                      final_is_spam = True
                      final_reason += f" | [URL: DETECTED SPAM]"
                      
                      # 확률 업데이트
                      url_prob = isaa_result.get("spam_probability", 0.95)
                      if url_prob > final_prob:
                          final_prob = url_prob
                      
                      # 코드 업데이트 (본문이 HAM이었거나 코드가 없으면 URL 코드 사용, 기본값 '0')
                      if url_code and str(url_code) != "0":
                           final_code = url_code
                      elif not final_code or str(final_code).startswith("HAM"):
                           final_code = "0" 
                      
                      logger.info(f"[URL Override] Final Verdict: SPAM, Code: {final_code}")
            else:
                 # Confirmed Safe -> Force HAM (Override Content SPAM if confirmed as safe institution/service)
                 if final_is_spam:
                      logger.info("    -> URL Confirmed Safe! Overriding Content SPAM to HAM.")
                      final_is_spam = False
                      final_reason += " | [URL: CONFIRMED SAFE (Override)]"
                      final_code = None
        except Exception as e:
            error_msg = str(e).lower()
            if "quota" in error_msg or "429" in error_msg or "exhausted" in error_msg:
                logger.warning(f"    -> [Stage 3] URL Agent API Quota Exhausted! Falling back to Content Agent verdict.")
                final_reason = f"429 Quota Exhausted. Content Agent Fallback: {final_reason}"
                return build_result(final_is_spam, final_code, final_reason)
            else:
                logger.error(f"URL Deep Dive Failed: {e}")
                final_reason += f" | [URL Error: {e}]"


    logger.info("    -> Classified as HAM.")
    logger.info(f"\n[DEBUG RESULT]\nMessage: {message}\nIs Spam: {final_is_spam}\nProbability: {final_prob}\nCode: {final_code}\nReason: {final_reason}\nTokens: In={input_tokens}, Out={output_tokens}\n")
    return build_result(final_is_spam, final_code, final_reason)


@app.websocket("/ws/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    await manager.connect(websocket, client_id)
    try:
        while True:
            # Listen for client messages
            data_str = await websocket.receive_text()
            try:
                import json
                data = json.loads(data_str)
                logger.info(f"Received from {client_id}: {data}")

                if data.get("type") == "HITL_RESPONSE":
                    logger.info(f"HITL Response Received for {client_id}: {data}")
                    manager.resolve_hitl_request(client_id, data)
                
                elif data.get("type") == "CHAT_MESSAGE":
                    user_msg = data.get("content", "")
                    mode = data.get("mode", "Unified") # Default to Unified "Smart" Mode
                    
                    # 🔹 Reset token tracking purely for this chat run
                    key_manager.reset_token_usage()
                    
                    logger.info(f"\n{'='*60}\n[Chat] Mode: {mode}\n[Chat] 메시지 원문:\n  {user_msg}\n{'='*60}")
                    
                    if user_msg:
                        # Signal Start of Stream
                        await manager.send_personal_message({
                            "type": "CHAT_STREAM_START",
                            "content": ""
                        }, client_id)
                        
                        manager.clear_cancellation(client_id)

                        # Helper for Status Updates
                        async def send_status(text: str):
                            if manager.is_cancelled(client_id):
                                raise CancellationException("Chat processing cancelled by user")
                            await manager.send_personal_message({
                                "type": "PROCESS_STATUS",
                                "content": text
                            }, client_id)
                            
                        # Helper for generic text chunk
                        async def send_text_chunk(text: str):
                            if manager.is_cancelled(client_id):
                                raise CancellationException("Chat processing cancelled by user")
                            await manager.send_personal_message({
                                "type": "CHAT_STREAM_CHUNK",
                                "content": text
                            }, client_id)

                        async def run_with_cancellation(coro):
                            task = asyncio.create_task(coro)
                            async def watch_cancellation():
                                while not task.done():
                                    if manager.is_cancelled(client_id):
                                        logger.info(f"Cancellation watcher triggered for {client_id}")
                                        task.cancel()
                                        break
                                    await asyncio.sleep(0.5)
                            watcher = asyncio.create_task(watch_cancellation())
                            try:
                                return await task
                            except asyncio.CancelledError:
                                raise CancellationException("Chat processing cancelled by user")
                            finally:
                                watcher.cancel()

                        # --- MODE DISPATCHER ---
                        
                        # 1. IBSE Mode (Signature Only)
                        if mode == "IBSE":
                             await send_status("분석 준비 중...")
                             loop = asyncio.get_running_loop()
                             import re
                             spaceless_msg = re.sub(r'[ \t\r\n\f\v]+', '', user_msg)
                             
                             try:
                                 ibse_result = await run_with_cancellation(ibse_service.process_message(spaceless_msg, status_callback=send_status))
                                 
                                 sig = ibse_result.get('signature')
                                 decision = ibse_result.get('decision')
                                 reason = ibse_result.get('reason')
                                 byte_len = ibse_result.get('byte_len', 0)
                                 
                                 response_text = f"**[IBSE 시그니처 추출 결과]**\n"
                                 response_text += f"🔧 **전처리(공백제거)**: `{spaceless_msg}`\n\n"
                                 
                                 if sig:
                                     response_text += f"✅ **추출 성공**\n- **시그니처**: `{sig}`\n- **길이**: {byte_len} bytes (CP949)\n"
                                 elif decision == "unextractable":
                                     response_text += f"🛑 **추출 생략 (대상 아님)**\n- **사유**: {reason}\n"
                                 else:
                                     response_text += f"⚠️ **오류/실패** (Type: {decision})\n- **사유**: {reason}\n"

                                 await send_text_chunk(response_text)
                                 
                             except CancellationException:
                                 logger.info(f"IBSE Chat cancellation confirmed for {client_id}")
                                 await manager.send_personal_message({"type": "CHAT_STREAM_CHUNK", "content": "\n🚫 **사용자에 의해 분석이 중지되었습니다.**\n"}, client_id)
                             except Exception as e:
                                 logger.error(f"IBSE execution error: {e}")
                                 await send_text_chunk(f"⚠️ **오류 발생**: {str(e)}")

                        # 2. URL Mode (URL Only)
                        elif mode == "URL":
                            # URL 모드에서도 난독화 체크
                            s1_url = rule_filter.check(user_msg)
                            decoded_text_url = s1_url.get("decoded_text")
                            try:
                                isaa_result = await run_with_cancellation(url_filter.acheck(user_msg, status_callback=send_status, decoded_text=decoded_text_url))
                                
                                analysis_text = f"**[ISAA URL 분석 결과]**\n"
                                if isaa_result["is_spam"]:
                                    analysis_text += f"🚫 **스팸 탐지됨** ({int(isaa_result.get('spam_probability',0)*100)}%)\n"
                                    analysis_text += f"- 분류: {isaa_result.get('classification_code', 'Unknown')}\n"
                                    analysis_text += f"- 사유: {isaa_result.get('reason')}\n"
                                else:
                                    analysis_text += f"✅ **정상 URL**\n- 사유: {isaa_result.get('reason')}\n"
                                    
                                details = isaa_result.get("details", {})
                                analysis_text += f"\n**[상세 분석 정보]**\n"
                                analysis_text += f"- URL 경로: {details.get('extracted_url', 'N/A')} → {details.get('final_url', 'N/A')}\n"
                                analysis_text += f"- 팝업/캡차: {details.get('popup_count', 0)}개 / {'있음' if details.get('captcha_detected') else '없음'}\n"

                                await send_text_chunk(analysis_text)
                            except CancellationException:
                                logger.info(f"URL Chat cancellation confirmed for {client_id}")
                                await manager.send_personal_message({"type": "CHAT_STREAM_CHUNK", "content": "\n🚫 **사용자에 의해 분석이 중지되었습니다.**\n"}, client_id)
                            except Exception as e:
                                error_msg = str(e).lower()
                                if "quota" in error_msg or "429" in error_msg or "exhausted" in error_msg:
                                    await send_text_chunk("\n⚠️ **오류**: API Quota 초과(429)로 인해 분석 불가.\n")
                                else:
                                    await send_text_chunk(f"\n⚠️ **URL 분석 오류**: {e}\n")

                        # 3. TEXT Mode (Content Only) - Isolated!
                        elif mode == "TEXT":
                            try:
                                # Use new async check with callbacks
                                s1 = rule_filter.check(user_msg) # Stage 1 is fast/local
                                
                                # Rule-based HAM (e.g., Non-Korean message)
                                if s1.get("is_spam") is False:
                                    s1_code = s1.get("classification_code", "HAM-5")
                                    s1_reason = s1.get("reason", "Rule-based HAM")
                                    code_map = SPAM_CODE_MAP
                                    code_desc = code_map.get(s1_code, "외국어 메시지")
                                    
                                    msg_text = f"✅ **정상 문자** - {s1_code}. {code_desc}\n- 사유: {s1_reason}\n"
                                    await send_text_chunk(msg_text)
                                    
                                    # Signal End of Stream
                                    await manager.send_personal_message({
                                        "type": "CHAT_STREAM_END",
                                        "content": "",
                                        "token_usage": key_manager.get_token_usage()
                                    }, client_id)
                                    continue
                                
                                # Calls Content Agent asynchronously
                                s2_result = await run_with_cancellation(rag_filter.acheck(user_msg, s1, status_callback=send_status))
                                
                                # Format & Send Result
                                final_is_spam = s2_result.get("is_spam")
                                reason = s2_result.get("reason")
                                prob = s2_result.get("spam_probability", 0)
                                code = s2_result.get("classification_code", "Unk")
                                
                                code_map = SPAM_CODE_MAP
                                import re
                                msg_text = ""
                                
                                if final_is_spam:
                                    match = re.search(r'\d+', str(code))
                                    raw_code = match.group(0) if match else str(code)
                                    code_desc = code_map.get(raw_code, "기타")
                                    msg_text = f"🚫 **스팸 의심** ({int(prob*100)}%) - {raw_code}. {code_desc}\n- 사유: {reason}\n"
                                elif final_is_spam is None:
                                    msg_text = f"⚠️ **판단 보류 (HITL)**\n- 사유: {reason}\n"
                                else:
                                    msg_text = f"✅ **정상 문자**\n- 사유: {reason}\n"
                                    
                                await send_text_chunk(msg_text)
                                # NOTE: No Auto-IBSE or URL check here. Completely Isolated.
                            except CancellationException:
                                logger.info(f"TEXT Chat cancellation confirmed for {client_id}")
                                await manager.send_personal_message({"type": "CHAT_STREAM_CHUNK", "content": "\n🚫 **사용자에 의해 분석이 중지되었습니다.**\n"}, client_id)
                            except Exception as text_err:
                                logger.error(f"TEXT Mode Error: {text_err}")
                                await manager.send_personal_message({"type": "CHAT_STREAM_CHUNK", "content": f"\n⚠️ **분석 중 오류 발생**: {text_err}\n"}, client_id)

                        # 4. Unified / Smart Mode (Default)
                        else:
                            await send_status("🔍 통합 분석 시작...")
                            
                            # Step A: Rule-based Check (Stage 1)
                            s1 = rule_filter.check(user_msg)
                            
                            # Rule-based HAM (e.g., Non-Korean message) - Early Exit
                            if s1.get("is_spam") is False:
                                s1_code = s1.get("classification_code", "HAM-5")
                                s1_reason = s1.get("reason", "Rule-based HAM")
                                code_map = SPAM_CODE_MAP
                                code_desc = code_map.get(s1_code, "외국어 메시지")
                                
                                msg_text = f"✅ **정상 문자** - {s1_code}. {code_desc}\n- 사유: {s1_reason}\n\n"
                                await send_text_chunk(msg_text)
                                
                                await manager.send_personal_message({
                                    "type": "CHAT_STREAM_END", 
                                    "content": "",
                                    "token_usage": key_manager.get_token_usage()
                                }, client_id)
                                continue

                            # Step B: LangGraph Execution (Stage 2+)
                            from app.graphs.batch_flow import create_batch_graph
                            # Use one-off graph for chat
                            cv_graph = create_batch_graph(rag_filter, url_filter, ibse_service)
                            
                            # Prepare Input State
                            input_state = {
                                "message": user_msg,
                                "s1_result": s1,
                                "prefetched_context": None, # Chat mode does its own RAG retrieval
                                "content_result": None,
                                "url_result": None,
                                "ibse_result": None,
                                "final_result": None,
                                "status_callback": send_status
                            }
                            
                            try:
                                # Invoke Graph
                                graph_output = await run_with_cancellation(cv_graph.ainvoke(input_state))
                                final_res = graph_output.get("final_result", {})
                                
                                final_is_spam = final_res.get("is_spam")
                                reason = final_res.get("reason", "No reason provided")
                                prob = final_res.get("spam_probability", 0.0)
                                code = final_res.get("classification_code")
                                learning_label = final_res.get("learning_label", "N/A")
                                
                                code_map = SPAM_CODE_MAP
                                import re
                                
                                # 📡 [Display Logic]
                                msg_text = ""
                                if final_is_spam:
                                    match = re.search(r'\d+', str(code))
                                    raw_code = match.group(0) if match else str(code)
                                    code_desc = code_map.get(raw_code, "기타")
                                    
                                    is_fp_sensitive = (final_res.get('semantic_class') == 'Type_B') or ('[FP Sentinel Override]' in reason)
                                    
                                    if is_fp_sensitive:
                                        msg_text = f"🟠 **FP SENSITIVE** ({int(prob*100)}%) - {raw_code}. {code_desc}\n"
                                    else:
                                        msg_text = f"🚫 **스팸 확정** ({int(prob*100)}%) - {raw_code}. {code_desc}\n"
                                    msg_text += f"- **사유**: {reason}\n"
                                    
                                    # 학습 보호 라벨 안내
                                    if learning_label == "HAM":
                                        msg_text += f"💡 **학습 보호**: 해당 메시지는 정상 토큰 오염 방지를 위해 학습 데이터에서 제외(HAM 처리)되었습니다.\n"
                                elif final_is_spam is None:
                                    msg_text = f"⚠️ **판단 보류 (HITL)**\n- **사유**: {reason}\n"
                                else:
                                    if "Override" in reason or "무죄 추정" in reason:
                                        msg_text = f"🛡️ **정상 문자 (시스템 오탐 방어 발동 - Override)**\n- **사유**: {reason}\n"
                                    else:
                                        msg_text = f"✅ **정상 문자**\n- **사유**: {reason}\n"
                                
                                await send_text_chunk(msg_text)
                                
                                # Step B-1 (NEW): Append URL Data if extracted
                                url_res = graph_output.get("url_result")
                                if url_res:
                                    url_is_spam = url_res.get("is_spam")
                                    url_reason = url_res.get("reason", "")
                                    url_details = url_res.get("details", {})
                                    url_text = f"\n**[ISAA URL 분석 결과]**\n"
                                    if url_is_spam:
                                        url_text += f"🚫 **스팸 URL 탐지됨**\n- 사유: {url_reason}\n"
                                    else:
                                        url_text += f"✅ **정상 URL**\n- 사유: {url_reason}\n"
                                    attempted_urls = url_details.get("attempted_urls", [])
                                    if attempted_urls and len(attempted_urls) > 1:
                                        url_list_str = ", ".join(attempted_urls)
                                        url_text += f"- 접속 최종 URL: {url_details.get('extracted_url', 'N/A')} ==> URL: {url_list_str}\n"
                                    else:
                                        url_text += f"- 접속 최종 URL: {url_details.get('final_url', 'N/A')}\n"
                                    await send_text_chunk(url_text)
                                
                                # Step B-2: Append IBSE Data if extracted
                                ibse_sig = final_res.get("ibse_signature")
                                ibse_len = final_res.get("ibse_len")
                                if ibse_sig:
                                    ibse_text = f"\n**[IBSE 시그니처 추출 결과]**\n- **시그니처**: `{ibse_sig}`\n- **길이**: {ibse_len} bytes (CP949)\n"
                                    await send_text_chunk(ibse_text)

                                # Step C: Post-processing Summary (RAG-based)
                                if final_res.get("url_result") or final_is_spam:
                                    await send_text_chunk("\n---\n**📝 종합 의견**\n\n")
                                    url_res = graph_output.get("url_result")
                                    summary = await run_with_cancellation(rag_filter.generate_final_summary(user_msg, final_res, url_res))
                                    await send_text_chunk(summary)

                                # Send final success status
                                await send_status("🎉 통합 분석 완료")

                            except CancellationException:
                                logger.info(f"Unified Chat cancellation confirmed for {client_id}")
                                await manager.send_personal_message({"type": "CHAT_STREAM_CHUNK", "content": "\n🚫 **사용자에 의해 분석이 중지되었습니다.**\n"}, client_id)
                            except Exception as graph_err:
                                logger.error(f"Chat Graph Execution Error: {graph_err}")
                                await send_text_chunk(f"\n⚠️ **시스템 분석 오류**: {graph_err}\n")
                        
                        # Signal End of Stream
                        await manager.send_personal_message({
                            "type": "CHAT_STREAM_END",
                            "content": "",
                            "token_usage": key_manager.get_token_usage()
                        }, client_id)

            except Exception as e:
                logger.error(f"Error parsing WS message: {e}")

    except WebSocketDisconnect:
        manager.disconnect(client_id)

@app.post("/cancel/{client_id}")
async def cancel_processing(client_id: str):
    """파일 처리 취소 요청"""
    manager.request_cancellation(client_id)
       
    logger.info(f"Cancellation requested for client {client_id}")
    return {"message": f"Cancellation requested for {client_id}"}


    ws.cell(row=excel_row, column=output_columns["Reason"]).value = sanitize(reason)

# ==========================================
# [Helper] URL Detection for Smart Concurrency
# ==========================================
def has_potential_url(message: str) -> bool:
    """
    메시지에 URL이 포함되어 있는지 확인 (난독화된 URL 포함)
    Smart Concurrency: URL이 있으면 Browser Queue(Slow), 없으면 LLM Queue(Fast)로 배정하기 위한 판단
    """
    if not message:
        return False
        
    # nodes.py와 동일한 강력한 정규식 (http(s) 옵션 + 도메인 패턴)
    # 한글, 영문, 숫자, 특수문자 도메인 지원
    url_pattern = r'(?:http[s]?://)?(?:[a-zA-Z0-9\uac00-\ud7a3\u3131-\u3163-]+\.)+[a-zA-Z가-힣]{2,}'
    
    # 1. 원본 텍스트 검사
    if re.search(url_pattern, message):
        return True
        
    # 2. 난독화 해제 후 검사 (NFKC 정규화)
    # 예: "lⓔtⓩ.kr" -> "letz.kr"
    try:
        normalized = unicodedata.normalize('NFKC', message)
        if normalized != message:
            if re.search(url_pattern, normalized):
                # logger.debug(f"De-obfuscated URL detected: {message} -> {normalized}")
                return True
    except:
        pass
        
    return False

@app.post("/upload")
async def upload_file(client_id: str = Form(...), files: List[UploadFile] = File(...)):
    logger.info(f"DEBUG: Receive file upload request from Client {client_id}: {[f.filename for f in files]}")
    
    # 🔹 Reset token tracking at the start of a batch session to accumulate correctly across chunks
    from app.core.llm_manager import key_manager
    if hasattr(key_manager, 'reset_token_usage'):
        key_manager.reset_token_usage()

    # [Safe Cleanup] Zombie process killing via 'taskkill' removed.
    # New architecture uses localized PlaywrightManager with auto-cleanup (finally block).
    # This prevents accidental termination of user's local Chrome browser.
    # loop = asyncio.get_running_loop()
    # await loop.run_in_executor(None, kill_zombie_processes)

    try:
        # Clear any previous cancellation flags
        manager.clear_cancellation(client_id)
        
        # Group and determine files
        kisa_file = None
        trap_file = None
        excel_files = []
        
        for f in files:
            name_lower = f.filename.lower()
            if "trap" in name_lower and name_lower.endswith('.txt'):
                trap_file = f
            elif name_lower.endswith('.txt'):
                # General txt or explicit kisa
                kisa_file = f
            elif name_lower.endswith('.xlsx'):
                excel_files.append(f)
        
        # If no explicit kisa_file but multiple TXT, fallback
        if not trap_file and not kisa_file and not excel_files:
            raise HTTPException(status_code=400, detail="No suitable files uploaded")

        # 1. Base Output Naming (Based on Kisa file or first file)
        base_file = kisa_file or trap_file or (excel_files[0] if excel_files else files[0])
        original_name = os.path.splitext(base_file.filename)[0]
        
        # Check if it looks like kisa_YYYYMMDD_A...
        import re
        from datetime import datetime
        match = re.search(r'(?:kisa_|trap_)(\d{8}_[A-Za-z0-9]+)', original_name, re.IGNORECASE)
        if match:
            extracted_part = match.group(1)
        else:
            date_match = re.search(r'\d{8}', original_name)
            extracted_part = date_match.group(0) if date_match else datetime.now().strftime("%Y%m%d")
            
        base_filename = f"MMSC스팸추출_{extracted_part}"
        final_filename = f"{base_filename}.xlsx"
        output_path = os.path.join(OUTPUT_DIR, final_filename)
        
        # Handle duplicate output names
        counter = 1
        while os.path.exists(output_path):
            final_filename = f"{base_filename} ({counter}).xlsx"
            output_path = os.path.join(OUTPUT_DIR, final_filename)
            counter += 1
            
        # Define Progress Callback (Thread-safe)
        loop = asyncio.get_running_loop()
        
        def progress_callback(data: dict):
            asyncio.run_coroutine_threadsafe(
                manager.send_personal_message(data, client_id), loop
            )

        # Wrapper for process_message to inject HITL logic (Batch Compatible)
        def process_message_with_hitl(messages: list, start_index: int = 0, total_count: int = 0, pre_parsed_urls: list = None, is_trap: bool = False) -> list:
            """
            Processes a batch of messages.
            1. Rule Filter (Stage 1) - Individual
            2. RAG Filter (Stage 2) - Batch
            3. URL Filter (Stage 3) - Parallel for SPAM items
            4. HITL Check - Individual (Blocking)
            
            pre_parsed_urls: KISA TXT 파일에서 탭으로 파싱한 URL 목록 (있으면 본문 추출 대신 사용)
            """
            
            import asyncio
            from app.graphs.batch_flow import clear_signature_cache
            
            # [NEW] 이전 배치들의 시그니처 런타임 캐시 완전 소각 (오염 전이 눈덩이 현상 방지)
            clear_signature_cache()
            
            # Step 1: Stage 1 (Rule) for all
            s1_results = []
            for msg in messages:
                if not msg:
                    s1_results.append({"is_spam": False, "detected_pattern": "Empty", "exclude_from_excel": True})
                else:
                    s1_results.append(rule_filter.check(msg))
            
            # Async Batch Orchestrator
            async def run_batch_pipeline():
                # [Performance Fix] Use Global ThreadPoolExecutor
                # We do NOT create a new executor here.
                # loop.set_default_executor is NOT called. 
                # The global executor (size=70) handles all blocking I/O.
                try:

                    
                    
                    # [JIT Optimization] Global pre-processing removed to eliminate startup delay.
                    # Context will be prepared individually within process_single_item.
                    
                    # A. Define Single Item Processing Function (Wraps Content + URL Logic per item)
                    from app.graphs.batch_flow import create_batch_graph
                    from app.agents.url_agent.tools import PlaywrightManager
                    
                    # [Infrastructure] Create Local PlaywrightManager for this batch to ensure isolation
                    # This fixes race conditions where global manager was closed by other threads.
                    local_manager = PlaywrightManager() 

                    # Use global agents (Thread-safe/Async-safe assumed)
                    batch_graph = create_batch_graph(rag_filter, url_filter, ibse_service, playwright_manager=local_manager)

                    async def process_single_item(index, message, s1_res, url_from_file: str = "", from_kisa_txt: bool = False):
                        import time
                        start_time = time.time() # [Time Tracking] Start

                        # Set Batch ID for this context (automatically prefixes all logs in this task)
                        batch_id_context.set(f"Batch {index+1}")
                        
                        # Rule-based HAM (e.g., Non-Korean message, short msgs) - Skip Graph completely
                        if s1_res.get("is_spam") is False:
                            s1_code = s1_res.get("classification_code")
                            s1_reason = s1_res.get("reason", "Rule-based HAM")
                            logger.info(f"Rule-based HAM | code={s1_code}")
                            duration = round(time.time() - start_time, 2)
                            return index, {
                                "is_spam": False,
                                "classification_code": s1_code,
                                "spam_probability": 0.0,
                                "reason": s1_reason,
                                "duration_seconds": duration,
                                "exclude_from_excel": s1_res.get("exclude_from_excel", False)
                            }
                            
                        # Construct context Just-In-Time (JIT)
                        try:
                            # Use list wrap [message] to reuse batch logic for single item
                            jit_contexts = await rag_filter.prepare_batch_contexts([message])
                            context_data = jit_contexts[0] if jit_contexts else None
                        except Exception as e:
                            logger.error(f"Context Prep Error: {e}")
                            context_data = None

                        # 배치 Item 시작 로그
                        logger.info(f"분석 시작 | msg={message[:80]}{'...' if len(message) > 80 else ''}")
                        
                        # Construct Input State
                        input_state = {
                            "message": message,
                            "s1_result": s1_res,
                            "prefetched_context": context_data, # [Batch Optimization] Inject Context
                            "pre_parsed_url": url_from_file.strip() if url_from_file else None,  # KISA TXT 파싱 URL (본문 추출 대신 사용)
                            "pre_parsed_only_mode": False,  # [User Requirement] KISA TXT라도 URL 필드가 없으면 본문에서 추출해서 스크래핑해야 함. (단 중복제거 시트엔 안들어감)
                            "content_result": None,
                            "url_result": None,
                            "ibse_result": None,
                            "final_result": None
                        }
                        
                        try:
                            # Invoke Graph
                            # ainvoke is async
                            graph_output = await batch_graph.ainvoke(input_state)
                            final_res = graph_output.get("final_result", {})
                            
                            # Logging
                            final_is_spam = final_res.get("is_spam")
                            final_code = final_res.get("classification_code")
                            final_prob = final_res.get("spam_probability", 0.0)
                            verdict = "SPAM" if final_is_spam else ("HITL" if final_is_spam is None else "HAM")
                            
                            duration = round(time.time() - start_time, 2)
                            logger.info(f"완료 | {verdict} | code={final_code} | prob={final_prob} | {duration}s")
                            
                            # [Batch Mode] generate_final_summary 비활성화 - 1782회 LLM 호출 제거, 처리 시간 단축
                            # Excel/UI에는 reason, classification_code 등 구조화 결과만 사용됨
                            
                            # Add duration to result
                            final_res["duration_seconds"] = duration
                            return index, final_res
                            
                        except Exception as e:
                            logger.error(f"Graph Execution Error: {e}")
                            duration = round(time.time() - start_time, 2)
                            return index, {"is_spam": None, "reason": f"Graph Error: {e}", "duration_seconds": duration}

                    # Create Two Priority Queues (Semaphores)
                    # 1. Browser Queue: For messages with URLs (High Resource) -> Limited by MAX_BROWSER_CONCURRENCY
                    # 2. LLM Queue: For text-only messages (Low Resource) -> Limited by LLM_BATCH_SIZE
                    
                    max_browser = int(os.getenv("MAX_BROWSER_CONCURRENCY", 10))
                    max_llm = int(os.getenv("LLM_BATCH_SIZE", 50))
                    
                    sem_browser = asyncio.Semaphore(max_browser)
                    sem_llm = asyncio.Semaphore(max_llm)
                    
                    logger.info(f"Initialized Dual Concurrency: Browser={max_browser}, LLM-Only={max_llm}")
                    
                    # Monotonic counter for progress
                    completed_count = 0
                    
                    # pre_parsed_urls: KISA TXT에서 오면 사용, Excel이면 None
                    url_list = pre_parsed_urls if pre_parsed_urls else [""] * len(messages)
                    is_kisa_txt = pre_parsed_urls is not None  # KISA TXT면 URL 없을 때 본문 추출 스킵
                    def _url_at(i): return url_list[i] if i < len(url_list) else ""
                    
                    async def sem_task(index, msg, s1):
                        # Set Batch ID for this context task
                        batch_id_context.set(f"Batch {index+1}")
                        if manager.is_cancelled(client_id):
                            logger.info(f"Cancelled before start.")
                            return index, {"is_spam": None, "reason": "Cancelled"}
                            
                        from app.core.llm_manager import key_manager
                        provider = os.getenv("LLM_PROVIDER", "OPENAI").upper()
                        if key_manager.is_quota_exhausted(provider):
                            logger.info(f"Batch {index+1} skipped due to global Quota Exhaustion.")
                            return index, {"is_spam": None, "reason": f"Skipped (Global {provider} Quota Exhausted)"}

                            
                        # Smart Concurrency: Check for URL (본문 추출 또는 KISA TXT 파싱 URL)
                        is_url_msg = has_potential_url(msg) or bool(_url_at(index).strip())
                        selected_sem = sem_browser if is_url_msg else sem_llm
                        queue_type = "Browser" if is_url_msg else "LLM-Only"
                        
                        # Terminology: 'Queued' means created and waiting for worker slot
                        logger.debug(f"Queued in {queue_type} Queue (Waiting for semaphore...)")
                        
                        # [Phase 2] 세마포어(동시 대기열) 줄서기 시작
                        # 실행 타임아웃(300초) 적용 (태스크 단위 제한)
                        task_timeout = int(os.getenv("BATCH_TASK_TIMEOUT_SEC", "300"))
                        
                        await selected_sem.acquire()
                        
                        try:
                            # [NEW Phase 3] 세마포어 통과 직후 Jitter 부여 (Throttling)
                            # 대량의 태스크가 동시에 자리를 배정받았을 때, 정확히 같은 시간(ms)에 
                            # 수십 개의 HTTP 요청이 집중(Burst)되어 구글 API에서 커넥션이 블랙홀/Drop 
                            # 되는 현상을 막기 위해 0.1초 ~ 1.5초 사이의 무작위 난수를 쉬게 합니다.
                            import random
                            jitter = random.uniform(0.1, 1.5)
                            await asyncio.sleep(jitter)
                            if manager.is_cancelled(client_id):
                                logger.info(f"Cancelled after semaphore acquisition.")
                                return index, {"is_spam": None, "reason": "Cancelled"}
                                
                            if key_manager.is_quota_exhausted(provider):
                                logger.info(f"Batch {index+1} aborted after queueing due to Quota Exhaustion.")
                                return index, {"is_spam": None, "reason": f"Aborted (Global {provider} Quota Exhausted)"}
                                
                            logger.debug(f"Acquired {queue_type} semaphore. Starting process...")
                            # [Phase 4] Per-task timeout: 실행 시간만 별도 제한 (세마포어 대기와 분리)
                            try:
                                idx, res = await asyncio.wait_for(process_single_item(index, msg, s1, _url_at(index), is_kisa_txt), timeout=task_timeout)
                            except asyncio.TimeoutError:
                                logger.warning(f"Task {index+1} timed out after {task_timeout}s")
                                return index, {"is_spam": None, "reason": f"Timeout ({task_timeout}s)"}
                            # [Real-time Streaming] Send result to client immediately
                            try:
                                nonlocal completed_count
                                completed_count += 1
                                
                                # [Optimization] Strip large binary/b64 data that frontend doesn't need
                                ws_res = res.copy()
                                if "screenshot_b64" in ws_res:
                                    del ws_res["screenshot_b64"]
                                
                                # [UI Fix] Use run_coroutine_threadsafe to send to main event loop
                                asyncio.run_coroutine_threadsafe(
                                    manager.send_personal_message({
                                        "type": "BATCH_PROCESS_UPDATE",
                                        "index": idx + start_index,  # Use Global Index for updating specific row
                                        "message": msg,
                                        "status": "done",
                                        "result": ws_res,
                                        "current": start_index + completed_count, # Monotonic progress
                                        "total": total_count, # Total rows in file
                                        "is_trap": is_trap,
                                        "token_usage": getattr(key_manager, 'get_token_usage', lambda: {})()
                                    }, client_id), loop
                                )
                            except Exception as ws_ex:
                                logger.warning(f"WS Streaming Failed: {ws_ex}")
                            return idx, res
                        finally:
                            selected_sem.release()

                    # Create tasks (create_task으로 취소 가능하게)
                    tasks = [asyncio.create_task(sem_task(i, messages[i], s1_results[i])) for i in range(len(messages))]
                    
                    # [UI Fix] Send initial state so frontend immediately knows Total count
                    try:
                        asyncio.run_coroutine_threadsafe(
                            manager.send_personal_message({
                                "type": "BATCH_PROCESS_UPDATE",
                                "index": start_index,
                                "message": "",
                                "status": "started",
                                "result": None,
                                "current": start_index,
                                "total": total_count,
                                "is_trap": is_trap
                            }, client_id), loop
                        )
                    except Exception as ws_ex:
                        logger.warning(f"Initial WS Streaming Failed: {ws_ex}")

                    # [Phase 4] 취소 모니터: 5초마다 is_cancelled 또는 Ctrl+C 확인 → 취소 시 모든 태스크 cancel
                    cancelled_by_user = False
                    async def cancel_checker():
                        nonlocal cancelled_by_user
                        while True:
                            await asyncio.sleep(5)
                            if shutdown_requested or (manager and client_id and manager.is_cancelled(client_id)):
                                cancelled_by_user = True
                                logger.info("Cancellation requested. Cancelling all batch tasks...")
                                for t in tasks:
                                    if not t.done():
                                        t.cancel()
                                return
                    
                    cancel_task = asyncio.create_task(cancel_checker())
                    try:
                        logger.info(f"Starting asyncio.gather for {len(tasks)} tasks (with cancellation monitor)...")
                        results_with_idx = await asyncio.gather(*tasks, return_exceptions=True)
                    finally:
                        cancel_task.cancel()
                        try:
                            await cancel_task
                        except asyncio.CancelledError:
                            pass
                    
                    if cancelled_by_user:
                        raise CancellationException("Processing cancelled by user")
                    
                    # return_exceptions=True이므로 CancelledError 등이 결과로 올 수 있음 → 정규화
                    normalized = []
                    for i, r in enumerate(results_with_idx):
                        if isinstance(r, BaseException):
                            logger.error(f"[Async Gather Trap] Index {i} failed: {type(r).__name__} - {r}")
                            if isinstance(r, asyncio.CancelledError):
                                normalized.append((i, {"is_spam": None, "reason": "Cancelled"}))
                            else:
                                # [User Request] 풍선효과 차단: excel_handler.py가 참조하는 기본 Schema 구성요소를 강력하게 삽입
                                err_type = type(r).__name__
                                normalized.append((i, {
                                    "is_spam": None, 
                                    "reason": f"Async Exception [{err_type}]: {str(r)}",
                                    "classification_code": "ERROR", 
                                    "spam_probability": 0.0,
                                    "input_tokens": 0,
                                    "output_tokens": 0
                                }))
                        else:
                            normalized.append(r)
                    results_with_idx = normalized
                    logger.info("asyncio.gather finished.")
                    
                    # Sort just in case (though gather guarantees order) and extract
                    # results_with_idx is list of (index, result)
                    sorted_results = sorted(results_with_idx, key=lambda x: x[0])
                    completed_results = [r[1] for r in sorted_results]
                        
                    return completed_results

                finally:
                    # Cleanup Playwright after batch loop finishes
                    try:
                        if local_manager:
                             await local_manager.stop()
                        logger.info("Local PlaywrightManager cleaned up.")
                    except Exception as cleanup_err:
                        logger.warning(f"Cleanup warning: {cleanup_err}")

            # Run Async Pipeline
            try:
                s2_results = asyncio.run(run_batch_pipeline())
            except CancellationException:
                raise  # [Phase 4] 취소는 상위로 재전파
            except Exception as e:
                logger.error(f"Batch Async Error: {e}")
                s2_results = [{"is_spam": None, "reason": f"Async Error: {e}"} for _ in messages]

            # Step 4: HITL & Finalize
            final_results = []
            for i, result in enumerate(s2_results):
                msg = messages[i]
                # Logging already done immediately above. 
                # Just HITL logic remains.
                
                # Check for HITL Condition (Code 30)
                # User Request (2026-02-06): Non-blocking notification only. No waiting.
                # If prob >= 0.9, auto-confirm SPAM. Otherwise, mark as HAM (30) with [확인 필요].
                spam_prob = result.get("spam_probability", 0.0)
                
                if result.get("classification_code") == "30":
                    if spam_prob >= 0.9:
                        logger.info(f"[HITL Override] Probability {spam_prob} >= 0.9. Marking as SPAM-2 without user check.")
                        result["is_spam"] = True
                        result["classification_code"] = "2" # Default to Catch-all/Gambling/Other Spam
                        result["reason"] += " [Auto-Confirmed due to High Probability]"
                    else:
                        logger.info(f"[HITL] Non-blocking Notification: {msg[:20]}...")
                    
                        # Notify Client (Visual Notification Only)
                        hitl_notification = {
                            "type": "HITL_NOTIFICATION", # Changed from REQUEST to NOTIFICATION
                            "message": msg,
                            "spam_probability": result.get("spam_probability"),
                            "reason": result.get("reason"),
                            "index": i + start_index # Global Index for Frontend Highlight
                        }
                        try:
                            # Send asynchronously without waiting for result blocking heavily
                            asyncio.run_coroutine_threadsafe(
                                manager.send_personal_message(hitl_notification, client_id), loop
                            )
                        except Exception as noti_err:
                            logger.warning(f"Failed to send HITL notification: {noti_err}")
                        
                        # ** Non-blocking Default Decision **
                        # Mark as HAM (is_spam=False) but keep Code 30 so user can review later.
                        result["is_spam"] = False
                        result["classification_code"] = "30" 
                        result["reason"] += " [판단 보류 - 확인 필요]"
                        logger.info("[HITL] Defaulted to HAM(30) (Non-blocking mode).")
                
                final_results.append(result)
                
            return final_results
 
        
        # Determine Batch Size
        batch_size_env = int(os.getenv("LLM_BATCH_SIZE", 10))
        batch_chunk_size = 1000 
        
        total_rows = 0
        uploaded_temp_files = []  # 임시 파일 추적용 리스트 (마지막 디스크 클린업)
        
        # If Excel file uploaded (Legacy logic bypass)
        if excel_files:
            for ex_file in excel_files:
                file_id = str(uuid.uuid4())
                file_ext = os.path.splitext(ex_file.filename)[1]
                input_path = os.path.join(UPLOAD_DIR, f"{file_id}{file_ext}")
                uploaded_temp_files.append(input_path)
                with open(input_path, "wb") as buffer:
                    shutil.copyfileobj(ex_file.file, buffer)
                    
                result = await loop.run_in_executor(
                    None, 
                    lambda path=input_path: excel_handler.process_file(path, output_path, process_message_with_hitl, progress_callback, batch_size=batch_chunk_size)
                )
                if isinstance(result, dict) and "total_rows" in result:
                    total_rows += result["total_rows"]
            output_filename = final_filename
            
        else:
            # It's KISA/TRAP txt files.
            # a. Create template 
            await loop.run_in_executor(None, excel_handler.create_template_workbook, output_path)
            
            # Base counters
            kisa_row_count = 0 
            trap_row_count = 0
            
            input_path_kisa = None
            input_path_trap = None

            # b. Pre-calculate KISA & Save
            if kisa_file:
                file_id = str(uuid.uuid4())
                input_path_kisa = os.path.join(UPLOAD_DIR, f"{file_id}.txt")
                uploaded_temp_files.append(input_path_kisa)
                with open(input_path_kisa, "wb") as buffer:
                    shutil.copyfileobj(kisa_file.file, buffer)
                
                with open(input_path_kisa, 'r', encoding='utf-8', errors='replace') as f:
                    kisa_row_count = len([line for line in f if line.strip()])

            # c. Pre-calculate TRAP & Save
            if trap_file:
                file_id = str(uuid.uuid4())
                input_path_trap = os.path.join(UPLOAD_DIR, f"{file_id}.txt")
                uploaded_temp_files.append(input_path_trap)
                with open(input_path_trap, "wb") as buffer:
                    shutil.copyfileobj(trap_file.file, buffer)

                with open(input_path_trap, 'r', encoding='utf-8', errors='replace') as f:
                    trap_row_count = len([line for line in f if line.strip()])

            global_total_rows = kisa_row_count + trap_row_count
            total_rows = global_total_rows

            # d. Process KISA
            if input_path_kisa:
                logger.info(f"Processing KISA component from {kisa_file.filename} (found {kisa_row_count} lines)")
                
                def run_kisa(in_path=input_path_kisa, orig_name=kisa_file.filename):
                    return excel_handler.process_kisa_txt(
                        in_path, OUTPUT_DIR, process_message_with_hitl, progress_callback, 
                        batch_size=batch_chunk_size, original_filename=orig_name,
                        manager=manager, client_id=client_id, is_trap=False, override_output_path=output_path, 
                        index_offset=0, global_total_rows=global_total_rows
                    )
                    
                result = await loop.run_in_executor(None, run_kisa)

            # e. Process TRAP
            if input_path_trap:
                logger.info(f"Processing TRAP component from {trap_file.filename} (found {trap_row_count} lines) with index_offset={kisa_row_count}")
                
                def run_trap(in_path=input_path_trap, orig_name=trap_file.filename):
                    return excel_handler.process_kisa_txt(
                        in_path, OUTPUT_DIR, process_message_with_hitl, progress_callback, 
                        batch_size=batch_chunk_size, original_filename=orig_name,
                        manager=manager, client_id=client_id, is_trap=True, override_output_path=output_path, 
                        index_offset=kisa_row_count, global_total_rows=global_total_rows
                    )
                    
                result = await loop.run_in_executor(None, run_trap)
                        
            output_filename = final_filename
        
        # [Proactive Rotation] 대량 배치 완료 시 예방적 로테이션 (사용자 요청 반영)
        # 500개 이상의 배치 처리가 오류 없이(또는 복구되며) 모두 끝난 직후, 
        # 특정 키의 일일 할당량(RPD) 고갈을 막기 위해 선제적으로 다음 키로 회전시켜 줍니다.
        if total_rows >= 500:
            provider = os.getenv("LLM_PROVIDER", "OPENAI").upper()
            from app.core.llm_manager import key_manager
            logger.info(f"[Batch Complete] Processed {total_rows} items. Proactively rotating {provider} key to prevent RPD exhaustion.")
            key_manager.rotate_key(provider)
            
        # File ID reference can be anything, picking final_filename base for UI
        return {
            "id": final_filename, 
            "filename": output_filename, 
            "message": "Processing complete", 
            "total_processed": total_rows,
            "kisa_filename": kisa_file.filename if kisa_file else (excel_files[0].filename if excel_files else base_file.filename),
            "trap_filename": trap_file.filename if trap_file else ""
        }
    
    except CancellationException as e:
        logger.info(f"Processing cancelled: {e}")
        try:
            await manager.send_personal_message({
                "type": "cancellation_confirmed",
                "message": "처리가 중지되었습니다."
            }, client_id)
        except Exception as ws_error:
            logger.warning(f"Failed to send cancellation confirmation: {ws_error}")
            
        # Safe variable access
        safe_file_id = locals().get('final_filename', 'unknown')
        safe_filename = locals().get('output_filename', None)
        
        return {
            "id": safe_file_id, 
            "filename": safe_filename,
            "message": "Processing cancelled by user",
            "status": "cancelled"
        }
        
    except Exception as e:
        logger.error(f"Error during upload/processing: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # [User Request] 업로드 원본(임시파일) 즉시 파기 (Disk Full 방지)
        for tmp_file in locals().get('uploaded_temp_files', []):
            if os.path.exists(tmp_file):
                try:
                    os.remove(tmp_file)
                    logger.info(f"Temporary file deleted immediately to prevent disk leak: {tmp_file}")
                except Exception as del_err:
                    logger.warning(f"Could not delete temp file {tmp_file}: {del_err}")

class ExcelRowUpdate(BaseModel):
    """엑셀 행 업데이트 요청"""
    filename: str
    excel_row_number: int  # Required: Direct row number for update
    message: str  # For validation only
    is_spam: bool
    classification_code: str
    reason: str
    spam_probability: float = 0.95
    is_trap: bool = False

@app.put("/api/excel/update-row")
async def update_excel_row(update: ExcelRowUpdate):
    """엑셀 파일의 모든 메인 시트 행을 업데이트 + 부가 시트(URL/시그니처)의 연쇄 동기화(삭제 등) 수행"""
    import re
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment
    
    file_path = os.path.join(OUTPUT_DIR, update.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"File not found: {update.filename}")
    
    def extract_urls_from_message(message: str) -> list:
        url_pattern = r'(?:https?://|www\.)[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}|[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        urls = re.findall(url_pattern, message)
        shortener_domains = {
            "a.to", "abit.ly", "adf.ly", "adfoc.us", "aka.ms", "amzn.to", "apple.co", "asq.kr", 
            "bit.do", "bit.ly", "bitly.com", "bitly.kr", "bl.ink", "blow.pw", "buff.ly", "buly.kr", 
            "c11.kr", "clic.ke", "cogi.cc", "coupa.ng", "cutt.it", "cutt.ly", 
            "di.do", "dokdo.in", "dub.co", 
            "fb.me", 
            "gmarket.it", "goo.gl", "goo.su", "gooal.kr", 
            "han.gl", "horturl.at", 
            "ii.ad", "iii.ad", "instagr.am", "is.gd", 
            "j.mp", 
            "kakaolink.com", "ko.gl", "koe.kr", 
            "link24.kr", "linktr.ee", "lrl.kr", 
            "mcaf.ee", "me2.do", "muz.so", "myip.kr", 
            "naver.me", 
            "ouo.io", "ow.ly", 
            "qrco.de", 
            "rb.gy", "rebrand.ly", "reurl.kr", 
            "sbz.kr", "short.io", "shorter.me", "shorturl.at", "shrl.me", "shrtco.de", 
            "t.co", "t.ly", "t.me", "t2m.kr", "tiny.cc", "tinyurl.com", "tne.kr", "tny.im", "tr.ee", "tuney.kr",
            "url.kr", "uto.kr", 
            "v.gd", "vo.la", "vvd.bz", "vvd.im", 
            "wp.me", 
            "youtu.be", "yun.kr", 
            "zrr.kr"
        }
        
        try:
            list_path = os.path.join(os.path.dirname(__file__), "utils", "shorteners_list.txt")
            if os.path.exists(list_path):
                with open(list_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip().lower()
                        if line and not line.startswith("#"):
                            shortener_domains.add(line)
        except Exception as e:
            logger.warning(f"Failed to load shorteners_list.txt in main.py: {e}")
            
        result = []
        for url in urls:
            url = url.rstrip('.,;!?)]}"\'')
            clean_url = re.sub(r'^https?://', '', url.lower())
            clean_url = re.sub(r'^www\.', '', clean_url)
            is_short = any(clean_url.startswith(domain) for domain in shortener_domains)
            if not is_short and url:
                result.append(url)
        return result
    
    def _lenb(text: str) -> int:
        if not isinstance(text, str):
            text = str(text) if text is not None else ""
        try:
            return len(text.encode('cp949'))
        except UnicodeEncodeError:
            return len(text.encode('utf-8'))
            
    def sanitize(val):
        if isinstance(val, str) and val.startswith(('=', '+', '-', '@')):
             return "'" + val
        return val

    try:
        wb = load_workbook(file_path)
        
        # 1. 대상 메인 시트 탐침 (원천 데이터 + 결과 데이터)
        target_sheet_names = ["TRAP.육안분석(시뮬결과35_150)", "TRAP.시뮬결과전체"] if update.is_trap else ["육안분석(시뮬결과35_150)", "시뮬결과전체"]
        valid_sheets = [wb[name] for name in target_sheet_names if name in wb.sheetnames]
        if not valid_sheets:
            valid_sheets = [wb.active] # Fallback
            
        global_was_spam = False
        global_new_code = ""
        global_old_code = ""
        global_urls_in_message = extract_urls_from_message(update.message)
        sync_run = False
        
        spam_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        no_fill = PatternFill(fill_type=None)
        wrap_vcenter_align = Alignment(wrap_text=True, vertical='center')
        
        final_row_idx = update.excel_row_number
        
        # 2. 메인 데이터 업데이트 및 정렬
        for ws in valid_sheets:
            headers = [cell.value for cell in ws[1]]
            def get_col_idx(name):
                try: return headers.index(name) + 1
                except ValueError: return None
            
            msg_col = get_col_idx("메시지")
            gubun_col = get_col_idx("구분")
            code_col = get_col_idx("분류")
            prob_col = get_col_idx("Probability")
            reason_col = get_col_idx("Reason")
            
            if not msg_col: continue
            
            # 행 번호 복구(Smart Recovery)
            found_row = update.excel_row_number
            if found_row < 2 or found_row > ws.max_row: found_row = None
            if found_row is not None:
                cell_value = ws.cell(row=found_row, column=msg_col).value
                if not cell_value or str(cell_value).strip() != update.message.strip():
                    found_row = None
            if found_row is None:
                candidates = []
                for r in range(2, ws.max_row + 1):
                    r_msg = ws.cell(row=r, column=msg_col).value
                    if r_msg and str(r_msg).strip() == update.message.strip():
                        candidates.append(r)
                if candidates:
                    found_row = min(candidates, key=lambda r: abs(r - update.excel_row_number))
                else: continue
            
            # 상태 기록 (첫 시트 기준)
            if not sync_run:
                gubun_val = ws.cell(row=found_row, column=gubun_col).value if gubun_col else None
                global_was_spam = (gubun_val == "o")
                global_old_code = str(ws.cell(row=found_row, column=code_col).value or "") if code_col else ""
                
                if update.is_spam:
                    match = re.search(r'\d+', str(update.classification_code))
                    global_new_code = match.group(0) if match else update.classification_code
                else:
                    global_new_code = ""

            # 메인 값 변경 (SPAM 처리)
            if gubun_col: ws.cell(row=found_row, column=gubun_col, value=sanitize("o" if update.is_spam else ""))
            if code_col: ws.cell(row=found_row, column=code_col, value=sanitize(global_new_code))
            if prob_col: ws.cell(row=found_row, column=prob_col, value=sanitize(f"{int(update.spam_probability * 100)}%"))
            if reason_col: ws.cell(row=found_row, column=reason_col, value=sanitize(update.reason))
            
            # 메인 값 셀 채우기(황금색/투명) 변환
            if msg_col:
                cell = ws.cell(row=found_row, column=msg_col)
                if not global_was_spam and update.is_spam:
                    cell.fill, cell.alignment = spam_fill, wrap_vcenter_align
                elif global_was_spam and not update.is_spam:
                    cell.fill, cell.alignment = no_fill, wrap_vcenter_align

            # [메인 시트 재정렬] - 오직 구분(o) 기준
            if global_was_spam != update.is_spam and gubun_col:
                data_rows = []
                for row_idx in range(2, ws.max_row + 1):
                    row_data = []
                    for col_idx in range(1, ws.max_column + 1):
                        row_data.append(ws.cell(row=row_idx, column=col_idx).value)
                    data_rows.append(row_data)
                
                def sort_key(row):
                    g_val = row[gubun_col - 1] if len(row) >= gubun_col else ""
                    return 0 if g_val == "o" else 1
                data_rows.sort(key=sort_key)
                
                for i, row_data in enumerate(data_rows):
                    r_idx = i + 2
                    for c_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 재정렬 후 색상 서식 보정
                for r_idx in range(2, ws.max_row + 1):
                    g_val = ws.cell(row=r_idx, column=gubun_col).value
                    if msg_col:
                        cell = ws.cell(row=r_idx, column=msg_col)
                        cell.fill = spam_fill if g_val == "o" else no_fill
                        cell.alignment = wrap_vcenter_align

            sync_run = True

        # =======================================================
        # 3. 부가 시트(URL/시그니처) 연쇄 동기화 (루프 밖 1회)
        # =======================================================
        
        # 3.1) URL 중복 제거 시트 동기화
        url_sheet_names = ["TRAP.URL중복 제거"] if update.is_trap else ["URL중복 제거"]
        for u_sheet in url_sheet_names:
            if u_sheet in wb.sheetnames and global_urls_in_message:
                url_ws = wb[u_sheet]
                
                if global_was_spam and not update.is_spam: # SPAM -> HAM (URL 삭제)
                    # 타 스팸 메시지의 URL 점유 여부 확인
                    other_spam_urls = set()
                    for check_ws in valid_sheets:
                        h = [c.value for c in check_ws[1]]
                        if "구분" in h and "메시지" in h:
                            c_g, c_m = h.index("구분") + 1, h.index("메시지") + 1
                            for r_idx in range(2, check_ws.max_row + 1):
                                if check_ws.cell(row=r_idx, column=c_g).value == "o":
                                    o_msg = check_ws.cell(row=r_idx, column=c_m).value
                                    if o_msg: other_spam_urls.update(extract_urls_from_message(str(o_msg)))
                    
                    rows_to_delete = []
                    for r in range(2, url_ws.max_row + 1):
                        val = url_ws.cell(row=r, column=1).value
                        if val and val in global_urls_in_message and val not in other_spam_urls:
                            rows_to_delete.append(r)
                    for r in sorted(rows_to_delete, reverse=True):
                        url_ws.delete_rows(r)
                        
                elif not global_was_spam and update.is_spam: # HAM -> SPAM (URL 추가)
                    existing_urls = set(url_ws.cell(row=r, column=1).value for r in range(2, url_ws.max_row + 1) if url_ws.cell(row=r, column=1).value)
                    for url in global_urls_in_message:
                        if url not in existing_urls:
                            url_ws.append([url, _lenb(url), global_new_code])
                            
                elif global_was_spam and update.is_spam and global_old_code != global_new_code:
                    for r in range(2, url_ws.max_row + 1):
                        val = url_ws.cell(row=r, column=1).value
                        if val and val in global_urls_in_message:
                            url_ws.cell(row=r, column=3, value=global_new_code)

        # 3.2) 시그니처 연쇄 동기화 (문자문장차단 + 문자열 + 문장열)
        clean_msg = re.sub(r'[ \t\r\n\f\v]+', '', update.message)
        bl_sheet_name = "TRAP.문자문장차단등록" if update.is_trap else "문자문장차단등록"
        
        signatures_to_remove = set()
        
        if bl_sheet_name in wb.sheetnames:
            bl_ws = wb[bl_sheet_name]
            
            if global_was_spam and not update.is_spam: # SPAM -> HAM (차단등록 삭제 및 시그니처 확보)
                rows_to_delete = []
                for r in range(2, bl_ws.max_row + 1):
                    bl_msg = bl_ws.cell(row=r, column=1).value
                    if bl_msg and str(bl_msg).strip() == clean_msg.strip():
                        s_str = bl_ws.cell(row=r, column=2).value
                        s_sen = bl_ws.cell(row=r, column=4).value
                        if s_str: signatures_to_remove.add(str(s_str))
                        if s_sen: signatures_to_remove.add(str(s_sen))
                        rows_to_delete.append(r)
                for r in sorted(rows_to_delete, reverse=True):
                    bl_ws.delete_rows(r)
                    
            elif global_was_spam and update.is_spam and global_old_code != global_new_code:
                for r in range(2, bl_ws.max_row + 1):
                    bl_msg = bl_ws.cell(row=r, column=1).value
                    if bl_msg and str(bl_msg).strip() == clean_msg.strip():
                        bl_ws.cell(row=r, column=6, value=global_new_code) # 분류 업데이트
                        s_str = bl_ws.cell(row=r, column=2).value
                        s_sen = bl_ws.cell(row=r, column=4).value
                        if s_str: signatures_to_remove.add(str(s_str))
                        if s_sen: signatures_to_remove.add(str(s_sen))

        # 3.3) 문자열/문장열 순수 중복제거 시트 동기화 (연쇄 삭제)
        if signatures_to_remove:
            derive_sheets = [
                "TRAP.문자열 중복제거" if update.is_trap else "문자열중복제거",
                "TRAP.문장 중복제거" if update.is_trap else "문장중복제거"
            ]
            
            # 삭제 전: 타 스팸이 이 시그니처를 쓰고 있는지 확인(문자문장차단등록 시트 스캔)
            other_signatures = set()
            if bl_sheet_name in wb.sheetnames:
                for r in range(2, wb[bl_sheet_name].max_row + 1):
                    o_str = wb[bl_sheet_name].cell(row=r, column=2).value
                    o_sen = wb[bl_sheet_name].cell(row=r, column=4).value
                    if o_str: other_signatures.add(str(o_str))
                    if o_sen: other_signatures.add(str(o_sen))
            
            for d_sheet in derive_sheets:
                if d_sheet in wb.sheetnames:
                    d_ws = wb[d_sheet]
                    if global_was_spam and not update.is_spam: # 연쇄 삭제 적용
                        rows_to_delete = []
                        for r in range(2, d_ws.max_row + 1):
                            key_val = d_ws.cell(row=r, column=1).value
                            if key_val and str(key_val) in signatures_to_remove and str(key_val) not in other_signatures:
                                rows_to_delete.append(r)
                        for r in sorted(rows_to_delete, reverse=True):
                            d_ws.delete_rows(r)
                            
                    elif global_was_spam and update.is_spam and global_old_code != global_new_code:
                        for r in range(2, d_ws.max_row + 1):
                            key_val = d_ws.cell(row=r, column=1).value
                            if key_val and str(key_val) in signatures_to_remove:
                                d_ws.cell(row=r, column=3, value=global_new_code)
        
        # 4. 저장 및 결과 반환
        wb.save(file_path)
        
        sync_info = []
        if global_was_spam != update.is_spam:
            sync_info.append(f"{'SPAM→HAM' if global_was_spam else 'HAM→SPAM'}")
        if global_was_spam and update.is_spam and global_old_code != global_new_code:
            sync_info.append(f"Code: {global_old_code}→{global_new_code}")
            
        return {
            "success": True,
            "message": f"Row updated successfully across sheets",
            "sync": sync_info if sync_info else None,
            "urls_affected": len(global_urls_in_message)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating Excel row: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{filename}")
async def download_file(filename: str, suggested_name: str = None):
    file_path = os.path.join(OUTPUT_DIR, filename)
    
    if os.path.exists(file_path):
        from urllib.parse import quote
        # Use suggested_name if provided, otherwise use original filename
        download_name = suggested_name if suggested_name else filename
        # Ensure it has the correct extension if suggested_name forgot it
        if download_name and not download_name.lower().endswith('.xlsx') and filename.lower().endswith('.xlsx'):
             download_name += '.xlsx'
             
        encoded_filename = quote(download_name, safe='')
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
        return FileResponse(file_path, filename=None, headers=headers)
    raise HTTPException(status_code=404, detail="File not found")


if __name__ == "__main__":
    import uvicorn
    # Use generic run config; reload might still be tricky in script but let's try without reload first or with simple config
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
