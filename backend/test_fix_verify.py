"""
버그 수정 검증 테스트 스크립트 v2
- 업로드 → 엑셀 결과 파싱으로 검증

Fix 1: Bare Domain URL SPAM 확정 시 drop_url 오탐 방지 (554: jusigmaenia.com)
Fix 2: Runtime Cache 수신 시 방패막이 drop_url 미전파 수정 (1122, 1124: 풍월2.o-r.kr)
"""

import asyncio
import httpx
import os
import tempfile
from openpyxl import load_workbook

# 테스트 메시지 (탭 구분, 4월 14일 B 원본)
TEST_MESSAGES = [
    # 554번: 루트 도메인 SPAM 확정 → URL이 엑셀에 표시되어야 함 (drop_url=False)
    ("[시장 흔들림, 계속 버티실 건가요?] 수익은 줄고, 불안한 투자만 반복되고 있습니다  지금 필요한 건 검증된 전략 + 따라가기만 하는 시스템  [주식매니아의 투자전략] 데이터 기반 종목 선별 초보도 가능한 완성형 매매 구조 실시간 리딩으로 핵심 타이밍 포착  복잡한 분석, 실패 반복 없이 [간편 참여 / 안정 설계]로 부담 없이 시작  지금도 늦지 않았지만 이 기회 놓치면 같은 혜택은 어렵습니다  이미 많은 투자자들이 혼자 하는 투자에서 벗어나고 있습니다  [리스크 최소화 + 수익 극대화 전략] 지금 바로 확인하세요   무료 참여 링크: https://www.jusigmaenia.com/",
     "https://www.jusigmaenia.com"),

    # 1036번 (리더): 방패막이 → drop_url=True, 시그니처 보존
    ("풍경 월마  월 일본NAR 20:50  문의사항 4963 - 9478  풍월2.o-r.kr",
     "2.o-r.kr"),

    # 1038번 (팔로워): Runtime Cache 수신 → 수정 후 drop_url=True, 시그니처 보존
    ("풍경 월마  국내 17:55  일본NAR 20:50  총판및신규모집  4963 - 9478  풍월2.o-r.kr",
     "2.o-r.kr"),
]

async def run_test():
    # 1. 임시 TXT 파일 생성 (CP949)
    content = "\r\n".join(f"{msg}\t{url}" for msg, url in TEST_MESSAGES) + "\r\n"
    with tempfile.NamedTemporaryFile(
        suffix=".txt", prefix="kisa_test_fix_",
        delete=False, mode='wb'
    ) as tmp:
        tmp.write(content.encode('cp949', errors='replace'))
        tmp_path = tmp.name

    print(f"\n{'='*60}")
    print("📋 버그 수정 검증 테스트 (4월 14일 B 데이터 3건)")
    print(f"{'='*60}")

    output_filename = None
    try:
        base_url = "http://localhost:8000"
        client_id = "test_fix_verify_001"

        async with httpx.AsyncClient(timeout=300.0) as client:
            print("[1/2] 📤 파일 업로드 및 분석 시작...")
            with open(tmp_path, 'rb') as f:
                resp = await client.post(
                    f"{base_url}/upload",
                    data={"client_id": client_id},
                    files={"files": (os.path.basename(tmp_path), f, "text/plain")}
                )

            if resp.status_code != 200:
                print(f"❌ 업로드 실패: {resp.status_code} - {resp.text[:200]}")
                return

            resp_data = resp.json()
            output_filename = resp_data.get("filename")
            total = resp_data.get("total_processed", 0)
            print(f"✅ 분석 완료: {total}건 처리")
            print(f"   출력파일: {output_filename}")

        # 2. 생성된 엑셀 파일 다운로드
        output_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", "data", "outputs"
        )
        output_path = os.path.join(output_dir, output_filename)

        if not os.path.exists(output_path):
            print(f"❌ 출력 파일 없음: {output_path}")
            return

        print(f"\n[2/2] 📊 엑셀 결과 분석 중...")

        # 3. 육안분석 시트에서 결과 확인
        wb = load_workbook(output_path, read_only=True)
        ws = None
        for name in wb.sheetnames:
            if "육안분석" in name or "시뮬결과" in name:
                ws = wb[name]
                break

        if not ws:
            print(f"❌ 육안분석 시트 없음. 시트 목록: {wb.sheetnames}")
            return

        print(f"\n   시트: {ws.title}")
        print(f"   {'='*55}")

        EXPECTED = [
            {"desc": "[554] jusigmaenia.com (루트 도메인 SPAM)", "check": "URL 있어야 함 (drop_url=False)"},
            {"desc": "[1122] 풍월2.o-r.kr (리더, 방패막이)",     "check": "URL 없고 시그니처 있어야 함"},
            {"desc": "[1124] 풍월2.o-r.kr (팔로워, Cache)",      "check": "URL 없고 시그니처 있어야 함"},
        ]

        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # 메시지 컬럼이 있는 행
                data_rows.append(row)

        print(f"\n   총 {len(data_rows)}행 발견\n")

        for i, (row, exp) in enumerate(zip(data_rows[:3], EXPECTED)):
            msg_val = str(row[0])[:50] if row[0] else ""
            url_val = str(row[1]) if row[1] else ""
            gubun   = str(row[2]) if row[2] else ""
            code    = str(row[3]) if row[3] else ""
            reason  = str(row[9])[:80] if len(row) > 9 and row[9] else ""

            print(f"   [{i+1}] {exp['desc']}")
            print(f"        메시지  : {msg_val}...")
            print(f"        URL     : {url_val or '(없음)'}")
            print(f"        구분/코드: {gubun} / {code}")
            print(f"        사유(축) : {reason}")
            print(f"        기대     : {exp['check']}")

            # 검증
            if i == 0:  # 554: URL 있어야 함
                if url_val and "jusigmaenia" in url_val.lower():
                    print(f"        ✅ PASS - URL 정상 기록됨")
                else:
                    print(f"        ❌ FAIL - URL이 비어있음! (Bare Domain drop_url 오탐 미수정)")
            else:  # 1122, 1124: URL 없어야 함
                if not url_val or url_val == "None":
                    print(f"        ✅ PASS - URL 없음 (방패막이 정상 처리)")
                else:
                    print(f"        ⚠️  WARN - URL 있음: {url_val} (방패막이인데 URL 남음)")
            print()

        wb.close()
        print(f"   출력 파일: {output_path}")

    finally:
        os.unlink(tmp_path)
        print(f"\n{'='*60}")
        print("테스트 완료")
        print(f"{'='*60}\n")

if __name__ == "__main__":
    asyncio.run(run_test())
