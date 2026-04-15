import sys
import glob
import openpyxl

files = glob.glob("./spams/*20260413_A.xlsx")
target = None
for f in files:
    if "SD Output" not in f:
        target = f
        break

if not target:
    print("파일을 찾을 수 없습니다.")
    sys.exit(1)

print(f"Reading: {target}")
wb = openpyxl.load_workbook(target, data_only=True)

sheets_to_read = ["문자열중복제거", "문장중복제거"]
for s in sheets_to_read:
    print(f"\n### {s}")
    if s not in wb.sheetnames:
        print(f"시트 '{s}' 가 없습니다.")
        continue
        
    ws = wb[s]
    headers = [c.value for c in ws[1][:3]]
    print(f"| {' | '.join(str(h) for h in headers)} |")
    print("|---" * len(headers) + "|")
    
    for row in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=col).value for col in range(1, 4)]
        if any(row_vals):
            print(f"| {' | '.join(str(v) if v is not None else '' for v in row_vals)} |")
