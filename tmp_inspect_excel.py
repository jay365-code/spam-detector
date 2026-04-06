import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\leejo\Project\AI Agent\Spam Detector\spams\MMSC스팸추출_20260327_A.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n--- Sheet: {sheet_name} ---')
    print('Column Widths:')
    for col in range(1, min(7, ws.max_column+1)):
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f'  Col {col_letter}: width={ws.column_dimensions[col_letter].width}')
    print('Header (Row 1):')
    for cell in list(ws[1])[:5]:
        font = cell.font
        fill = cell.fill.start_color.index if cell.fill.start_color else None
        print(f'  {cell.coordinate} ({str(cell.value)[:15]}): font_name={font.name}, size={font.sz}, bold={font.b}, fill={fill}, align={cell.alignment.horizontal}, align_v={cell.alignment.vertical}')
    print('Data (Row 2):')
    row = list(ws.iter_rows(min_row=2, max_row=2))
    if row:
        for cell in row[0][:5]:
             font = cell.font
             fill = cell.fill.start_color.index if cell.fill.start_color else None
             print(f'  {cell.coordinate} ({str(cell.value)[:15]}): font_name={font.name}, size={font.sz}, bold={font.b}, fill={fill}, align={cell.alignment.horizontal}, align_v={cell.alignment.vertical}')
