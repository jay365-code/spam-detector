# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook

wb = load_workbook(r'c:\Users\leejo\Project\AI Agent\Spam Detector\spams\MMSC스팸추출_20260327_A.xlsx', data_only=True)
ws = wb['육안분석(시뮬결과35_150)']

data = []
for row in ws.iter_rows(min_row=2):
    if not row[0].value: break
    color = str(row[0].fill.start_color.rgb) if getattr(row[0].fill, 'start_color', None) else 'None'
    # We dump to investigate sorting
    data.append({
        'color': color,
        'msg': str(row[0].value).replace('\n', ' ')[:20],
        'url': str(row[1].value) if row[1].value else '',
        'gubun': str(row[2].value) if row[2].value else '',
        'cls': str(row[3].value) if row[3].value else '',
        'msglen': row[4].value if len(row)>4 else '',
        'urllen': row[5].value if len(row)>5 else '',
    })

df = pd.DataFrame(data)
df.to_csv('tmp_analysis.csv', index=False, encoding='utf-8-sig')
print("Saved to tmp_analysis.csv")
