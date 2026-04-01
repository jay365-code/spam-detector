import unicodedata
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Dict, Any, Optional
import io
import hashlib
import warnings

# 로깅 설정 (다른 import 전에 초기화)
from logging_config import setup_logging, get_logger
setup_logging()
logger = get_logger(__name__)

# Suppress warnings
warnings.filterwarnings("ignore")

# Import metrics
from metrics import calculate_advanced_metrics, interpret_policy, generate_summary_text
import monitor

app = FastAPI()

# CORS for frontend (Port 5173 and 5174)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(monitor.router)

import re

def normalize_text(text: Any) -> str:
    if pd.isna(text):
        return ""
    text = str(text)
    # NFKC normalization
    text = unicodedata.normalize('NFKC', text)
    
    # KISA 공통 접두어 제거 ([Web발신], [국제발신], (광고) 등)
    text = re.sub(r'^(\W*(web발신|웹발신|국제발신|로밍발신|광고|fw|fwd)\W*)+', '', text, flags=re.IGNORECASE)
    
    # Generic whitespace normalization (remove all spaces for robust matching)
    text_no_space = "".join(text.split())
    # Remove all non-alphanumeric characters (including punctuation) for robust matching
    norm_t = re.sub(r'[^\w]', '', text_no_space)
    
    # 텍스트가 특수기호/구두점으로만 이루어진 경우 (예: 박스 기호), 원본 기호 집합(공백 제거)을 유지
    if not norm_t:
        return text_no_space
        
    return norm_t

def normalize_label(val: Any) -> bool:
    if pd.isna(val):
        return False
    # Check for 'o' (spam indicator)
    s = str(val).strip()
    return s == 'o'


# ========== 로그 레벨 런타임 변경 API ==========
from logging_config import get_log_levels, set_log_level, set_console_enabled
from pydantic import BaseModel

class ExportDiffRequest(BaseModel):
    summary: dict
    human_based_diffs: list
    filename: str

class OpenFileRequest(BaseModel):
    path: str

class LogLevelChange(BaseModel):
    target: str  # "console" 또는 "file"
    level: str   # "DEBUG", "INFO", "WARNING", "ERROR"

class ConsoleToggle(BaseModel):
    enabled: bool  # True=ON, False=OFF

@app.get("/api/log-level")
async def get_current_log_level():
    """현재 로그 레벨 및 콘솔 상태 조회"""
    return get_log_levels()

@app.post("/api/log-level")
async def change_log_level(request: LogLevelChange):
    """런타임에 로그 레벨 변경"""
    result = set_log_level(request.target, request.level)
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    return result

@app.post("/api/log-console")
async def toggle_console_log(request: ConsoleToggle):
    """콘솔 로그 출력 ON/OFF"""
    return set_console_enabled(request.enabled)


import os

@app.post("/compare")
async def compare_results(
    human_file: UploadFile = File(...),
    llm_file: UploadFile = File(...),
    sheet_name: str = Form("육안분석(시뮬결과35_150)")
):
    logger.info(f"비교 요청 | Human: {human_file.filename} | LLM: {llm_file.filename} | Sheet: {sheet_name}")
    
    try:
        human_content = await human_file.read()
        llm_content = await llm_file.read()
        df_human = pd.read_excel(io.BytesIO(human_content), sheet_name=sheet_name)
        
        # Parse all sheets for LLM to get the signature sheet
        xl_llm = pd.ExcelFile(io.BytesIO(llm_content))
        df_llm = xl_llm.parse(sheet_name)
        
        # Find Signature sheet ('문자문장차단등록' or similar)
        df_llm_sig = None
        for s in xl_llm.sheet_names:
            if '문자문장차단등록' in s or '차단' in s:
                df_llm_sig = xl_llm.parse(s)
                break
                
    except ValueError as e:
        logger.warning(f"Sheet not found: {sheet_name} - {e}")
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found or error: {str(e)}")
    except Exception as e:
        logger.exception("Excel 로드 중 오류")
        raise HTTPException(status_code=400, detail=f"Error reading excel file: {str(e)}")

    return _process_dataframes(df_human, df_llm, sheet_name, df_llm_sig)

@app.post("/compare/auto")
async def compare_auto(
    date: str = Form(...),
    file_type: str = Form(...),
    sheet_name: str = Form("육안분석(시뮬결과35_150)")
):
    logger.info(f"자동 비교 요청 | Date: {date} | Type: {file_type} | Sheet: {sheet_name}")
    
    curr_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir = os.path.dirname(os.path.dirname(curr_dir))
    
    human_path = os.path.join(base_dir, "spams", f"MMSC스팸추출_{date}_{file_type}.xlsx")
    llm_path = os.path.join(base_dir, "spams", "SD Output", f"MMSC스팸추출_{date}_{file_type}.xlsx")
    
    if not os.path.exists(human_path):
        raise HTTPException(status_code=404, detail=f"Human file not found: {human_path}")
    if not os.path.exists(llm_path):
        raise HTTPException(status_code=404, detail=f"LLM file not found: {llm_path}")
        
    try:
        df_human = pd.read_excel(human_path, sheet_name=sheet_name)
        
        xl_llm = pd.ExcelFile(llm_path)
        df_llm = xl_llm.parse(sheet_name)
        
        # Find Signature sheet
        df_llm_sig = None
        for s in xl_llm.sheet_names:
            if '문자문장차단등록' in s or '차단' in s:
                df_llm_sig = xl_llm.parse(s)
                break
    except ValueError as e:
        logger.warning(f"Sheet not found: {sheet_name} - {e}")
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found or error: {str(e)}")
    except Exception as e:
        logger.exception("Excel 로드 중 오류")
        raise HTTPException(status_code=400, detail=f"Error reading excel file: {str(e)}")

    return _process_dataframes(df_human, df_llm, sheet_name, df_llm_sig)

def _process_dataframes(df_human, df_llm, sheet_name, df_llm_sig=None):

    # Check Columns
    required_cols = ["메시지", "구분"]
    for col in required_cols:
        if col not in df_human.columns:
            raise HTTPException(status_code=400, detail=f"Human file missing column: {col}")
        if col not in df_llm.columns:
            raise HTTPException(status_code=400, detail=f"LLM file missing column: {col}")

    # Process Data with index preservation
    def process_df(df):
        df = df.copy()
        # Keep original index
        df['original_index'] = df.index
        df['norm_msg'] = df['메시지'].apply(normalize_text)
        df['is_spam'] = df['구분'].apply(normalize_label)
        
        # Determine Reason column (Reason or 사유)
        reason_col = None
        for col in ['Reason', '사유', 'reason']:
            if col in df.columns:
                reason_col = col
                break
        
        if reason_col:
            df['reason'] = df[reason_col].fillna("").astype(str)
        else:
            df['reason'] = ""

        # Determine Code column (분류 - Classification Code)
        if '분류' in df.columns:
             df['code'] = df['분류'].fillna("").astype(str).replace(r'\.0$', '', regex=True) # Remove .0 for integers
        else:
             df['code'] = ""

        # Determine Semantic Class column
        if 'Semantic Class' in df.columns:
            df['semantic_class'] = df['Semantic Class'].fillna("").astype(str)
        else:
            df['semantic_class'] = ""

        # Occurrence Index for Duplicates
        # This handles multiple identical messages by assigning 1, 2, 3...
        df['cc_idx'] = df.groupby('norm_msg').cumcount() + 1
        return df

    df_h = process_df(df_human)
    df_l = process_df(df_llm)

    # Combined Merge for Match and Missing detection
    merged_all = pd.merge(
        df_h, 
        df_l, 
        on=['norm_msg', 'cc_idx'], 
        how='outer', 
        suffixes=('_human', '_llm'),
        indicator=True
    )
    
    # ----------------------------------------------------
    # Type B (FP Sentinel Override - 학습 보호) 처리:
    # 1. Human이 SPAM (True) 이고 AI가 Type B인 경우 -> AI도 SPAM (True) 으로 간주 (뒷단 검출 예정이므로 FN 제외)
    # 2. Human이 HAM (False) 이고 AI가 Type B인 경우 -> AI도 HAM (False) 으로 간주 (FP로 잡히지 않도록)
    # ----------------------------------------------------
    type_b_mask = merged_all['reason_llm'].astype(str).str.contains(r'\[FP Sentinel Override\]', case=False, na=False) | \
                  merged_all['semantic_class_llm'].astype(str).str.contains(r'Type_B', case=False, na=False)
    
    # Human이 SPAM인 케이스
    merged_all.loc[type_b_mask & (merged_all['is_spam_human'] == True), 'is_spam_llm'] = True
    # Human이 HAM인 케이스
    merged_all.loc[type_b_mask & (merged_all['is_spam_human'] == False), 'is_spam_llm'] = False


    # Split into Matched / Missing
    merged = merged_all[merged_all['_merge'] == 'both'].copy()
    missing_in_llm_df = merged_all[merged_all['_merge'] == 'left_only'].copy()
    missing_in_human_df = merged_all[merged_all['_merge'] == 'right_only'].copy()

    def format_missing(df, side):
        msg_col = f'메시지_{side}'
        label_col = f'구분_{side}'
        idx_col = f'original_index_{side}'
        code_col = f'code_{side}'
        reason_col = f'reason_{side}'
        
        return [
            {
                "index": int(row[idx_col]),
                "message": str(row[msg_col]),
                "label": str(row[label_col]),
                "code": str(row.get(code_col, '')),
                "reason": str(row.get(reason_col, ''))
            } for _, row in df.iterrows()
        ]

    missing_in_llm = format_missing(missing_in_llm_df, 'human')
    missing_in_human = format_missing(missing_in_human_df, 'llm')

    # Calculate Metrics
    tp = len(merged[(merged['is_spam_human'] == True) & (merged['is_spam_llm'] == True)])
    fn = len(merged[(merged['is_spam_human'] == True) & (merged['is_spam_llm'] == False)])
    fp = len(merged[(merged['is_spam_human'] == False) & (merged['is_spam_llm'] == True)])
    tn = len(merged[(merged['is_spam_human'] == False) & (merged['is_spam_llm'] == False)])
    
    matched_count = len(merged)
    same_label_count = len(merged[merged['is_spam_human'] == merged['is_spam_llm']])
    
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0.0
    
    # 고급 지표 계산
    advanced_metrics = calculate_advanced_metrics(tp, tn, fp, fn, matched_count)
    
    # Generate Diffs (Mismatched Labels)
    diffs = []
    # Filter for mismatches
    mismatch = merged[merged['is_spam_human'] != merged['is_spam_llm']]

    
    for _, row in mismatch.iterrows():
        diff_type = "FN" if row['is_spam_human'] else "FP"
        policy_tag = interpret_policy(
            diff_type, 
            row.get('reason_llm', ''), 
            row.get('reason_human', '')
        )
        
        diffs.append({
            "diff_id": hashlib.md5(f"{row['original_index_human']}_{row['original_index_llm']}".encode()).hexdigest(),
            "diff_type": diff_type,
            "message_preview": str(row['메시지_human'])[:80] + "...",
            "message_full": str(row['메시지_human']),
            "human_label_raw": str(row['구분_human']),
            "llm_label_raw": str(row['구분_llm']),
            "human_code": str(row.get('code_human', '')),
            "llm_code": str(row.get('code_llm', '')),
            "human_is_spam": bool(row['is_spam_human']),
            "llm_is_spam": bool(row['is_spam_llm']),
            "human_reason": str(row.get('reason_human', '')),
            "llm_reason": str(row.get('reason_llm', '')),
            "match_key": f"{row['norm_msg'][:10]}... (idx: {row['cc_idx']})",
            "policy_interpretation": policy_tag
        })

    # Pre-process signature signatures mapped by matching message text
    signature_map = {}
    if df_llm_sig is not None and not df_llm_sig.empty:
        # Looking for '메시지' and '문자열'/'문장열' columns
        msg_col = next((c for c in df_llm_sig.columns if '메시지' in str(c)), None)
        sig_col = next((c for c in df_llm_sig.columns if str(c) == '문자열'), None)
        sentence_col = next((c for c in df_llm_sig.columns if str(c) == '문장열'), None)
        
        if msg_col:
            for _, sig_row in df_llm_sig.iterrows():
                try:
                    s_msg = normalize_text(str(sig_row[msg_col]))
                    s_val = ""
                    
                    # 1. Check 짧은 문자열 (<= 20)
                    if sig_col and str(sig_row[sig_col]) and str(sig_row[sig_col]) != 'nan':
                        s_val = str(sig_row[sig_col])
                    # 2. Check 긴 문장열 (> 20)
                    elif sentence_col and str(sig_row[sentence_col]) and str(sig_row[sentence_col]) != 'nan':
                        s_val = str(sig_row[sentence_col])
                        
                    if s_val and s_val != 'nan':
                        # CSV Injection 방어용으로 추가된 선행 홑따옴표 제거 ('=이*영*아= -> =이*영*아=)
                        if s_val.startswith("'") and len(s_val) > 1 and s_val[1] in ('=', '+', '-', '@'):
                            s_val = s_val[1:]
                        signature_map[s_msg] = s_val
                except Exception:
                    pass

    # Generate Human-based Integrated Diffs
    human_based_diffs = []
    
    # We want ALL rows from human (both matched and missing in LLM)
    human_all = pd.concat([merged, missing_in_llm_df])
    
    # Sort by original index to keep the sequence of the human file
    human_all = human_all.sort_values(by='original_index_human')

    for _, row in human_all.iterrows():
        is_missing_in_llm = pd.isna(row.get('메시지_llm'))
        
        # Determine Match Status
        match_status = "MISSING_IN_LLM"
        if not is_missing_in_llm:
            match_status = "MATCH" if row['is_spam_human'] == row['is_spam_llm'] else ("FN" if row['is_spam_human'] else "FP")
            
        # Get URL & SIGNATURE
        # Get URL & SIGNATURE
        llm_url = ""
        llm_signature = ""
        
        if not is_missing_in_llm:
            llm_url = str(row.get('URL_llm', '')) if 'URL_llm' in row and pd.notna(row['URL_llm']) else ""
            norm_msg = normalize_text(str(row['메시지_human']))
            llm_signature = signature_map.get(norm_msg, "")
            
            is_llm_spam = bool(row.get('is_spam_llm', False))
            
        human_based_diffs.append({
            "index": int(row['original_index_human']),
            "message_full": str(row['메시지_human']),
            "human_is_spam": bool(row['is_spam_human']),
            "human_code": str(row.get('code_human', '')),
            "human_reason": str(row.get('reason_human', '')),
            "llm_is_spam": bool(row.get('is_spam_llm', False)) if not is_missing_in_llm else None,
            "llm_semantic_class": str(row.get('semantic_class_llm', '')) if not is_missing_in_llm else "",
            "llm_code": str(row.get('code_llm', '')) if not is_missing_in_llm and is_llm_spam else "",
            "llm_reason": str(row.get('reason_llm', '')) if not is_missing_in_llm else "",
            "llm_url": llm_url,
            "llm_signature": llm_signature,
            "match_status": match_status
        })

    # Generate Type B List
    type_b_items = []

    type_b_df = df_l[(df_l['semantic_class'].str.startswith("Type_B", na=False)) & (df_l['is_spam'] == False)]
    for _, row in type_b_df.iterrows():
        # Get URL if it exists
        extracted_url = str(row['URL']) if 'URL' in row and pd.notna(row['URL']) else ""
        
        # Look up signature based on message content
        norm_msg = normalize_text(str(row['메시지']))
        extracted_signature = signature_map.get(norm_msg, "")
        
        type_b_items.append({
            "message_preview": str(row['메시지'])[:80] + "...",
            "message_full": str(row['메시지']),
            "semantic_class": str(row['semantic_class']),
            "llm_reason": str(row.get('reason', '')),
            "llm_code": str(row.get('code', '')),
            "is_spam": bool(row['is_spam']),
            "extracted_url": extracted_url,
            "extracted_signature": extracted_signature
        })

    # 자동 요약 생성
    type_b_url_count = len(type_b_df[type_b_df['semantic_class'] == "Type_B (URL)"])
    type_b_sig_count = len(type_b_df[type_b_df['semantic_class'] == "Type_B (SIGNATURE)"])
    type_b_both_count = len(type_b_df[type_b_df['semantic_class'] == "Type_B (URL, SIGNATURE)"])
    type_b_none_count = len(type_b_df[type_b_df['semantic_class'] == "Type_B (NONE)"])
    type_b_total_count = len(type_b_df)

    # Generate Type A List
    type_a_items = []
    type_a_df = df_l[df_l['is_spam'] == True]
    for _, row in type_a_df.iterrows():
        type_a_items.append({
            "message_preview": str(row['메시지'])[:80] + "...",
            "message_full": str(row['메시지']),
            "semantic_class": str(row.get('semantic_class', '')),
            "llm_reason": str(row.get('reason', '')),
            "llm_code": str(row.get('code', '')),
            "is_spam": bool(row['is_spam'])
        })

    summary_dict = {
        "sheet_used": sheet_name,
        "total_human": len(df_h),
        "total_llm": len(df_l),
        "type_b_total_count": type_b_total_count,
        "type_b_url_count": type_b_url_count,
        "type_b_sig_count": type_b_sig_count,
        "type_b_both_count": type_b_both_count,
        "type_b_none_count": type_b_none_count,
        "human_spam_count": int(df_h['is_spam'].sum()),
        "llm_spam_count": int(df_l['is_spam'].sum()),
        "human_spam_rate": float(df_h['is_spam'].mean()) if len(df_h) > 0 else 0.0,
        "llm_spam_rate": float(df_l['is_spam'].mean()) if len(df_l) > 0 else 0.0,
        "matched": matched_count,
        "match_rate": matched_count / len(df_h) if len(df_h) > 0 else 0.0,
        "agreement_rate": same_label_count / matched_count if matched_count > 0 else 0.0,
        "tp": tp, "fp": fp, "fn": fn, "tn": tn,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
        # 주요 지표
        "accuracy": advanced_metrics["accuracy"],
        "kappa": advanced_metrics["kappa"],
        "kappa_status": advanced_metrics["kappa_status"],
        "mcc": advanced_metrics["mcc"],
        "disagreement_rate": advanced_metrics["disagreement_rate"],
        # 주요 판정 (Accuracy + Kappa 기반)
        "primary_status": advanced_metrics["primary_status"],
        "primary_color": advanced_metrics["primary_color"],
        "primary_description": advanced_metrics["primary_description"],
        # 보조 지표 (HEI)
        "hei": advanced_metrics["hei"],
        "hei_status": advanced_metrics["hei_status"],
        "hei_color": advanced_metrics["hei_color"]
    }
    
    auto_summary = generate_summary_text(summary_dict, advanced_metrics)
    
    logger.info(f"비교 완료 | Matched: {matched_count} | TP: {tp} | FP: {fp} | FN: {fn} | TN: {tn} | Accuracy: {advanced_metrics['accuracy']} | MissingH: {len(missing_in_human)} | MissingL: {len(missing_in_llm)}")
    
    return {
        "summary": summary_dict,
        "diffs": diffs,
        "human_based_diffs": human_based_diffs,
        "type_b_items": type_b_items,
        "type_a_items": type_a_items,
        "missing_in_human": missing_in_human,
        "missing_in_llm": missing_in_llm,
        "auto_summary": auto_summary
    }

@app.post("/export/diff")
async def export_diff(request: ExportDiffRequest):
    try:
        base_dir = r"c:\Users\leejo\Project\AI Agent\Spam Detector\data\reports\DIFF"
        os.makedirs(base_dir, exist_ok=True)
        save_path = os.path.join(base_dir, request.filename)
        
        # 1. (Dashboard generation is moved inside ExcelWriter)
        
        # 2. Create Diff DataFrame
        def safe_str(val):
            s = str(val).strip() if val is not None else ""
            # Escape strings that might be interpreted as formulas
            if s.startswith(('=', '+', '-', '@')):
                return "'" + s
            return s
            
        diff_items = []
        for idx, d in enumerate(request.human_based_diffs):
            llm_val = "누락"
            if d.get("llm_is_spam") is True:
                sem_class = d.get("llm_semantic_class", "").replace("_", " ") # "Type_A" -> "Type A"
                if sem_class:
                    llm_val = f"SPAM: {sem_class}"
                else:
                    llm_val = "SPAM"
            elif d.get("llm_is_spam") is False:
                llm_val = "HAM"
            
            m_status = d.get("match_status", "")
            m_status_disp = m_status
            if m_status == "FP":
                m_status_disp = "FP (오탐)"
            elif m_status == "FN":
                m_status_disp = "FN (미탐)"
                
            # 엑셀 출력용 URL (규칙 적용: Type B인데 URL 구분이 아니면 빈칸 처리)
            llm_url_for_excel = safe_str(d.get("llm_url", ""))
            sem_class = d.get("llm_semantic_class", "")
            if sem_class and sem_class.startswith("Type_B") and "(URL" not in sem_class:
                llm_url_for_excel = ""
                
            diff_items.append({
                "Row 순번": d.get("index", idx) + 1,
                "메시지 원본": safe_str(d.get("message_full", "")),
                "정답 (Human)": "SPAM" if d.get("human_is_spam") else "HAM",
                "Human 분류코드": safe_str(d.get("human_code", "")),
                "AI 판단 (LLM)": llm_val,
                "AI 분류코드": safe_str(d.get("llm_code", "")),
                "AI 제출용 URL (입력 URL 기준)": llm_url_for_excel,
                "AI 추출 SIGNATURE": safe_str(d.get("llm_signature", "")),
                "AI 사유": safe_str(d.get("llm_reason", "")),
                "매칭 상태": safe_str(m_status_disp)
            })
        df_diff = pd.DataFrame(diff_items)
        
        from openpyxl.styles import PatternFill, Font, Alignment

        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            # First write Diff table to create workbook
            df_diff.to_excel(writer, sheet_name='Diff', index=False)
            
            # Now build Dashboard interactively
            wb = writer.book
            ws_dash = wb.create_sheet('Dashboard', 0)
            
            sum_data = request.summary
            
            # Dashboard Styles
            title_font = Font(size=14, bold=True, color="FFFFFF")
            title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
            section_font = Font(size=11, bold=True, color="1F497D")
            section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            
            def write_pair(r, c, label, val):
                c_lbl = ws_dash.cell(row=r, column=c, value=label)
                c_lbl.font = bold_font
                c_val = ws_dash.cell(row=r, column=c+1, value=val)
                # Text = Center, Number = Right
                if isinstance(val, str):
                    c_val.alignment = center_align
                else:
                    c_val.alignment = right_align

            # Layout: Title
            ws_dash.merge_cells('A1:F2')
            cell_title = ws_dash['A1']
            cell_title.value = f"🤖 AI Spam Validator Dashboard [{request.filename}]"
            cell_title.font = title_font
            cell_title.fill = title_fill
            cell_title.alignment = center_align

            # Layout: 1. 핵심 성과 지표 (KPI 달성 현황)
            ws_dash.merge_cells('A4:F4')
            ws_dash['A4'] = "▶ 1. 핵심 성과 지표 (KPI 달성 현황)"
            ws_dash['A4'].font = section_font
            ws_dash['A4'].fill = section_fill
            
            # KPI Calculations
            tn = sum_data.get("tn", 0)
            fp = sum_data.get("fp", 0)
            fpr = (fp / (fp + tn)) if (fp + tn) > 0 else 0.0
            recall = float(sum_data.get("recall", 0) or 0)
            agr = float(sum_data.get("agreement_rate", 0) or 0)
            kappa = float(sum_data.get("kappa", 0) or 0)

            fpr_status = "✅ Pass (목표: 1.0% 미만)" if fpr < 0.01 else "❌ Fail (과탐 주의)"
            recall_status = "✅ Pass (목표: 95.0% 이상)" if recall >= 0.95 else "❌ Fail (방어력 부족)"
            agr_status = "✅ Pass (목표: 90.0% 이상)" if agr >= 0.90 else "❌ Fail (기준 불일치)"
            kappa_status = "✅ Pass (목표: 0.8 이상)" if kappa >= 0.8 else "⚠️ Warning (0.8 미만)"

            write_pair(5, 1, "과탐률 (FPR) [치명적 지표]", f"{fpr*100:.2f}%")
            write_pair(5, 4, "달성 여부", fpr_status)
            ws_dash.cell(row=5, column=5).font = Font(bold=True, color="008000" if "Pass" in fpr_status else "C00000")

            write_pair(6, 1, "스팸 방어율 (Recall)", f"{recall*100:.2f}%")
            write_pair(6, 4, "달성 여부", recall_status)
            ws_dash.cell(row=6, column=5).font = Font(bold=True, color="008000" if "Pass" in recall_status else "C00000")

            write_pair(7, 1, "인간-AI 일치율 (Agreement)", f"{agr*100:.2f}%")
            write_pair(7, 4, "달성 여부", agr_status)
            ws_dash.cell(row=7, column=5).font = Font(bold=True, color="008000" if "Pass" in agr_status else "C00000")

            write_pair(8, 1, "신뢰도 (Cohen's Kappa)", round(kappa, 4))
            write_pair(8, 4, "평가", kappa_status)
            ws_dash.cell(row=8, column=5).font = Font(bold=True, color="008000" if "Pass" in kappa_status else "D97500")

            # Layout: 2. 전체 데이터 처리 요약 (Overview)
            ws_dash.merge_cells('A10:F10')
            ws_dash['A10'] = "▶ 2. 전체 데이터 처리 요약 (Overview)"
            ws_dash['A10'].font = section_font
            ws_dash['A10'].fill = section_fill
            
            write_pair(11, 1, "분석 기준 파일", sum_data.get("sheet_used", "N/A"))
            write_pair(12, 1, "Human 총 레코드", sum_data.get("total_human", 0))
            write_pair(12, 4, "AI 처리 레코드", sum_data.get("total_llm", 0))
            write_pair(13, 1, "Human 기준 SPAM 수", sum_data.get("human_spam_count", 0))
            write_pair(13, 4, "AI 판별 SPAM 수", sum_data.get("llm_spam_count", 0))
            write_pair(14, 1, "매칭된 1:1 레코드 수", sum_data.get("matched", 0))
            
            # Layout: 3. 상세 분류 성능 지표 (Performance Metrics)
            ws_dash.merge_cells('A16:F16')
            ws_dash['A16'] = "▶ 3. 상세 분류 성능 지표 (Performance Metrics)"
            ws_dash['A16'].font = section_font
            ws_dash['A16'].fill = section_fill
            
            acc = sum_data.get('accuracy', 0)
            acc_val = f"{acc*100:.2f}%" if isinstance(acc, (int, float)) else acc

            write_pair(17, 1, "정확도 (Accuracy)", acc_val)
            write_pair(17, 4, "F1 Score", round(float(sum_data.get("f1", 0) or 0), 4))
            write_pair(18, 1, "정밀도 (Precision)", round(float(sum_data.get("precision", 0) or 0), 4))
            write_pair(18, 4, "재현율 (Recall)", round(float(sum_data.get("recall", 0) or 0), 4))
            
            # Layout: 4. 오차 행렬 (Confusion Matrix)
            ws_dash.merge_cells('A20:F20')
            ws_dash['A20'] = "▶ 4. 오차 행렬 (Confusion Matrix)"
            ws_dash['A20'].font = section_font
            ws_dash['A20'].fill = section_fill
            
            write_pair(21, 1, "TP (정답:SPAM, AI:SPAM)", sum_data.get("tp", 0))
            write_pair(21, 4, "TN (정답:HAM, AI:HAM)", sum_data.get("tn", 0))
            write_pair(22, 1, "FP (정답:HAM, AI:SPAM) [오탐]", sum_data.get("fp", 0))
            write_pair(22, 4, "FN (정답:SPAM, AI:HAM) [미탐]", sum_data.get("fn", 0))
            ws_dash.cell(row=22, column=2).font = Font(bold=True, color="C00000")
            ws_dash.cell(row=22, column=5).font = Font(bold=True, color="D97500")
            
            # Layout: 5. 잠재적 위험 (Type B) 모니터링
            ws_dash.merge_cells('A25:F25')
            ws_dash['A25'] = "▶ 5. 잠재적 위험 (Type B - Poisoning Risk) 모니터링"
            ws_dash['A25'].font = section_font
            ws_dash['A25'].fill = section_fill
            
            write_pair(26, 1, "Type B 전체 총계", sum_data.get("type_b_total_count", 0))
            ws_dash.cell(row=26, column=1).font = Font(bold=True, color="1F497D")
            
            write_pair(27, 1, " ↳ Type B (URL)", sum_data.get("type_b_url_count", 0))
            write_pair(27, 4, " ↳ Type B (Signature)", sum_data.get("type_b_sig_count", 0))
            write_pair(28, 1, " ↳ Type B (URL+Sig)", sum_data.get("type_b_both_count", 0))
            write_pair(28, 4, " ↳ Type B (기타)", sum_data.get("type_b_none_count", 0))
            
            # Dashboard Column Width Adjustments
            ws_dash.column_dimensions['A'].width = 32
            ws_dash.column_dimensions['B'].width = 15
            ws_dash.column_dimensions['C'].width = 5
            ws_dash.column_dimensions['D'].width = 30
            ws_dash.column_dimensions['E'].width = 15

            # ==============================
            # Formatting Diff Sheet
            # ==============================
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Colors for match status
            color_fp = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
            color_fn = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
            color_missing = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid") # Purple/Grey
            
            # Formatting Diff
            ws_diff = writer.sheets['Diff']
            ws_diff.auto_filter.ref = ws_diff.dimensions
            ws_diff.freeze_panes = "A2"
            
            ws_diff.column_dimensions['A'].width = 10
            ws_diff.column_dimensions['B'].width = 80
            ws_diff.column_dimensions['C'].width = 15
            ws_diff.column_dimensions['D'].width = 18
            ws_diff.column_dimensions['E'].width = 30
            ws_diff.column_dimensions['F'].width = 18
            ws_diff.column_dimensions['G'].width = 40
            ws_diff.column_dimensions['H'].width = 40
            ws_diff.column_dimensions['I'].width = 80
            ws_diff.column_dimensions['J'].width = 20

            for cell in ws_diff[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            for row in ws_diff.iter_rows(min_row=2, max_row=ws_diff.max_row, min_col=1, max_col=10):
                row[0].alignment = center_align   # 순번
                row[1].alignment = left_align     # 메시지 원본 (No Wrap)
                row[2].alignment = center_align   # 정답
                row[3].alignment = center_align   # Human 분류코드
                row[4].alignment = center_align   # AI 판단
                row[5].alignment = center_align   # AI 분류코드
                row[6].alignment = left_align     # AI 제출용 URL
                row[7].alignment = left_align     # AI 추출 SIGNATURE
                row[8].alignment = left_align     # AI 사유
                row[9].alignment = center_align   # 매칭 상태

                # Add color to match status IF it's an error (Apply to entire row)
                status_val = str(row[9].value).upper()
                target_fill = None
                
                if "FP" in status_val:
                    target_fill = color_fp
                elif "FN" in status_val:
                    target_fill = color_fn
                elif "MISSING" in status_val:
                    target_fill = color_missing
                    
                if target_fill:
                    for cell in row:
                        cell.fill = target_fill
        
        logger.info(f"Diff Excel exported successfully to {save_path}")
        return {"success": True, "path": save_path, "filename": request.filename}
    except PermissionError:
        logger.warning(f"Excel Export PermissionError: {save_path} is opened by another program.")
        raise HTTPException(status_code=400, detail="엑셀 파일이 열려있어서 덮어쓸 수 없습니다. 열려있는 엑셀 파일을 닫고 다시 내보내기를 시도해주세요.")
    except Exception as e:
        logger.exception("Excel Export Error")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/open")
async def open_excel_file(req: OpenFileRequest):
    try:
        if os.path.exists(req.path) and req.path.endswith('.xlsx'):
            os.startfile(req.path)
            return {"success": True}
        else:
            raise HTTPException(status_code=404, detail="File not found or invalid format")
    except Exception as e:
        logger.error(f"Failed to open file: {e}")
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)
