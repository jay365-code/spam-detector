import sys
import os
import openpyxl

def validate_excel(file_path):
    print(f"=== 엑셀 검수 리포트: {os.path.basename(file_path)} ===")
    if not os.path.exists(file_path):
        print("파일을 찾을 수 없습니다.")
        return
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    errors = []
    
    # 1. 육안분석 시트 검수
    sheet_name = "육안분석(시뮬결과35_150)"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        
        try:
            msg_col = headers.index("메시지") + 1
            gubun_col = headers.index("구분") + 1
            reason_col = headers.index("Reason") + 1
            code_col = headers.index("분류") + 1
        except ValueError as e:
            errors.append(f"[{sheet_name}] 필수 헤더를 찾지 못했습니다: {e}")
        else:
            for row in range(2, ws.max_row + 1):
                gubun = ws.cell(row=row, column=gubun_col).value
                reason = ws.cell(row=row, column=reason_col).value or ""
                cell_fill = ws.cell(row=row, column=msg_col).fill
                
                # Check Separated logic
                is_separated = "[텍스트 HAM + 악성 URL 분리 감지" in str(reason)
                
                color_hex = cell_fill.start_color.index if cell_fill else None
                if color_hex is not None and isinstance(color_hex, str):
                    color_hex = color_hex.upper()

                # Rule 1: 구분 'o' -> FFF2CC
                if str(gubun).lower() == "o":
                    if color_hex not in ["FFF2CC", "FFFFFF00", "00FFF2CC", "FFF2CCFF", 3]: # 3 is sometimes index for indexed colors, but FFF2CC is ARGB
                        # openpyxl returns ARGB (e.g. 00FFF2CC)
                        if not (isinstance(color_hex, str) and "FFF2CC" in color_hex):
                            errors.append(f"[{sheet_name}] 행 {row}: 구분이 'o'인데 배경색이 황금색(FFF2CC)이 아닙니다. (현재 색: {color_hex})")
                
                # Rule 2: Separated -> FFCCCC
                if is_separated:
                    if str(gubun).lower() == "o":
                        errors.append(f"[{sheet_name}] 행 {row}: URL 분리 감지(핑크색) 대상인데 구분에 'o'가 잘못 입력되었습니다.")
                    if not (isinstance(color_hex, str) and "FFCCCC" in color_hex):
                        errors.append(f"[{sheet_name}] 행 {row}: URL 분리 감지 대상인데 배경색이 핑크색(FFCCCC)이 아닙니다. (현재 색: {color_hex})")

    # 2. URL 중복제거 시트 검수 (40바이트 초과 및 IP 주소 확인)
    url_sheet = "URL중복 제거"
    import re
    ip_pattern = re.compile(r'\b\d{1,3}(?:\.\d{1,3}){3}\b')
    
    if url_sheet in wb.sheetnames:
        ws_url = wb[url_sheet]
        for row in range(2, ws_url.max_row + 1):
            url_val = str(ws_url.cell(row=row, column=1).value or "")
            len_val = ws_url.cell(row=row, column=2).value
            
            if url_val and url_val.strip() and url_val.strip() != "None":
                # IP 주소 검사
                if ip_pattern.search(url_val):
                    errors.append(f"[{url_sheet}] 일반 URL 행 {row}: IP 형식의 URL이 검출되었습니다. (삭제 요망: {url_val})")
                try:
                    if int(len_val) > 40:
                        errors.append(f"[{url_sheet}] 일반 URL 행 {row}: 40바이트를 초과하는 URL이 검출되었습니다. ({len_val} bytes)")
                except: pass
            
            # 단축 URL 컬럼 검사 (20번째 열)
            s_url_val = str(ws_url.cell(row=row, column=20).value or "")
            s_len_val = ws_url.cell(row=row, column=21).value
            
            if s_url_val and s_url_val.strip() and s_url_val.strip() != "None":
                if ip_pattern.search(s_url_val):
                    errors.append(f"[{url_sheet}] 단축 URL 행 {row}: IP 형식의 URL이 검출되었습니다. (삭제 요망: {s_url_val})")
                try:
                    if int(s_len_val) > 40:
                        errors.append(f"[{url_sheet}] 단축 URL 행 {row}: 40바이트를 초과하는 URL이 검출되었습니다. ({s_len_val} bytes)")
                except: pass

    # 3. 차단등록 시트 검수 (데드존 확인)
    block_sheet = "문자문장차단등록"
    if block_sheet in wb.sheetnames:
        ws_block = wb[block_sheet]
        for row in range(2, ws_block.max_row + 1):
            str_len = ws_block.cell(row=row, column=3).value
            sen_len = ws_block.cell(row=row, column=5).value
            
            # 문자열 (9~20)
            if str_len and str_len != 0:
                try:
                    val = int(str_len)
                    if val < 9:
                        errors.append(f"[{block_sheet}] 행 {row}: 문자열 최소길이(9바이트) 위반 ({val} bytes)")
                    elif val > 20:
                        errors.append(f"[{block_sheet}] 행 {row}: 문자열 최대길이(20바이트 초과) 위반 ({val} bytes)")
                except: pass
                
            # 문장열 (39~40)
            if sen_len and sen_len != 0:
                try:
                    val = int(sen_len)
                    if val < 39:
                        errors.append(f"[{block_sheet}] 행 {row}: 문장 데드존(21~38바이트) 및 길이 미달 위반 ({val} bytes)")
                    elif val > 40:
                        errors.append(f"[{block_sheet}] 행 {row}: 문장 최대길이(40바이트 초과) 위반 ({val} bytes)")
                except: pass

    # 4. 통계표 데이터 교차 검증
    target_summary_sheet = "TRAP.문장 중복제거"
    if target_summary_sheet in wb.sheetnames:
        ws_sum = wb[target_summary_sheet]
        spam_val = ws_sum.cell(row=3, column=6).value
        url_val = ws_sum.cell(row=3, column=7).value
        str_val = ws_sum.cell(row=3, column=8).value
        sen_val = ws_sum.cell(row=3, column=9).value
        
        # 실제 SPAM 카운트 계산
        actual_spam = 0
        if "육안분석(시뮬결과35_150)" in wb.sheetnames:
            ws_main = wb["육안분석(시뮬결과35_150)"]
            headers_main = [c.value for c in ws_main[1]]
            try:
                g_col = headers_main.index("구분") + 1
                try: red_col = headers_main.index("Red Group") + 1
                except ValueError: red_col = None
                try: rea_col = headers_main.index("Reason") + 1
                except ValueError: rea_col = None
                    
                for r in range(2, ws_main.max_row + 1):
                    is_spam = str(ws_main.cell(row=r, column=g_col).value).lower() == "o"
                    is_red = False
                    if red_col:
                        is_red = str(ws_main.cell(row=r, column=red_col).value).upper() == "O"
                    is_pink = False
                    if rea_col:
                        is_pink = "[텍스트 HAM + 악성 URL 분리 감지" in str(ws_main.cell(row=r, column=rea_col).value)
                    
                    if is_spam or is_red or is_pink:
                        actual_spam += 1
            except Exception as e:
                pass
            
        # 실제 URL 개수 산정
        actual_url = 0
        if "URL중복 제거" in wb.sheetnames:
            ws_u = wb["URL중복 제거"]
            for r in range(2, ws_u.max_row + 1):
                if ws_u.cell(row=r, column=1).value: actual_url += 1
                if ws_u.cell(row=r, column=20).value: actual_url += 1
                
        # 실제 문자열, 문장 카운트
        actual_str = max(0, wb["문자열중복제거"].max_row - 1) if "문자열중복제거" in wb.sheetnames else 0
        actual_sen = max(0, wb["문장중복제거"].max_row - 1) if "문장중복제거" in wb.sheetnames else 0
        
        # 검증
        if str(spam_val) != "-" and str(spam_val).isdigit() and int(spam_val) != actual_spam:
            errors.append(f"[{target_summary_sheet}] 표 수치 불일치: SPAM 수 (표: {spam_val} vs 실제: {actual_spam})")
        if str(url_val) != "-" and str(url_val).isdigit() and int(url_val) != actual_url:
            errors.append(f"[{target_summary_sheet}] 표 수치 불일치: URL 수 (표: {url_val} vs 실제: {actual_url})")
        if str(str_val) != "-" and str(str_val).isdigit() and int(str_val) != actual_str:
            errors.append(f"[{target_summary_sheet}] 표 수치 불일치: 문자열 수 (표: {str_val} vs 실제: {actual_str})")
        if str(sen_val) != "-" and str(sen_val).isdigit() and int(sen_val) != actual_sen:
            errors.append(f"[{target_summary_sheet}] 표 수치 불일치: 문장 수 (표: {sen_val} vs 실제: {actual_sen})")

    if not errors:
        print(">> 완벽합니다! 설정된 룰에 위배되는 1차 검수 에러가 없습니다.")
    else:
        print(f">> 총 {len(errors)}건의 위반 사항이 발견되었습니다:\n")
        for e in errors:
            print("-", e)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        target = sys.argv[1]
    else:
        target = "./spams/SD Output/MMSC스팸추출_20260413_C.xlsx"
    validate_excel(target)
