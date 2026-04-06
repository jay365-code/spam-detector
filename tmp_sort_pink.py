# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook(r'c:\Users\leejo\Project\AI Agent\Spam Detector\spams\MMSC스팸추출_20260327_A.xlsx', data_only=True)
ws = wb['육안분석(시뮬결과35_150)']
for row in ws.iter_rows(min_row=2):
    if not row[0].value: break
    color = str(row[0].fill.start_color.rgb) if getattr(row[0].fill, 'start_color', None) else 'None'
    if color == 'FFFFD1D1' or 'D1' in color:
        url_text = str(row[1].value) if row[1].value else 'None'
        msg_text = str(row[0].value).replace('\n', ' ')
        print(f"URL: {url_text[:30]} | 메시지: {msg_text[:20]}")
